"""
ow_hk.py  —  ONEWEST Housekeeping Module Blueprint
Prefix: ow_hk_  | Data: static/data/ow_hk_data.json
Routes:
  GET  /ow_hk                    → page
  GET  /api/ow_hk/data           → full data JSON (global live)
  POST /api/ow_hk/task           → add/update task
  POST /api/ow_hk/task/<id>/status → update task status
  POST /api/ow_hk/request        → add on-demand request
  POST /api/ow_hk/inspection     → add inspection record
  POST /api/ow_hk/staff          → add/update staff
  GET  /api/ow_hk/template       → download Excel template
  POST /api/ow_hk/upload         → upload Excel data
  DELETE /api/ow_hk/task/<id>    → delete task
"""

from flask import (Blueprint, render_template, jsonify, request,
                   session, send_file, make_response)
from pathlib import Path
from datetime import datetime, timedelta
import json, uuid, traceback, io

ow_hk_bp = Blueprint("ow_hk", __name__)

BASE_DIR    = Path(__file__).parent.resolve()
DATA_DIR    = BASE_DIR / "static" / "data"
OW_HK_JSON  = DATA_DIR / "ow_hk_data.json"

# ── Seed / default data structure ────────────────────────────────────────────
DEFAULT_DATA = {
    "tasks": [],
    "requests": [],
    "staff": [
        {"id":"s1","name":"Ravi Kumar","shift":"Morning","zone":"Lobby & Atrium","status":"active","current_task":None,"phone":""},
        {"id":"s2","name":"Priya Devi","shift":"Morning","zone":"Washrooms B1-GF","status":"active","current_task":None,"phone":""},
        {"id":"s3","name":"Suresh Naidu","shift":"Afternoon","zone":"Floors 1-3","status":"active","current_task":None,"phone":""},
        {"id":"s4","name":"Lakshmi Bai","shift":"Afternoon","zone":"Floors 4-6","status":"active","current_task":None,"phone":""},
        {"id":"s5","name":"Ramesh Rao","shift":"Night","zone":"External Areas","status":"off_shift","current_task":None,"phone":""},
    ],
    "inspections": [],
    "schedules": [
        {"id":"sc1","zone":"Lobby & Atrium","task":"Full sweep & mop","frequency":"Every 2 hrs","next_due":"08:00","sla_mins":30,"active":True},
        {"id":"sc2","zone":"Washrooms B1-GF","task":"Deep clean + restock","frequency":"Hourly","next_due":"08:00","sla_mins":20,"active":True},
        {"id":"sc3","zone":"Parking Level","task":"Sweep & debris removal","frequency":"Daily 06:00","next_due":"06:00","sla_mins":60,"active":True},
        {"id":"sc4","zone":"Lifts & Lobbies","task":"Wipe panels + mop","frequency":"Every 3 hrs","next_due":"09:00","sla_mins":15,"active":True},
        {"id":"sc5","zone":"Glass Facade","task":"External glass clean","frequency":"Weekly Mon","next_due":"Monday 07:00","sla_mins":120,"active":True},
    ],
    "materials": [
        {"id":"m1","name":"Floor Cleaner (Lizol)","unit":"Litres","stock":24,"min_stock":10,"category":"chemical"},
        {"id":"m2","name":"Glass Cleaner","unit":"Litres","stock":8,"min_stock":5,"category":"chemical"},
        {"id":"m3","name":"Toilet Cleaner","unit":"Litres","stock":18,"min_stock":8,"category":"chemical"},
        {"id":"m4","name":"Tissue Rolls","unit":"Rolls","stock":120,"min_stock":50,"category":"consumable"},
        {"id":"m5","name":"Trash Liners (L)","unit":"Nos","stock":200,"min_stock":80,"category":"consumable"},
        {"id":"m6","name":"Mop Head","unit":"Nos","stock":12,"min_stock":4,"category":"equipment"},
        {"id":"m7","name":"Scrub Pad","unit":"Nos","stock":30,"min_stock":10,"category":"equipment"},
        {"id":"m8","name":"Hand Gloves","unit":"Pairs","stock":40,"min_stock":15,"category":"ppe"},
    ],
    "material_log": [],
    "inout": [],
    "meta": {"last_updated": "", "property": "ONEWEST"}
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _load():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not OW_HK_JSON.exists():
        _save(DEFAULT_DATA)
        return DEFAULT_DATA
    try:
        with open(OW_HK_JSON, "r", encoding="utf-8") as f:
            d = json.load(f)
        # ensure all keys present
        for k, v in DEFAULT_DATA.items():
            if k not in d:
                d[k] = v
        return d
    except Exception:
        return DEFAULT_DATA

def _save(data):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    data["meta"]["last_updated"] = datetime.now().isoformat()
    with open(OW_HK_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _now_ist():
    return (datetime.utcnow() + timedelta(hours=5, minutes=30)).isoformat()

def _ist_str():
    t = datetime.utcnow() + timedelta(hours=5, minutes=30)
    return t.strftime("%d %b %Y %H:%M")

def _sla_status(created_at_iso, sla_mins):
    """Return 'ok'|'warning'|'breached' based on SLA."""
    try:
        created = datetime.fromisoformat(created_at_iso)
        elapsed = (datetime.utcnow() + timedelta(hours=5,minutes=30) - created).total_seconds() / 60
        if elapsed > sla_mins:         return "breached"
        elif elapsed > sla_mins * 0.7: return "warning"
        return "ok"
    except Exception:
        return "ok"

# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════════════════════
def _ow_hk_make_routes(login_required, require_property):

    # ── Page ──────────────────────────────────────────────────────────────────
    @ow_hk_bp.route("/ow_hk")
    @login_required
    @require_property("ONEWEST")
    def ow_hk_page():
        session["active_property"] = "ONEWEST"
        print(f"\n🧹 OW HK — User: {session.get('user')} | Role: {session.get('role')}")
        return render_template("ow_hk.html")

    # ── Full data (global live) ───────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/data")
    @login_required
    def ow_hk_data():
        d = _load()
        # Enrich tasks with live SLA status
        for t in d.get("tasks", []):
            if t.get("status") not in ("completed", "closed"):
                t["sla_status"] = _sla_status(t.get("created_at",""), t.get("sla_mins", 60))
        for r in d.get("requests", []):
            if r.get("status") not in ("completed", "closed"):
                r["sla_status"] = _sla_status(r.get("created_at",""), r.get("sla_mins", 30))
        if "materials"    not in d: d["materials"]    = DEFAULT_DATA["materials"]
        if "material_log" not in d: d["material_log"] = []
        if "inout"        not in d: d["inout"]         = []
        return jsonify(d)

    # ── Add / Update Task ─────────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/task", methods=["POST"])
    @login_required
    def ow_hk_task_add():
        body = request.get_json(force=True, silent=True) or {}
        d = _load()
        tid = body.get("id") or str(uuid.uuid4())[:8]
        existing = next((t for t in d["tasks"] if t["id"]==tid), None)
        sla_map  = {"critical":10,"high":20,"medium":60,"low":240}
        priority = body.get("priority","medium").lower()
        task = {
            "id":         tid,
            "title":      body.get("title","Unnamed Task"),
            "zone":       body.get("zone","General"),
            "area":       body.get("area",""),
            "task_type":  body.get("task_type","routine"),
            "priority":   priority,
            "status":     body.get("status","pending"),
            "assigned_to":body.get("assigned_to",""),
            "sla_mins":   body.get("sla_mins") or sla_map.get(priority, 60),
            "checklist":  body.get("checklist",[]),
            "notes":      body.get("notes",""),
            "created_at": body.get("created_at") or _now_ist(),
            "updated_at": _now_ist(),
            "created_by": session.get("user",""),
            "source":     body.get("source","manual"),
        }
        if existing:
            idx = d["tasks"].index(existing)
            d["tasks"][idx] = task
        else:
            d["tasks"].insert(0, task)
        _save(d)
        return jsonify({"success":True,"task":task})

    # ── Update task status ────────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/task/<tid>/status", methods=["POST"])
    @login_required
    def ow_hk_task_status(tid):
        body   = request.get_json(force=True, silent=True) or {}
        d      = _load()
        task   = next((t for t in d["tasks"] if t["id"]==tid), None)
        if not task:
            return jsonify({"success":False,"error":"Task not found"}), 404
        task["status"]     = body.get("status", task["status"])
        task["updated_at"] = _now_ist()
        if task["status"] == "completed":
            task["completed_at"] = _now_ist()
        _save(d)
        return jsonify({"success":True,"task":task})

    # ── Delete task ───────────────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/task/<tid>", methods=["DELETE"])
    @login_required
    def ow_hk_task_delete(tid):
        d = _load()
        d["tasks"] = [t for t in d["tasks"] if t["id"] != tid]
        _save(d)
        return jsonify({"success":True})

    # ── Add on-demand request ─────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/request", methods=["POST"])
    @login_required
    def ow_hk_request_add():
        body = request.get_json(force=True, silent=True) or {}
        d    = _load()
        sla_map = {"critical":5,"high":15,"medium":30,"low":60}
        priority = body.get("priority","high").lower()
        req = {
            "id":          str(uuid.uuid4())[:8],
            "title":       body.get("title","Service Request"),
            "location":    body.get("location",""),
            "zone":        body.get("zone",""),
            "priority":    priority,
            "status":      "open",
            "source":      body.get("source","tenant"),
            "notes":       body.get("notes",""),
            "sla_mins":    sla_map.get(priority, 30),
            "created_at":  _now_ist(),
            "updated_at":  _now_ist(),
            "created_by":  session.get("user",""),
        }
        d["requests"].insert(0, req)
        _save(d)
        return jsonify({"success":True,"request":req})

    # ── Convert request → task ────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/request/<rid>/convert", methods=["POST"])
    @login_required
    def ow_hk_request_convert(rid):
        body = request.get_json(force=True, silent=True) or {}
        d    = _load()
        req  = next((r for r in d["requests"] if r["id"]==rid), None)
        if not req:
            return jsonify({"success":False,"error":"Request not found"}), 404
        req["status"] = "converted"
        sla_map = {"critical":10,"high":20,"medium":60,"low":240}
        task = {
            "id":          str(uuid.uuid4())[:8],
            "title":       req["title"],
            "zone":        req.get("zone",""),
            "area":        req.get("location",""),
            "task_type":   "on_demand",
            "priority":    req.get("priority","high"),
            "status":      "pending",
            "assigned_to": body.get("assigned_to",""),
            "sla_mins":    sla_map.get(req.get("priority","high"),20),
            "checklist":   [],
            "notes":       req.get("notes",""),
            "created_at":  _now_ist(),
            "updated_at":  _now_ist(),
            "created_by":  session.get("user",""),
            "source":      "request",
            "request_ref": rid,
        }
        d["tasks"].insert(0, task)
        _save(d)
        return jsonify({"success":True,"task":task})

    # ── Add inspection ────────────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/inspection", methods=["POST"])
    @login_required
    def ow_hk_inspection_add():
        body = request.get_json(force=True, silent=True) or {}
        d    = _load()
        ins  = {
            "id":          str(uuid.uuid4())[:8],
            "zone":        body.get("zone",""),
            "inspector":   body.get("inspector", session.get("user","")),
            "score":       body.get("score", 0),
            "remarks":     body.get("remarks",""),
            "items":       body.get("items",[]),
            "result":      "pass" if body.get("score",0) >= 70 else "fail",
            "rework":      body.get("rework", False),
            "inspected_at":_now_ist(),
            "task_ref":    body.get("task_ref",""),
        }
        d["inspections"].insert(0, ins)
        _save(d)
        return jsonify({"success":True,"inspection":ins})

    # ── Add / update staff ────────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/staff", methods=["POST"])
    @login_required
    def ow_hk_staff_add():
        body = request.get_json(force=True, silent=True) or {}
        d    = _load()
        sid  = body.get("id") or str(uuid.uuid4())[:8]
        existing = next((s for s in d["staff"] if s["id"]==sid), None)
        member = {
            "id":           sid,
            "name":         body.get("name",""),
            "shift":        body.get("shift","Morning"),
            "zone":         body.get("zone",""),
            "status":       body.get("status","active"),
            "current_task": body.get("current_task",None),
            "phone":        body.get("phone",""),
        }
        if existing:
            d["staff"][d["staff"].index(existing)] = member
        else:
            d["staff"].append(member)
        _save(d)
        return jsonify({"success":True,"member":member})

    # ── Material: add/update item ────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/material", methods=["POST"])
    @login_required
    def ow_hk_material_add():
        body=request.get_json(force=True,silent=True) or {}
        d=_load()
        if "materials" not in d: d["materials"]=[]
        mid=body.get("id") or str(uuid.uuid4())[:8]
        existing=next((m for m in d["materials"] if m["id"]==mid),None)
        mat={"id":mid,"name":body.get("name",""),"unit":body.get("unit","Nos"),
             "stock":float(body.get("stock",0)),"min_stock":float(body.get("min_stock",0)),
             "category":body.get("category","consumable")}
        if existing: d["materials"][d["materials"].index(existing)]=mat
        else: d["materials"].append(mat)
        _save(d); return jsonify({"success":True,"material":mat})

    @ow_hk_bp.route("/api/ow_hk/material/<mid>/issue", methods=["POST"])
    @login_required
    def ow_hk_material_issue(mid):
        body=request.get_json(force=True,silent=True) or {}
        d=_load()
        if "material_log" not in d: d["material_log"]=[]
        mat=next((m for m in d.get("materials",[]) if m["id"]==mid),None)
        if not mat: return jsonify({"success":False,"error":"Not found"}),404
        qty=float(body.get("qty",0))
        if qty<=0: return jsonify({"success":False,"error":"Qty must be > 0"}),400
        mat["stock"]=max(0,mat["stock"]-qty)
        log={"id":str(uuid.uuid4())[:8],"material_id":mid,"material":mat["name"],"type":"issue",
             "qty":qty,"unit":mat["unit"],"zone":body.get("zone",""),"issued_to":body.get("issued_to",""),
             "notes":body.get("notes",""),"logged_at":_now_ist(),"logged_by":session.get("user","")}
        d["material_log"].insert(0,log); _save(d)
        return jsonify({"success":True,"material":mat,"log":log})

    @ow_hk_bp.route("/api/ow_hk/material/<mid>/receive", methods=["POST"])
    @login_required
    def ow_hk_material_receive(mid):
        body=request.get_json(force=True,silent=True) or {}
        d=_load()
        if "material_log" not in d: d["material_log"]=[]
        mat=next((m for m in d.get("materials",[]) if m["id"]==mid),None)
        if not mat: return jsonify({"success":False,"error":"Not found"}),404
        qty=float(body.get("qty",0))
        if qty<=0: return jsonify({"success":False,"error":"Qty must be > 0"}),400
        mat["stock"]+=qty
        log={"id":str(uuid.uuid4())[:8],"material_id":mid,"material":mat["name"],"type":"receive",
             "qty":qty,"unit":mat["unit"],"zone":"","issued_to":"",
             "notes":body.get("notes",""),"logged_at":_now_ist(),"logged_by":session.get("user","")}
        d["material_log"].insert(0,log); _save(d)
        return jsonify({"success":True,"material":mat,"log":log})

    # ── In & Out register ─────────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/inout", methods=["POST"])
    @login_required
    def ow_hk_inout_add():
        body=request.get_json(force=True,silent=True) or {}
        d=_load()
        if "inout" not in d: d["inout"]=[]
        entry={"id":str(uuid.uuid4())[:8],"name":body.get("name",""),
               "type":body.get("type","staff"),"direction":body.get("direction","in"),
               "zone":body.get("zone",""),"purpose":body.get("purpose",""),
               "items":body.get("items",""),"qty":body.get("qty",""),
               "authorized_by":body.get("authorized_by",""),"notes":body.get("notes",""),
               "timestamp":_now_ist(),"logged_by":session.get("user","")}
        d["inout"].insert(0,entry); _save(d)
        return jsonify({"success":True,"entry":entry})

    @ow_hk_bp.route("/api/ow_hk/inout/<eid>", methods=["DELETE"])
    @login_required
    def ow_hk_inout_delete(eid):
        d=_load(); d["inout"]=[e for e in d.get("inout",[]) if e["id"]!=eid]; _save(d)
        return jsonify({"success":True})

    # ── Excel template download ───────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/template")
    @login_required
    def ow_hk_template():
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()

            cyan   = "FF06B6D4"
            dark   = "FF0A1628"
            hdr_fg = "FFFFFFFF"
            alt    = "FFF0F9FF"

            def _sheet(wb, name, headers, examples, first=False):
                ws = wb.active if first else wb.create_sheet(name)
                ws.title = name
                thin = Side(style="thin", color="FFCCCCCC")
                border = Border(left=thin,right=thin,top=thin,bottom=thin)
                # Header row
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(1, ci, h)
                    c.font      = Font(bold=True, color=hdr_fg, size=10)
                    c.fill      = PatternFill("solid", fgColor=dark)
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border    = border
                ws.row_dimensions[1].height = 32
                # Sub-header (allowed values)
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(2, ci, "")
                    c.fill   = PatternFill("solid", fgColor="FF1E3A5F")
                    c.font   = Font(color="FF94A3B8", italic=True, size=8)
                    c.border = border
                    c.alignment = Alignment(horizontal="center")
                # Example rows
                for ri, row in enumerate(examples, 3):
                    for ci, val in enumerate(row, 1):
                        c = ws.cell(ri, ci, val)
                        c.alignment = Alignment(vertical="center", wrap_text=True)
                        c.border    = border
                        if ri % 2 == 1:
                            c.fill = PatternFill("solid", fgColor="FFF8FAFF")
                    ws.row_dimensions[ri].height = 22
                # Column widths
                for col in ws.columns:
                    maxw = max(len(str(c.value or "")) for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max(maxw + 4, 14), 42)
                return ws

            # ── Sheet 1: Tasks ──────────────────────────────────────────────
            _sheet(wb, "Tasks",
                headers=["Title","Zone","Area","Task Type","Priority","Assigned To","SLA (mins)","Checklist (;-sep)","Notes"],
                examples=[
                    ["Lobby full sweep","Lobby & Atrium","Main Entrance","routine","high","Ravi Kumar","30","Sweep floor;Mop;Empty bins;Wipe surfaces","Post peak hours"],
                    ["Washroom Level 1","Washrooms","1st Floor Gents","routine","critical","Priya Devi","20","Clean WC;Mop floor;Restock tissue;Sanitise","Check sensor"],
                    ["Spill cleanup Cafe","Cafeteria","Seating Area","on_demand","critical","Suresh Naidu","10","Cordon area;Clean spill;Dry floor;Inspect","Tenant call"],
                    ["Glass facade wash","External","North Face","periodic","low","Ramesh Rao","120","Safety harness;Squeegee top-down;Check seals","Weekly"],
                    ["Parking sweep","Parking","B1 Level","routine","medium","Lakshmi Bai","60","Leaf blower;Manual sweep;Collect debris","Morning shift"],
                ],
                first=True
            )

            # ── Sheet 2: On-Demand Requests ──────────────────────────────────
            _sheet(wb, "Requests",
                headers=["Title","Location","Zone","Priority","Source","Notes"],
                examples=[
                    ["Washroom attention","2nd Floor Gents","Washrooms","critical","tenant","Dirty after event"],
                    ["Spill at reception","Ground Floor Lobby","Lobby","high","fm","Water spill near lift"],
                    ["Pantry cleaning","4th Floor Pantry","Floors 4-6","medium","tenant","Post lunch cleaning required"],
                ]
            )

            # ── Sheet 3: Staff ───────────────────────────────────────────────
            _sheet(wb, "Staff",
                headers=["Name","Shift","Zone Assignment","Phone","Status"],
                examples=[
                    ["Ravi Kumar","Morning","Lobby & Atrium","9876543210","active"],
                    ["Priya Devi","Morning","Washrooms B1-GF","9876543211","active"],
                    ["Suresh Naidu","Afternoon","Floors 1-3","9876543212","active"],
                    ["Lakshmi Bai","Afternoon","Floors 4-6","9876543213","active"],
                    ["Ramesh Rao","Night","External Areas","9876543214","off_shift"],
                ]
            )

            # ── Sheet 4: Inspections ─────────────────────────────────────────
            _sheet(wb, "Inspections",
                headers=["Zone","Inspector","Score (0-100)","Remarks","Floor Clean?","Washroom OK?","Bins Empty?","Glass Clean?","Rework (yes/no)"],
                examples=[
                    ["Lobby & Atrium","Supervisor 1","92","Good condition","yes","yes","yes","yes","no"],
                    ["Washrooms B1","Supervisor 1","74","Tissue needs restock","yes","no","yes","yes","yes"],
                ]
            )


            # ── Sheet 5: Materials ──────────────────────────────────────────────
            _sheet(wb, "Materials",
                headers=["Item Name","Unit","Opening Stock","Min Stock Level","Category"],
                examples=[
                    ["Floor Cleaner (Lizol)","Litres","24","10","chemical"],
                    ["Glass Cleaner","Litres","8","5","chemical"],
                    ["Toilet Cleaner","Litres","18","8","chemical"],
                    ["Tissue Rolls","Rolls","120","50","consumable"],
                    ["Trash Liners (L)","Nos","200","80","consumable"],
                    ["Mop Head","Nos","12","4","equipment"],
                    ["Scrub Pad","Nos","30","10","equipment"],
                    ["Hand Gloves","Pairs","40","15","ppe"],
                ]
            )

            # ── Sheet 6: Material Issue / Receive Log ────────────────────────
            _sheet(wb, "Material_Log",
                headers=["Item Name","Type (issue/receive)","Qty","Zone","Issued To","Notes"],
                examples=[
                    ["Floor Cleaner (Lizol)","issue","2","Lobby & Atrium","Ravi Kumar","Morning shift"],
                    ["Tissue Rolls","issue","20","Washrooms B1-GF","Priya Devi","Daily restock"],
                    ["Trash Liners (L)","receive","100","Store","","New delivery"],
                    ["Mop Head","issue","2","Floors 1-3","Suresh Naidu","Replacement"],
                ]
            )

            # ── Sheet 7: In & Out Register ───────────────────────────────────
            _sheet(wb, "InOut_Register",
                headers=["Name / Item","Type","Direction (in/out)","Zone","Purpose","Items / Qty","Authorized By","Notes"],
                examples=[
                    ["Ravi Kumar","staff","in","Lobby & Atrium","Morning shift start","","FM",""],
                    ["Cleaning Vendor ABC","vendor","in","Parking Level","Periodic glass clean","Equipment x3","Supervisor",""],
                    ["Mop Trolley #2","equipment","out","Store","Issued for Floor 3","1 unit","FM","Return by EOD"],
                    ["Priya Devi","staff","out","Washrooms","Shift end","","",""],
                    ["Chemical Drums x2","material","in","Store","Monthly delivery","Floor Cleaner 20L","FM","Invoice #INV-234"],
                ]
            )


            # ── Sheet 5: Instructions ────────────────────────────────────────
            ws5 = wb.create_sheet("Instructions")
            ws5.title = "Instructions"
            instructions = [
                ("ONEWEST Housekeeping Data Upload Template",""),
                ("",""),
                ("Sheet","Instructions"),
                ("Tasks","Fill task details. Task Type: routine / on_demand / periodic. Priority: critical / high / medium / low"),
                ("Requests","On-demand service requests raised by tenants or FM. Source: tenant / fm / iot"),
                ("Staff","Housekeeping staff roster. Shift: Morning / Afternoon / Night. Status: active / busy / off_shift"),
                ("Inspections","Quality audit records. Score 0-100. Rework: yes/no"),
                ("Materials","Opening stock. Category: chemical / consumable / equipment / ppe"),
                ("Material_Log","Issue = consumed, receive = stocked. Qty numeric."),
                ("InOut_Register","Gate register. Type: staff/vendor/material/equipment. Direction: in/out"),
                ("",""),
                ("Notes","- Do not change column headers"),
                ("","- Date format: DD MMM YYYY or leave blank (system will use current time)"),
                ("","- Upload this file via the Upload button on the HK module"),
                ("","- All existing data will be preserved; new rows will be added"),
            ]
            for ri, (a, b) in enumerate(instructions, 1):
                ca = ws5.cell(ri, 1, a)
                cb = ws5.cell(ri, 2, b)
                if ri == 1:
                    ca.font = Font(bold=True, size=13, color="FF0A1628")
                elif ri == 3:
                    for c in [ca, cb]:
                        c.font = Font(bold=True, color=hdr_fg)
                        c.fill = PatternFill("solid", fgColor=dark)
                ws5.column_dimensions["A"].width = 22
                ws5.column_dimensions["B"].width = 70

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="OW_Housekeeping_Template.xlsx"
            )
        except Exception as e:
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ── Upload Excel data ─────────────────────────────────────────────────────
    @ow_hk_bp.route("/api/ow_hk/upload", methods=["POST"])
    @login_required
    def ow_hk_upload():
        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"success":False,"error":"No file"}), 400
        if not f.filename.lower().endswith((".xlsx",".xls")):
            return jsonify({"success":False,"error":"Only .xlsx/.xls accepted"}), 400
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            d  = _load()
            added = {"tasks":0,"requests":0,"staff":0,"inspections":0}
            sla_map = {"critical":10,"high":20,"medium":60,"low":240}

            # ── Tasks sheet ──────────────────────────────────────────────────
            if "Tasks" in wb.sheetnames:
                ws = wb["Tasks"]
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or not row[0]: continue
                    priority = str(row[4] or "medium").lower()
                    task = {
                        "id":str(uuid.uuid4())[:8],"title":str(row[0] or ""),
                        "zone":str(row[1] or ""),"area":str(row[2] or ""),
                        "task_type":str(row[3] or "routine"),
                        "priority":priority,"status":"pending",
                        "assigned_to":str(row[5] or ""),
                        "sla_mins":int(row[6] or 0) or sla_map.get(priority,60),
                        "checklist":[x.strip() for x in str(row[7] or "").split(";") if x.strip()],
                        "notes":str(row[8] or ""),
                        "created_at":_now_ist(),"updated_at":_now_ist(),
                        "created_by":session.get("user","upload"),"source":"upload",
                    }
                    d["tasks"].insert(0, task); added["tasks"]+=1

            # ── Requests sheet ───────────────────────────────────────────────
            if "Requests" in wb.sheetnames:
                ws = wb["Requests"]
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or not row[0]: continue
                    priority = str(row[3] or "medium").lower()
                    req = {
                        "id":str(uuid.uuid4())[:8],"title":str(row[0] or ""),
                        "location":str(row[1] or ""),"zone":str(row[2] or ""),
                        "priority":priority,"status":"open",
                        "source":str(row[4] or "manual"),
                        "notes":str(row[5] or ""),
                        "sla_mins":{"critical":5,"high":15,"medium":30,"low":60}.get(priority,30),
                        "created_at":_now_ist(),"updated_at":_now_ist(),
                        "created_by":session.get("user","upload"),
                    }
                    d["requests"].insert(0, req); added["requests"]+=1

            # ── Staff sheet ──────────────────────────────────────────────────
            if "Staff" in wb.sheetnames:
                ws = wb["Staff"]
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or not row[0]: continue
                    name = str(row[0] or "")
                    existing = next((s for s in d["staff"] if s["name"].lower()==name.lower()), None)
                    member = {
                        "id": existing["id"] if existing else str(uuid.uuid4())[:8],
                        "name":name,"shift":str(row[1] or "Morning"),
                        "zone":str(row[2] or ""),"phone":str(row[3] or ""),
                        "status":str(row[4] or "active"),"current_task":None,
                    }
                    if existing:
                        d["staff"][d["staff"].index(existing)] = member
                    else:
                        d["staff"].append(member); added["staff"]+=1

            # ── Inspections sheet ─────────────────────────────────────────────
            if "Inspections" in wb.sheetnames:
                ws = wb["Inspections"]
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or not row[0]: continue
                    score = int(row[2] or 0)
                    ins = {
                        "id":str(uuid.uuid4())[:8],"zone":str(row[0] or ""),
                        "inspector":str(row[1] or ""),"score":score,
                        "remarks":str(row[3] or ""),
                        "items":[str(row[i] or "") for i in range(4,9) if i<len(row)],
                        "result":"pass" if score>=70 else "fail",
                        "rework":str(row[8] or "no").lower() in ("yes","true","1") if len(row)>8 else False,
                        "inspected_at":_now_ist(),"task_ref":"",
                    }
                    d["inspections"].insert(0, ins); added["inspections"]+=1

            # ── Materials sheet ──────────────────────────────────────────────
            if "Materials" in wb.sheetnames:
                ws=wb["Materials"]
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or not row[0]: continue
                    name=str(row[0] or "")
                    existing=next((m for m in d.get("materials",[]) if m["name"].lower()==name.lower()),None)
                    mat={"id":existing["id"] if existing else str(uuid.uuid4())[:8],"name":name,
                         "unit":str(row[1] or "Nos"),"stock":float(row[2] or 0),
                         "min_stock":float(row[3] or 0),"category":str(row[4] or "consumable").lower()}
                    if existing: d["materials"][d["materials"].index(existing)]=mat
                    else: d["materials"].append(mat)
            if "Material_Log" in wb.sheetnames:
                ws=wb["Material_Log"]
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or not row[0]: continue
                    mat_name=str(row[0] or "")
                    mat_obj=next((m for m in d.get("materials",[]) if m["name"].lower()==mat_name.lower()),None)
                    qty=float(row[2] or 0)
                    if mat_obj and qty>0:
                        if str(row[1] or "").lower()=="issue": mat_obj["stock"]=max(0,mat_obj["stock"]-qty)
                        else: mat_obj["stock"]+=qty
                    d["material_log"].insert(0,{"id":str(uuid.uuid4())[:8],"material_id":mat_obj["id"] if mat_obj else "",
                        "material":mat_name,"type":str(row[1] or "issue").lower(),"qty":qty,
                        "unit":mat_obj["unit"] if mat_obj else "Nos","zone":str(row[3] or ""),
                        "issued_to":str(row[4] or ""),"notes":str(row[5] or ""),
                        "logged_at":_now_ist(),"logged_by":session.get("user","upload")})
            if "InOut_Register" in wb.sheetnames:
                ws=wb["InOut_Register"]
                for row in list(ws.iter_rows(values_only=True))[2:]:
                    if not row or not row[0]: continue
                    d["inout"].insert(0,{"id":str(uuid.uuid4())[:8],"name":str(row[0] or ""),
                        "type":str(row[1] or "staff").lower(),"direction":str(row[2] or "in").lower(),
                        "zone":str(row[3] or ""),"purpose":str(row[4] or ""),
                        "items":str(row[5] or ""),"qty":str(row[5] or ""),
                        "authorized_by":str(row[6] if len(row)>6 else ""),
                        "notes":str(row[7] if len(row)>7 else ""),
                        "timestamp":_now_ist(),"logged_by":session.get("user","upload")})
            _save(d)
            return jsonify({"success":True,"added":added})
        except Exception as e:
            return jsonify({"success":False,"error":str(e),"trace":traceback.format_exc()}), 500


def ow_hk_register(app, login_required=None, require_property=None):
    """Register blueprint — call AFTER login_required/require_property defined."""
    if login_required is None or require_property is None:
        from functools import wraps
        from flask import session as _s, redirect as _r, url_for as _u, request as _req, abort as _a
        def login_required(f):
            @wraps(f)
            def w(*a,**k):
                if "user" not in _s: return _r(_u("login")+"?next="+_req.path)
                return f(*a,**k)
            return w
        def require_property(prop):
            def dec(f):
                @wraps(f)
                def w(*a,**k):
                    if "user" not in _s: return _r(_u("login")+"?next="+_req.path)
                    bypass={"admin","management","general manager","property manager"}
                    if (_s.get("role") or "").lower() in bypass: return f(*a,**k)
                    if prop in _s.get("properties",[]): _s["active_property"]=prop; return f(*a,**k)
                    _a(403)
                return w
            return dec
        print("⚠️  ow_hk: using fallback auth guards")
    else:
        print("✅ ow_hk: using server's login_required + require_property")

    _ow_hk_make_routes(login_required, require_property)
    app.register_blueprint(ow_hk_bp)
    print("✅ Registered: ow_hk_bp  →  /ow_hk")