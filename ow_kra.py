"""
ow_kra.py  —  ONEWEST KRA Score Card Blueprint
Prefix: ow_  |  File: static/data/ow_kra.xlsx
Sheets: pm, afm, fire_m, T_Exe_1..3, MEP_Sup_1..4,
        HK_Exe_1..2, Sec_Exe, HK_Sup_1..2,
        Tech_1..5, HVAC_1..5, Plum_1..5,
        Ass_Tech_1..2, Carp, Paint  (34 sheets total)

Global: /api/ow_kra/data  → parsed JSON of all sheets
        /api/ow_kra/upload → replace xlsx globally
"""

from flask import (Blueprint, render_template, jsonify, request,
                   send_file, session, redirect, url_for, make_response)
from pathlib import Path
from datetime import datetime
import traceback

ow_kra_bp = Blueprint("ow_kra", __name__)

BASE_DIR = Path(__file__).parent.resolve()
OW_KRA_XLSX = BASE_DIR / "static" / "data" / "ow_kra.xlsx"
DATA_DIR    = BASE_DIR / "static" / "data"


# ── Parse Excel ───────────────────────────────────────────────────────────────
def _ow_kra_parse(sheet_name=None):
    """
    Parse ow_kra.xlsx.
    If sheet_name given → return that sheet's data dict.
    If None → return {sheet_name: meta_dict, ...} for all sheets.
    """
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl not installed"
    if not OW_KRA_XLSX.exists():
        return None, f"File not found: {OW_KRA_XLSX}"

    try:
        wb = openpyxl.load_workbook(str(OW_KRA_XLSX), data_only=True)

        def _parse_sheet(ws):
            rows = list(ws.iter_rows(values_only=True))
            info = {
                "name": "", "designation": "", "code": "",
                "site": "", "location": "", "period": "YEAR 2025-26"
            }
            kras = []

            for i, row in enumerate(rows):
                if not row or not any(c is not None for c in row):
                    continue
                cells = [str(c).strip() if c is not None else "" for c in row]
                joined = " ".join(cells).lower()

                # Header info rows (rows 0-7)
                if i < 8:
                    if "year" in cells[0].lower():
                        info["period"] = cells[0].strip()
                    for ci, cell in enumerate(cells):
                        cl = cell.lower()
                        nv = cells[ci+1] if ci+1 < len(cells) else ""
                        if "name of employee" in cl and nv:
                            info["name"] = nv
                        elif "designation" in cl and nv:
                            info["designation"] = nv
                        elif "employee code" in cl and nv:
                            info["code"] = nv
                        elif "site details" in cl and nv:
                            info["site"] = nv
                        elif "location" in cl and nv:
                            info["location"] = nv

                # KRA data rows — first col is numeric serial
                first = row[0]
                if first is not None and isinstance(first, (int, float)) and int(first) >= 1:
                    kra_text = str(row[1]).strip() if row[1] else ""
                    if not kra_text:
                        continue
                    # Weight — col 2 (may be 0.15 or 15 or "15%")
                    w_raw = row[2]
                    if w_raw is not None:
                        try:
                            w = float(str(w_raw).replace("%","").strip())
                            # stored as 0.15 → convert to 15
                            if 0 < w <= 1:
                                w = round(w * 100, 1)
                        except:
                            w = 0
                    else:
                        w = 0
                    benchmark = str(row[3]).strip() if row[3] else ""
                    self_rating  = row[4]   # may be None
                    lead_rating  = row[6]   # may be None
                    comments     = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                    kras.append({
                        "n": int(first),
                        "kra": kra_text,
                        "weight": w,
                        "benchmark": benchmark,
                        "selfRating": float(self_rating) if self_rating is not None else None,
                        "leadRating": float(lead_rating) if lead_rating is not None else None,
                        "comments": comments,
                    })

            return {"info": info, "kras": kras}

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return None, f"Sheet '{sheet_name}' not found"
            return _parse_sheet(wb[sheet_name]), None
        else:
            result = {
                "sheets": wb.sheetnames,
                "data": {}
            }
            for sn in wb.sheetnames:
                result["data"][sn] = _parse_sheet(wb[sn])
            return result, None

    except Exception as e:
        return None, f"{e}\n{traceback.format_exc()}"


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — injected at registration time with server's decorators
# ══════════════════════════════════════════════════════════════════════════════

def _ow_kra_make_routes(login_required, require_property):

    @ow_kra_bp.route("/ow_kra")
    @login_required
    @require_property("ONEWEST")
    def ow_kra_page():
        session["active_property"] = "ONEWEST"
        print(f"\n📊 OW KRA — User: {session.get('user')} | Role: {session.get('role')}")
        return render_template("ow_kra.html")

    @ow_kra_bp.route("/api/ow_kra/sheets")
    @login_required
    def ow_kra_sheets():
        """Return list of sheet names."""
        try:
            import openpyxl
            if not OW_KRA_XLSX.exists():
                return jsonify({"sheets": [], "error": "File not found"})
            wb = openpyxl.load_workbook(str(OW_KRA_XLSX), data_only=True)
            return jsonify({"sheets": wb.sheetnames})
        except Exception as e:
            return jsonify({"sheets": [], "error": str(e)})

    @ow_kra_bp.route("/api/ow_kra/data")
    @login_required
    def ow_kra_data():
        """
        Return parsed data for all sheets (or single sheet via ?sheet=name).
        Called by ow_kra.html on every load → global live data.
        """
        sheet = request.args.get("sheet")
        data, err = _ow_kra_parse(sheet_name=sheet)
        if err:
            return jsonify({"error": err}), 200
        return jsonify(data)

    @ow_kra_bp.route("/api/ow_kra/upload", methods=["POST"])
    @login_required
    def ow_kra_upload():
        """Replace ow_kra.xlsx — all clients see updated data on next load."""
        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"success": False, "error": "No file provided"}), 400
        if not f.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({"success": False, "error": "Only .xlsx/.xls accepted"}), 400
        try:
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            f.save(str(OW_KRA_XLSX))
            kb = OW_KRA_XLSX.stat().st_size // 1024
            return jsonify({
                "success": True,
                "message": f"ow_kra.xlsx updated ({kb} KB). All clients see new data on next load."
            })
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    @ow_kra_bp.route("/api/ow_kra/file")
    @login_required
    def ow_kra_file():
        """Download raw xlsx."""
        if not OW_KRA_XLSX.exists():
            return jsonify({"error": "File not found"}), 404
        resp = make_response(send_file(
            str(OW_KRA_XLSX),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="ow_kra.xlsx"
        ))
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp


def ow_kra_register(app, login_required=None, require_property=None):
    """
    Register blueprint. Call AFTER login_required/require_property are defined:
        from ow_kra import ow_kra_register
        ow_kra_register(app, login_required, require_property)
    """
    if login_required is None or require_property is None:
        from functools import wraps
        from flask import session as _s, redirect as _r, url_for as _u, request as _req, abort as _a

        def login_required(f):
            @wraps(f)
            def w(*a, **k):
                if "user" not in _s:
                    return _r(_u("login") + "?next=" + _req.path)
                return f(*a, **k)
            return w

        def require_property(prop):
            def dec(f):
                @wraps(f)
                def w(*a, **k):
                    if "user" not in _s:
                        return _r(_u("login") + "?next=" + _req.path)
                    bypass = {"admin","management","general manager","property manager"}
                    if (_s.get("role") or "").lower() in bypass:
                        return f(*a, **k)
                    if prop in _s.get("properties", []):
                        _s["active_property"] = prop
                        return f(*a, **k)
                    _a(403)
                return w
            return dec
        print("⚠️  ow_kra: using fallback auth guards")
    else:
        print("✅ ow_kra: using server's login_required + require_property")

    _ow_kra_make_routes(login_required, require_property)
    app.register_blueprint(ow_kra_bp)
    print("✅ Registered: ow_kra_bp  →  /ow_kra")