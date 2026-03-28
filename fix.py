import pandas as pd
import openpyxl

df = pd.read_excel('static/data/inventory_master.xlsx')
df['Item_Code'] = df['Item_Code'].astype(str).str.strip().str.lstrip("'")

wb = openpyxl.load_workbook('static/data/inventory_master.xlsx')
ws = wb.active

header = [cell.value for cell in ws[1]]
col_idx = header.index('Item_Code') + 1

for i, code in enumerate(df['Item_Code'].tolist(), start=2):
    ws.cell(row=i, column=col_idx).value = str(code)

wb.save('static/data/inventory_master.xlsx')
print('Fixed! First 10 codes:')
print(df['Item_Code'].tolist()[:10])
