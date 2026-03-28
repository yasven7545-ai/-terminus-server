"""
OW Asset Debug Script
Run this DIRECTLY on the server (same folder as server.py):
    python ow_debug_assets.py

It will tell you exactly what files exist, what sheets they have,
and how many assets are readable — without needing Flask.
"""
import sys, os
from pathlib import Path
import io

print("=" * 60)
print("OW ASSET DEBUG TOOL")
print("=" * 60)

# ── Locate the OW data directory ──
BASE_DIR = Path(__file__).parent
OW_DIR   = BASE_DIR / "static" / "data" / "OW"

print(f"\n📁 Looking in: {OW_DIR}")
print(f"   Exists: {OW_DIR.exists()}")

if OW_DIR.exists():
    files = list(OW_DIR.iterdir())
    print(f"   Files: {[f.name for f in files]}")
else:
    print("   ❌ Directory does not exist!")
    sys.exit(1)

# ── Check each possible Excel file ──
candidates = [
    OW_DIR / "Asset.xls",
    OW_DIR / "Asset.xlsx",
    OW_DIR / "asset.xls",
    OW_DIR / "asset.xlsx",
]

asset_file = None
for c in candidates:
    if c.exists():
        asset_file = c
        print(f"\n✅ Found: {c.name}  ({c.stat().st_size:,} bytes)")
        with open(c, 'rb') as f:
            magic = f.read(8)
        print(f"   Magic bytes: {magic.hex()}")
        is_zip  = magic[:2] == b'PK'
        is_biff = magic == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
        print(f"   Format: {'xlsx/zip' if is_zip else 'xls/BIFF8' if is_biff else 'UNKNOWN (possibly HTML)'}")
        break

if not asset_file:
    print("\n❌ NO ASSET FILE FOUND")
    print("   Upload Asset.xls via the dashboard sidebar first.")
    sys.exit(1)

# ── Try to read it ──
import subprocess

# Install xlrd if missing
try:
    import xlrd
    print(f"\n✅ xlrd version: {xlrd.__VERSION__}")
except ImportError:
    print("\n⚠️  xlrd not found — installing...")
    subprocess.run([sys.executable, "-m", "pip", "install", "xlrd>=2.0.1", "--quiet"])
    import xlrd

try:
    import pandas as pd
    print(f"✅ pandas version: {pd.__version__}")
except ImportError:
    print("❌ pandas not installed!")
    sys.exit(1)

try:
    import openpyxl
    print(f"✅ openpyxl version: {openpyxl.__version__}")
except ImportError:
    print("⚠️  openpyxl not installed")

# ── Discover sheet names ──
print(f"\n📋 Sheet names in {asset_file.name}:")
sheet_names = []

with open(asset_file, 'rb') as f:
    raw = f.read()
is_zip = raw[:2] == b'PK'

if is_zip:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        sheet_names = wb.sheetnames
        print(f"   openpyxl sheets: {sheet_names}")
    except Exception as e:
        print(f"   openpyxl failed: {e}")
else:
    try:
        wb = xlrd.open_workbook(str(asset_file))
        sheet_names = wb.sheet_names()
        print(f"   xlrd sheets: {sheet_names}")
    except Exception as e:
        print(f"   xlrd failed: {e}")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
            sheet_names = wb.sheetnames
            print(f"   openpyxl fallback sheets: {sheet_names}")
        except Exception as e2:
            print(f"   openpyxl fallback failed: {e2}")

# ── Try reading each sheet ──
print(f"\n📊 Reading sheets:")

def try_read_sheet(path, sheet):
    raw_data = open(path,'rb').read()
    is_zip_ = raw_data[:2] == b'PK'
    errors = []

    # openpyxl via BytesIO
    if is_zip_:
        try:
            df = pd.read_excel(io.BytesIO(raw_data), sheet_name=sheet, engine='openpyxl')
            return df, 'openpyxl'
        except Exception as e:
            errors.append(f"openpyxl: {e}")

    # xlrd
    try:
        df = pd.read_excel(path, sheet_name=sheet, engine='xlrd')
        return df, 'xlrd'
    except Exception as e:
        errors.append(f"xlrd: {e}")

    # openpyxl BytesIO (for renamed xls)
    if not is_zip_:
        try:
            df = pd.read_excel(io.BytesIO(raw_data), sheet_name=sheet, engine='openpyxl')
            return df, 'openpyxl-bytes'
        except Exception as e:
            errors.append(f"openpyxl-bytes: {e}")

    # calamine
    try:
        df = pd.read_excel(path, sheet_name=sheet, engine='calamine')
        return df, 'calamine'
    except Exception as e:
        errors.append(f"calamine: {e}")

    return None, ' | '.join(errors)


# Try named sheets
for sname in ['inhouse_ppm', 'vendor_ppm'] + (sheet_names or []):
    df, eng = try_read_sheet(asset_file, sname)
    if df is not None:
        cols = list(df.columns)
        rows = len(df)
        asset_rows = len([1 for _, r in df.iterrows() 
                          if str(r.get('Asset Code','')).strip() not in ('','nan','None')])
        print(f"   ✅ sheet='{sname}' via {eng}: {rows} rows, {asset_rows} assets")
        print(f"      Columns: {cols[:8]}")
        if asset_rows > 0:
            sample = df[df['Asset Code'].notna()].iloc[0]
            print(f"      Sample row: Code={sample.get('Asset Code','')} | Name={sample.get('Asset Name','')} | Location={sample.get('Location','')}")
    else:
        print(f"   ❌ sheet='{sname}': {eng}")

# Try by index
for idx in [0, 1]:
    df, eng = try_read_sheet(asset_file, idx)
    if df is not None:
        asset_col = 'Asset Code' if 'Asset Code' in df.columns else df.columns[0] if len(df.columns) > 0 else None
        rows = len(df)
        print(f"   ✅ sheet index={idx} via {eng}: {rows} rows, columns={list(df.columns[:6])}")

print("\n" + "=" * 60)
print("DONE — Share this output to diagnose the issue.")
print("=" * 60)