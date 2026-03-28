"""
ENERGY ROUTES
SLN Terminus energy analytics: serve Excel, upload replacement, parse JSON.
"""
from flask import Blueprint, jsonify, request, send_file, make_response
from datetime import datetime as _dt, timedelta as _td
import traceback

from decorators import login_required
from config import BASE_DIR

energy_bp = Blueprint("energy", __name__)


# ── Serve dashboard Excel with no-cache headers ───────────────────────────────
@energy_bp.route("/api/energy/data")
@login_required
def energy_data_redirect():
    """API alias for the dashboard Excel — returns with no-cache headers."""
    path = BASE_DIR / "static" / "data" / "SLN_Terminus_Dashboard_Data.xlsx"
    if not path.exists():
        return jsonify({"error": "SLN_Terminus_Dashboard_Data.xlsx not found in static/data/"}), 404
    resp = make_response(send_file(
        str(path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"]        = "no-cache"
    return resp


# ── Replace dashboard Excel via upload ────────────────────────────────────────
@energy_bp.route("/api/energy/upload", methods=["POST"])
@login_required
def energy_upload():
    """Allow replacing SLN_Terminus_Dashboard_Data.xlsx via file upload."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"success": False, "error": "No file provided"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "error": "Only .xlsx / .xls files accepted"}), 400
    try:
        dest = BASE_DIR / "static" / "data" / "Energy_analysys.xlsx"
        dest.parent.mkdir(parents=True, exist_ok=True)
        f.save(str(dest))
        return jsonify({
            "success": True,
            "message": f"Dashboard data updated ({dest.stat().st_size // 1024} KB)",
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── Parse Excel → JSON (all sheets incl. water_consumption) ──────────────────
@energy_bp.route("/api/energy/json")
@login_required
def energy_json():
    """
    Parse Energy_analysys.xlsx server-side and return structured JSON.
    Handles sheet names with spaces (e.g. 'water _consumption').
    Called by energy.html on DOMContentLoaded so all modules get live data.
    """
    import openpyxl

    XLSX_NAME = "Energy_analysys.xlsx"
    path = BASE_DIR / "static" / "data" / XLSX_NAME
    if not path.exists():
        path = BASE_DIR / "static" / "data" / "SLN_Terminus_Dashboard_Data.xlsx"
    if not path.exists():
        return jsonify({"error": f"{XLSX_NAME} not found in static/data/"}), 404

    try:
        wb     = openpyxl.load_workbook(str(path), data_only=True)
        result = {}

        # ── ele_consumption ──────────────────────────────────────────────────
        ele_sheet = next((s for s in wb.sheetnames
                          if "ele" in s.lower() and "sft" not in s.lower()), None)
        if ele_sheet:
            ele = []
            for row in list(wb[ele_sheet].iter_rows(values_only=True))[1:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                ele.append({"n": int(row[0]), "m": m,
                             "u": float(row[2]) if row[2] is not None else None})
            result["ele"] = ele

        # ── dg_consumption ───────────────────────────────────────────────────
        dg_sheet = next((s for s in wb.sheetnames if "dg" in s.lower()), None)
        if dg_sheet:
            dg   = []
            rows = list(wb[dg_sheet].iter_rows(values_only=True))
            start = next((i + 1 for i, r in enumerate(rows) if r[0] == "Sno"), 2)
            for row in rows[start:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m  = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                rh = row[2]
                if isinstance(rh, _td):
                    ts  = int(rh.total_seconds())
                    hd  = round(ts / 3600, 2)
                    h_str = f"{ts // 3600:02d}:{(ts % 3600) // 60:02d}"
                elif hasattr(rh, "hour"):
                    hd    = round(rh.hour + rh.minute / 60, 2)
                    h_str = f"{rh.hour:02d}:{rh.minute:02d}"
                else:
                    hd    = float(rh or 0)
                    hh    = int(hd)
                    h_str = f"{hh:02d}:{int((hd - hh) * 60):02d}"
                dg.append({"n": int(row[0]), "m": m, "h": h_str, "hd": hd,
                            "k": float(row[3] or 0), "d": float(row[4] or 0)})
            result["dg"] = dg

        # ── water _consumption (sheet name may have a space) ─────────────────
        water_sheet = next((s for s in wb.sheetnames
                            if "water" in s.lower() and "tank" not in s.lower()), None)
        if water_sheet:
            water = []
            for row in list(wb[water_sheet].iter_rows(values_only=True))[1:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m   = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                kl  = float(row[2]) if row[2] is not None else None
                pay = float(row[3]) if len(row) > 3 and row[3] is not None else None
                water.append({"n": int(row[0]), "m": m, "v": kl, "pay": pay})
            result["water"] = water

        # ── ele_con_sft (benchmark / area data) ──────────────────────────────
        sft_sheet = next((s for s in wb.sheetnames if "sft" in s.lower()), None)
        if sft_sheet:
            bm = {}
            for row in wb[sft_sheet].iter_rows(values_only=True):
                if row[0] and row[1] is not None:
                    bm[str(row[0]).strip()] = row[1]
            result["sft"] = bm

        # ── Tank_cons ─────────────────────────────────────────────────────────
        tank_sheet = next((s for s in wb.sheetnames if "tank" in s.lower()), None)
        if tank_sheet:
            tanker  = []
            rows_t  = list(wb[tank_sheet].iter_rows(values_only=True))
            data_start = next((i + 1 for i, r in enumerate(rows_t) if r[0] == "Sno"), 1)
            for row in rows_t[data_start:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m      = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                trips  = int(row[2])   if len(row) > 2 and row[2] is not None else None
                vol    = float(row[3]) if len(row) > 3 and row[3] is not None else None
                rate   = float(row[5]) if len(row) > 5 and row[5] is not None else None
                cost   = float(row[6]) if len(row) > 6 and row[6] is not None else None
                vendor = str(row[7]).strip() if len(row) > 7 and row[7] is not None else "—"
                tanker.append({"n": int(row[0]), "m": m, "trips": trips, "vol": vol,
                                "rate": rate, "cost": cost, "vendor": vendor})
            result["tanker"] = tanker

        result["source"] = XLSX_NAME
        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
