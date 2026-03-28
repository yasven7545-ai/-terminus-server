"""
SLN BUDGET REVIEW — Flask Blueprint  v3.0
==========================================
• Reads Budget_review.xlsx from  static/data/Budget_review.xlsx  on first startup
• Persists edits globally in     static/data/sln_budget_data.json
• Every change is logged in      static/data/sln_budget_audit.json
• Frontend-editable: any cell in any sheet can be changed by authorised users
  (finance team / GM / management / admin / property manager)
• Changes are immediately visible to ALL users (no localStorage, no per-user state)

SHEET NAMES IN EXCEL (exact):
  INDEX, area_summary, area_mall, area_office, area_comm, common_services,
  cam_variables, cam_charges_summary, cam_charges, amc_cost,
  comm_electrical_load, load_breakup, hvac_mall, mall_ahu_load,
  manpower, salary_breakup, hk_secr_equipments, sinking_fund_estimation

REGISTER IN server.py:
    safe_register("sln_budget", "sln_budget_bp")
"""

from flask import Blueprint, render_template, request, jsonify, session, Response
from functools import wraps
from pathlib import Path
import json, threading, copy
from datetime import datetime

try:
    import openpyxl
    _XL_OK = True
except ImportError:
    _XL_OK = False
    print("⚠️  [sln_budget] openpyxl not installed — Excel load disabled. Run: pip install openpyxl")

# ── Blueprint ──────────────────────────────────────────────────────────────
sln_budget_bp = Blueprint("sln_budget", __name__)

# ── Paths ──────────────────────────────────────────────────────────────────
_BASE      = Path(__file__).parent.resolve()
_DATA_DIR  = _BASE / "static" / "data"
_XLSX      = _DATA_DIR / "Budget_review.xlsx"
_JSON_FILE = _DATA_DIR / "sln_budget_data.json"
_AUDIT     = _DATA_DIR / "sln_budget_audit.json"
_LOCK      = threading.Lock()

_DATA_DIR.mkdir(parents=True, exist_ok=True)

# ── Roles allowed to EDIT (finance team, GM, management, admin) ────────────
_EDIT_ROLES = {
    "admin", "management", "general manager",
    "property manager", "finance manager", "executive",
}

# ── Sheet names exactly as in the Excel workbook ──────────────────────────
SHEET_NAMES = [
    "INDEX", "area_summary", "area_mall", "area_office", "area_comm",
    "common_services", "cam_variables", "cam_charges_summary", "cam_charges",
    "amc_cost", "comm_electrical_load", "load_breakup", "hvac_mall",
    "mall_ahu_load", "manpower", "salary_breakup",
    "hk_secr_equipments", "sinking_fund_estimation",
]

# ══════════════════════════════════════════════════════════════════════════
#  EXCEL → JSON READER
# ══════════════════════════════════════════════════════════════════════════
def _cell_val(v):
    """Convert openpyxl cell value to a JSON-serialisable Python type."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        # round away floating-point noise
        r = round(v, 6)
        return int(r) if r == int(r) else r
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        return v.strip()
    # date / datetime / etc.
    try:
        return str(v)
    except Exception:
        return ""


def load_excel(xlsx_path: Path) -> dict:
    """
    Read every sheet in Budget_review.xlsx.
    Returns dict keyed by sheet name, each value:
        {
          "headers":    [col0_header, col1_header, ...],
          "rows":       [
              {"row_num": 2, "cells": [v0, v1, ...], "modified": False, "note": ""},
              ...
          ],
          "source":     "excel",
          "sheet_note": ""
        }
    row_num is the 1-based Excel row number so the UI can show the real address.
    """
    if not _XL_OK:
        print("❌ [sln_budget] openpyxl unavailable — cannot load Excel")
        return {}
    if not xlsx_path.exists():
        print(f"❌ [sln_budget] Excel file not found: {xlsx_path}")
        return {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"❌ [sln_budget] Cannot open workbook: {e}")
        return {}

    result = {}
    for sname in SHEET_NAMES:
        ws = None
        # exact match first
        if sname in wb.sheetnames:
            ws = wb[sname]
        else:
            # case-insensitive / underscore-tolerant fallback
            for real in wb.sheetnames:
                if real.lower().replace(" ", "_") == sname.lower():
                    ws = wb[real]
                    break
        if ws is None:
            print(f"⚠️  [sln_budget] Sheet '{sname}' not found in workbook")
            result[sname] = {"headers": [], "rows": [], "source": "missing", "sheet_note": ""}
            continue

        raw = []
        for row in ws.iter_rows(values_only=True):
            raw.append([_cell_val(c) for c in row])

        # find first non-blank row as header row
        header_idx = 0
        for i, row in enumerate(raw):
            if any(v not in ("", None) for v in row):
                header_idx = i
                break

        headers   = raw[header_idx] if raw else []
        data_rows = []
        for offset, row in enumerate(raw[header_idx + 1:], start=0):
            # Excel row number = header_idx(1-based) + 1(header itself) + 1(offset starts at 0) + offset
            excel_row = header_idx + 2 + offset   # 1-based
            # skip fully blank rows
            if all(v in ("", None) for v in row):
                continue
            data_rows.append({
                "row_num":  excel_row,
                "cells":    row,
                "modified": False,
                "hidden":   False,
                "note":     ""
            })

        result[sname] = {
            "headers":    headers,
            "rows":       data_rows,
            "source":     "excel",
            "sheet_note": ""
        }
        print(f"  ✅ [sln_budget] '{sname}': {len(data_rows)} rows × {len(headers)} cols  (Excel row {header_idx+1} = header)")

    wb.close()
    return result


# ══════════════════════════════════════════════════════════════════════════
#  PERSISTENCE
# ══════════════════════════════════════════════════════════════════════════
def _seed_from_excel() -> dict:
    print("📊 [sln_budget] Seeding JSON from Excel…")
    sheets = load_excel(_XLSX)
    data = {
        "meta": {
            "property":     "SLN Terminus",
            "title":        "CAM Budget Review",
            "fiscal_year":  "FY 2016",
            "excel_file":   _XLSX.name,
            "seeded_at":    datetime.now().isoformat(),
            "last_updated": "",
            "updated_by":   "",
            "notes":        ""
        },
        "sheets": sheets
    }
    _save_raw(data)
    return data


def _save_raw(data: dict) -> bool:
    """Atomic write — no lock, caller responsible."""
    try:
        tmp = _JSON_FILE.with_suffix(".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        tmp.replace(_JSON_FILE)
        return True
    except Exception as e:
        print(f"❌ [sln_budget] Save failed: {e}")
        return False


def _save(data: dict) -> bool:
    with _LOCK:
        return _save_raw(data)


def _load() -> dict:
    """Load from JSON; auto-seed from Excel on first run."""
    with _LOCK:
        if _JSON_FILE.exists():
            try:
                with open(_JSON_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                print(f"⚠️  [sln_budget] JSON read error: {e}")
        # First run — seed from Excel
        return _seed_from_excel()


# ── Audit log ──────────────────────────────────────────────────────────────
def _audit(action: str, sheet: str, detail: dict):
    try:
        log = []
        if _AUDIT.exists():
            try:
                with open(_AUDIT, "r", encoding="utf-8") as f:
                    log = json.load(f)
            except Exception:
                pass
        log.append({
            "ts":     datetime.now().isoformat(),
            "user":   session.get("user", "—"),
            "role":   session.get("role", "—"),
            "action": action,
            "sheet":  sheet,
            **detail
        })
        log = log[-500:]
        with open(_AUDIT, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ── Auth helpers ───────────────────────────────────────────────────────────
def _login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return jsonify({"success": False, "error": "Not authenticated"}), 401
        return fn(*args, **kwargs)
    return wrapper


def _can_edit() -> bool:
    return (session.get("role") or "").lower() in _EDIT_ROLES


# ══════════════════════════════════════════════════════════════════════════
#  PAGE ROUTE
# ══════════════════════════════════════════════════════════════════════════
@sln_budget_bp.route("/sln_budget")
@_login_required
def sln_budget_page():
    prop = session.get("active_property", "")
    role = (session.get("role") or "").lower()
    full = role in {"admin", "management", "general manager", "property manager"}
    if prop not in ("SLN Terminus",) and not full:
        from flask import abort
        abort(403)
    return render_template("sln_budget.html")


# ══════════════════════════════════════════════════════════════════════════
#  READ APIs
# ══════════════════════════════════════════════════════════════════════════
@sln_budget_bp.route("/api/sln_budget/data")
@_login_required
def api_data():
    """Full dataset — all sheets + meta."""
    return jsonify({"success": True, "data": _load(), "can_edit": _can_edit()})


@sln_budget_bp.route("/api/sln_budget/sheet/<sheet_name>")
@_login_required
def api_sheet(sheet_name):
    data  = _load()
    sheet = data.get("sheets", {}).get(sheet_name)
    if sheet is None:
        return jsonify({"success": False, "error": f"Sheet '{sheet_name}' not found"}), 404
    return jsonify({"success": True, "sheet_name": sheet_name,
                    "data": sheet, "can_edit": _can_edit()})


@sln_budget_bp.route("/api/sln_budget/sheet_names")
@_login_required
def api_sheet_names():
    sheets = _load().get("sheets", {})
    return jsonify({
        "success": True,
        "sheets": [
            {"name": k,
             "rows": len(v.get("rows", [])),
             "cols": len(v.get("headers", [])),
             "note": v.get("sheet_note", ""),
             "source": v.get("source", "")}
            for k, v in sheets.items()
        ]
    })


@sln_budget_bp.route("/api/sln_budget/meta_info")
@_login_required
def api_meta_get():
    return jsonify({"success": True, "meta": _load().get("meta", {}),
                    "can_edit": _can_edit()})


# ══════════════════════════════════════════════════════════════════════════
#  WRITE APIs  — authorised roles only
# ══════════════════════════════════════════════════════════════════════════

@sln_budget_bp.route("/api/sln_budget/edit_cell", methods=["POST"])
@_login_required
def api_edit_cell():
    """
    Edit a single cell.
    Body: { "sheet": "cam_charges_summary", "row_num": 4, "col_idx": 2,
            "value": 13.5, "note": "GM decision 2026-03-18" }
    Global — visible to all users immediately.
    """
    if not _can_edit():
        return jsonify({"success": False, "error": "No permission to edit"}), 403

    body    = request.get_json(force=True) or {}
    sheet   = body.get("sheet")
    row_num = body.get("row_num")
    col_idx = body.get("col_idx")
    value   = body.get("value")
    note    = body.get("note", "")

    if not sheet or row_num is None or col_idx is None:
        return jsonify({"success": False, "error": "sheet, row_num, col_idx required"}), 400

    data  = _load()
    sdata = data.get("sheets", {}).get(sheet)
    if sdata is None:
        return jsonify({"success": False, "error": f"Sheet '{sheet}' not found"}), 404

    rows   = sdata.get("rows", [])
    target = next((r for r in rows if r["row_num"] == row_num), None)
    if target is None:
        return jsonify({"success": False, "error": f"Row {row_num} not found in '{sheet}'"}), 404
    if col_idx >= len(target["cells"]):
        return jsonify({"success": False, "error": f"col_idx {col_idx} out of range"}), 400

    old_val = target["cells"][col_idx]
    target["cells"][col_idx] = value
    target["modified"]       = True
    if note:
        target["note"] = note

    data["meta"]["last_updated"] = datetime.now().isoformat()
    data["meta"]["updated_by"]   = session.get("user", "—")
    _save(data)

    header = sdata["headers"][col_idx] if col_idx < len(sdata["headers"]) else f"col_{col_idx}"
    _audit("edit_cell", sheet, {
        "row_num": row_num, "col_idx": col_idx,
        "header": str(header), "old": old_val, "new": value, "note": note
    })

    return jsonify({
        "success":      True,
        "sheet":        sheet,
        "row_num":      row_num,
        "col_idx":      col_idx,
        "header":       header,
        "old_value":    old_val,
        "new_value":    value,
        "last_updated": data["meta"]["last_updated"],
        "updated_by":   data["meta"]["updated_by"]
    })


@sln_budget_bp.route("/api/sln_budget/edit_row", methods=["POST"])
@_login_required
def api_edit_row():
    """
    Replace all cells in a row.
    Body: { "sheet": "...", "row_num": 4, "cells": [...], "note": "..." }
    """
    if not _can_edit():
        return jsonify({"success": False, "error": "No permission to edit"}), 403

    body    = request.get_json(force=True) or {}
    sheet   = body.get("sheet")
    row_num = body.get("row_num")
    cells   = body.get("cells")
    note    = body.get("note", "")

    if not sheet or row_num is None or not isinstance(cells, list):
        return jsonify({"success": False, "error": "sheet, row_num, cells[] required"}), 400

    data   = _load()
    sdata  = data.get("sheets", {}).get(sheet)
    if sdata is None:
        return jsonify({"success": False, "error": f"Sheet '{sheet}' not found"}), 404

    target = next((r for r in sdata["rows"] if r["row_num"] == row_num), None)
    if target is None:
        return jsonify({"success": False, "error": f"Row {row_num} not found"}), 404

    old_cells = list(target["cells"])
    for i, v in enumerate(cells):
        if i < len(target["cells"]):
            target["cells"][i] = v
    target["modified"] = True
    if note:
        target["note"] = note

    data["meta"]["last_updated"] = datetime.now().isoformat()
    data["meta"]["updated_by"]   = session.get("user", "—")
    _save(data)
    _audit("edit_row", sheet, {"row_num": row_num, "old": old_cells, "new": cells, "note": note})
    return jsonify({"success": True, "row_num": row_num,
                    "last_updated": data["meta"]["last_updated"]})


@sln_budget_bp.route("/api/sln_budget/add_row", methods=["POST"])
@_login_required
def api_add_row():
    """
    Append a new blank/filled row to a sheet.
    Body: { "sheet": "...", "cells": [...], "note": "..." }
    """
    if not _can_edit():
        return jsonify({"success": False, "error": "No permission to edit"}), 403

    body  = request.get_json(force=True) or {}
    sheet = body.get("sheet")
    cells = body.get("cells", [])
    note  = body.get("note", "Added by user")

    if not sheet:
        return jsonify({"success": False, "error": "sheet required"}), 400

    data  = _load()
    sdata = data.get("sheets", {}).get(sheet)
    if sdata is None:
        return jsonify({"success": False, "error": f"Sheet '{sheet}' not found"}), 404

    rows    = sdata["rows"]
    max_row = max((r["row_num"] for r in rows), default=1)
    new_row = {"row_num": max_row + 1, "cells": cells,
               "modified": True, "hidden": False, "note": note}
    rows.append(new_row)
    data["meta"]["last_updated"] = datetime.now().isoformat()
    data["meta"]["updated_by"]   = session.get("user", "—")
    _save(data)
    _audit("add_row", sheet, {"row_num": new_row["row_num"], "cells": cells})
    return jsonify({"success": True, "row_num": new_row["row_num"]})


@sln_budget_bp.route("/api/sln_budget/delete_row", methods=["POST"])
@_login_required
def api_delete_row():
    """
    Soft-delete a row (hidden=True).
    Body: { "sheet": "...", "row_num": 4 }
    """
    if not _can_edit():
        return jsonify({"success": False, "error": "No permission"}), 403

    body    = request.get_json(force=True) or {}
    sheet   = body.get("sheet")
    row_num = body.get("row_num")
    if not sheet or row_num is None:
        return jsonify({"success": False, "error": "sheet + row_num required"}), 400

    data   = _load()
    sdata  = data.get("sheets", {}).get(sheet)
    if sdata is None:
        return jsonify({"success": False, "error": "Sheet not found"}), 404

    target = next((r for r in sdata["rows"] if r["row_num"] == row_num), None)
    if target is None:
        return jsonify({"success": False, "error": f"Row {row_num} not found"}), 404

    target["hidden"]   = True
    target["modified"] = True
    data["meta"]["last_updated"] = datetime.now().isoformat()
    data["meta"]["updated_by"]   = session.get("user", "—")
    _save(data)
    _audit("delete_row", sheet, {"row_num": row_num})
    return jsonify({"success": True})


@sln_budget_bp.route("/api/sln_budget/update_meta", methods=["POST"])
@_login_required
def api_update_meta():
    """Update meta fields (title, fiscal_year, notes, property)."""
    if not _can_edit():
        return jsonify({"success": False, "error": "No permission"}), 403
    body = request.get_json(force=True) or {}
    data = _load()
    for k in ("fiscal_year", "notes", "title", "property"):
        if k in body:
            data["meta"][k] = body[k]
    data["meta"]["last_updated"] = datetime.now().isoformat()
    data["meta"]["updated_by"]   = session.get("user", "—")
    _save(data)
    _audit("update_meta", "meta", body)
    return jsonify({"success": True, "meta": data["meta"]})


@sln_budget_bp.route("/api/sln_budget/sheet_note", methods=["POST"])
@_login_required
def api_sheet_note():
    """Attach a note to a whole sheet."""
    if not _can_edit():
        return jsonify({"success": False, "error": "No permission"}), 403
    body  = request.get_json(force=True) or {}
    sheet = body.get("sheet")
    note  = body.get("note", "")
    if not sheet:
        return jsonify({"success": False, "error": "sheet required"}), 400
    data = _load()
    if sheet not in data.get("sheets", {}):
        return jsonify({"success": False, "error": "Sheet not found"}), 404
    data["sheets"][sheet]["sheet_note"] = note
    data["meta"]["last_updated"] = datetime.now().isoformat()
    data["meta"]["updated_by"]   = session.get("user", "—")
    _save(data)
    _audit("sheet_note", sheet, {"note": note})
    return jsonify({"success": True})


@sln_budget_bp.route("/api/sln_budget/reseed", methods=["POST"])
@_login_required
def api_reseed():
    """
    Re-read Excel and rebuild JSON (preserves user notes).
    Admin / PM only — use after Excel file is updated on disk.
    """
    role = (session.get("role") or "").lower()
    if role not in {"admin", "management", "general manager", "property manager"}:
        return jsonify({"success": False, "error": "Admin / PM role required"}), 403
    if not _XLSX.exists():
        return jsonify({"success": False, "error": f"Excel not found: {_XLSX}"}), 404

    old        = _load()
    old_sheets = old.get("sheets", {})
    new_data   = _seed_from_excel()

    # Restore user notes into freshly-seeded rows
    for sname, sheet in new_data.get("sheets", {}).items():
        old_s   = old_sheets.get(sname, {})
        old_map = {r["row_num"]: r for r in old_s.get("rows", [])}
        for row in sheet.get("rows", []):
            if row["row_num"] in old_map:
                old_r = old_map[row["row_num"]]
                if old_r.get("note"):
                    row["note"] = old_r["note"]
        if old_s.get("sheet_note"):
            sheet["sheet_note"] = old_s["sheet_note"]

    _save(new_data)
    _audit("reseed", "ALL", {"excel": str(_XLSX)})
    return jsonify({"success": True,
                    "msg": "Re-seeded from Excel. User notes preserved.",
                    "seeded_at": new_data["meta"]["seeded_at"]})


@sln_budget_bp.route("/api/sln_budget/audit")
@_login_required
def api_audit():
    """Return audit log (last 200 entries, newest first). Admin/PM only."""
    role = (session.get("role") or "").lower()
    if role not in {"admin", "management", "general manager", "property manager"}:
        return jsonify({"success": False, "error": "Admin role required"}), 403
    try:
        if _AUDIT.exists():
            with open(_AUDIT, "r", encoding="utf-8") as f:
                log = json.load(f)
            return jsonify({"success": True, "log": list(reversed(log[-200:]))})
    except Exception:
        pass
    return jsonify({"success": True, "log": []})


@sln_budget_bp.route("/api/sln_budget/export_json")
@_login_required
def api_export_json():
    """Download full budget JSON."""
    data = _load()
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        json.dumps(data, indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename=sln_budget_{ts}.json"}
    )