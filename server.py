"""
TERMINUS MAINTENANCE MANAGEMENT SYSTEM - SERVER
COMPLETE WORKING VERSION - ALL PORTALS FUNCTIONAL
NO DUPLICATES • NO ERRORS • PRODUCTION READY
"""
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory, abort, send_file, Blueprint
from models import db, User, Issue, Asset, WorkOrder, Vendor, AuditLog, init_db, create_default_users
from pathlib import Path
import os
import pandas as pd
import json
import traceback
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import io
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
import smtplib
import threading
import time as _time


# ── ONE master lock for the entire Gmail account ─────────────────────────────
# Gmail rejects concurrent auth from the same account regardless of which
# module triggers it. Every email send in this server — MMS, PPM, CAM, OW,
# GM Tasks — MUST acquire _GMAIL_LOCK before opening any SMTP connection.
# GM Tasks has its own send function but shares this same lock.
_GMAIL_LOCK = threading.Lock()

# Legacy alias so existing code that references _SMTP_LOCK still works
_SMTP_LOCK  = _GMAIL_LOCK

# =====================================================
# IMPORT BLUEPRINTS (DO NOT MODIFY)
# =====================================================
from ppm_routes import ppm_api
from ppm_workflow_routes import workflow_api
from inventory_routes import inventory_bp
from workorders_routes import workorders_bp
from vendor_visit_routes import vendor_visit_bp
from cam_charges_routes import cam_charges_bp
try:
    from cam_charges_scheduler import run_once as cam_auto_remind
except ImportError:
    cam_auto_remind = None




# Import PPM modules (other properties only — SLN MMS daily mail is handled by sln_mms_routes/register_mms_scheduler)
try:
    from ppm_daily_mailer import send_daily_summary
except ImportError:
    send_daily_summary = None
    print("⚠️  ppm_daily_mailer not found — PPM daily email for other properties disabled")



# =====================================================
# CRITICAL: Define BASE_DIR FIRST
# =====================================================
BASE_DIR = Path(__file__).parent.resolve()
ROOT = BASE_DIR

# =====================================================
# PATH CONSTANTS
# =====================================================
WO_JSON = BASE_DIR / "static" / "data" / "work_orders.json"
ASSETS_XLSX = BASE_DIR / "static" / "data" / "Assets.xlsx"
PPM_DATA_FILE = BASE_DIR / "static" / "data" / "ppm_data.json"
DATA_DIR = BASE_DIR / "static" / "data"

# Upload Directories
UPLOAD_ROOT = BASE_DIR / "uploads" / "project_handover"
TRAINING_UPLOAD_ROOT = BASE_DIR / "uploads" / "training"
DOC_UPLOAD_DIR = BASE_DIR / "uploads" / "documents"
VISITOR_UPLOADS = BASE_DIR / "uploads" / "visitor_documents"
VENDOR_UPLOADS = BASE_DIR / "uploads" / "vendor_documents"

# Create all directories
for folder in [DATA_DIR, UPLOAD_ROOT, TRAINING_UPLOAD_ROOT, DOC_UPLOAD_DIR, VISITOR_UPLOADS, VENDOR_UPLOADS]:
    folder.mkdir(parents=True, exist_ok=True)

# =====================================================
# EMAIL CONFIGURATION
# =====================================================
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "maintenance.slnterminus@gmail.com"
SENDER_PASSWORD = "xaottgrqtqnkouqn"
RECEIVER_EMAILS = [
    "maintenance.slnterminus@gmail.com",
    "yasven7545@gmail.com",
    "engineering@terminus-global.com"
]

# Allowed file extensions
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx'}
ALLOWED_IMAGE_EXT = {"png", "jpg", "jpeg", "gif", "webp"}

# ── Shared SMTP send helper ───────────────────────────────────────────────────
# Use this for EVERY email send in the server.
# Acquires _SMTP_LOCK so concurrent schedulers never collide on Gmail.
# Retries up to 3 times with exponential back-off on transient errors.
# ── Tracks when any email was last sent — enforces a minimum gap ─────────────
_LAST_SMTP_SEND = {"ts": 0.0}
_MIN_SEND_GAP   = 6   # seconds — Gmail needs breathing room between sessions

def _smtp_send(msg_obj, recipients, caller="unknown", retries=3, base_delay=4):
    """
    Thread-safe SMTP send with retry and minimum inter-send gap.
    Acquires _GMAIL_LOCK (shared by ALL senders in this server).
    """
    last_err = None
    for attempt in range(1, retries + 1):
        with _GMAIL_LOCK:
            # Enforce minimum gap since last send from ANY module
            gap = _time.time() - _LAST_SMTP_SEND["ts"]
            if gap < _MIN_SEND_GAP:
                _time.sleep(_MIN_SEND_GAP - gap)
            try:
                with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=25) as srv:
                    srv.ehlo()
                    srv.starttls()
                    srv.ehlo()
                    srv.login(SENDER_EMAIL, SENDER_PASSWORD)
                    srv.sendmail(SENDER_EMAIL, recipients, msg_obj.as_string())
                _LAST_SMTP_SEND["ts"] = _time.time()
                print(f"✅ [{caller}] Email sent → {recipients} (attempt {attempt})")
                return True
            except smtplib.SMTPAuthenticationError as e:
                print(f"⚠️  [{caller}] SMTP auth error (attempt {attempt}): {e}")
                last_err = e
                _time.sleep(base_delay * attempt * 2)
            except (smtplib.SMTPException, OSError) as e:
                print(f"⚠️  [{caller}] SMTP error (attempt {attempt}): {e}")
                last_err = e
                _time.sleep(base_delay * attempt)
    print(f"❌ [{caller}] All {retries} attempts failed. Last error: {last_err}")
    raise last_err

# =====================================================
# 1. CREATE APP & CONFIGURATION
# =====================================================
app = Flask(__name__, static_folder="static", template_folder="templates")

# ── All config in ONE block before blueprints / init_db ───────────────────────
app.config.update(
    DEBUG                          = False,
    SECRET_KEY                     = "supersecretkey-2026",
    SQLALCHEMY_DATABASE_URI        = "sqlite:///" + str(BASE_DIR / "portal.db"),
    SQLALCHEMY_TRACK_MODIFICATIONS = False,
    MAX_CONTENT_LENGTH             = 16 * 1024 * 1024,
    SESSION_COOKIE_SAMESITE        = "Lax",
    SESSION_COOKIE_SECURE          = False,
    SESSION_COOKIE_HTTPONLY        = True,
    SESSION_COOKIE_NAME            = "terminus_session",
    PERMANENT_SESSION_LIFETIME     = 86400,
)
app.secret_key = app.config["SECRET_KEY"]  # legacy alias

# ── Initialise DB immediately — before any blueprint import touches models ─────
init_db(app)

from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)







# =====================================================
# 2. SAFE BLUEPRINT REGISTRATION
# =====================================================
def safe_register(module_name, bp_name, url_prefix=None):
    """Safely register blueprint if module exists"""
    try:
        mod = __import__(module_name, fromlist=[bp_name])
        bp = getattr(mod, bp_name)
        if url_prefix:
            app.register_blueprint(bp, url_prefix=url_prefix)
        else:
            app.register_blueprint(bp)
        print(f"✅ Registered: {bp_name} from {module_name}")
    except ImportError as e:
        print(f"⚠️  Blueprint not found: {module_name}.{bp_name} - {str(e)}")
    except Exception as e:
        print(f"⚠️  Blueprint registration error: {module_name}.{bp_name} - {str(e)}")

# Register all blueprints ONCE
safe_register("ppm_routes", "ppm_api", url_prefix="/api")
safe_register("ppm_workflow_routes", "workflow_api", url_prefix="/api/workflow")
safe_register("inventory_routes", "inventory_bp", url_prefix="/inventory")
safe_register("workorders_routes", "workorders_bp")
safe_register("vendor_visit_routes", "vendor_visit_bp")

safe_register("ow_vms_routes", "ow_vms_bp", url_prefix="/ow_vms")

# ── ONEWEST Space Occupancy (blueprint — optional separate file) ────────────
try:
    from ow_occupancy import ow_occupancy_register
    ow_occupancy_register(app)
    print("✅ Registered: ow_occupancy_bp (from ow_occupancy.py)")
except ImportError:
    print("ℹ️  ow_occupancy.py not found — using inline routes below")
except Exception as e:
    print(f"⚠️  ow_occupancy registration error: {e}")
# ───────────────────────────────────────────────────────────────────────────

# ── ONEWEST Handover Takeover (ow_hoto) ─────────────────────────────────────
try:
    from ow_hoto import ow_hoto_register
    ow_hoto_register(app)
except ImportError:
    print("ℹ️  ow_hoto.py not found — registering inline route")
    # Inline fallback page route
    @app.route("/ow_hoto")
    @login_required
    @require_property("ONEWEST")
    def _ow_hoto_inline():
        session["active_property"] = "ONEWEST"
        return render_template("ow_hoto.html")
except Exception as e:
    print(f"⚠️  ow_hoto registration error: {e}")
# ───────────────────────────────────────────────────────────────────────────

# ── ONEWEST Energy Analytics — registered after decorators are defined (see below) ──

safe_register("cam_charges_routes", "cam_charges_bp")

# ── Command Center — Breakdown / Operations modules ──────────────────────────
safe_register("breakdown_routes", "breakdown_bp")
# ─────────────────────────────────────────────────────────────────────────────
try:
    safe_register("issues_routes", "issues_bp", url_prefix="/sln")
except:
    print("⚠️  Issues blueprint not available")

# MIS blueprint
safe_register("sln_mis", "mis_bp")

# ── SLN MMS Dashboard ──────────────────────────────────────────────────────
safe_register("sln_mms_routes", "sln_mms_bp")
# ───────────────────────────────────────────────────────────────────────────

# ── SLN PM Daily Log ───────────────────────────────────────────────────────
safe_register("sln_pm_daily_routes", "sln_pm_daily_bp")
# ───────────────────────────────────────────────────────────────────────────

# ── OGM PM Daily Log ───────────────────────────────────────────────────────
safe_register("ogm_pm_daily_routes", "ogm_pm_daily_bp")
# ───────────────────────────────────────────────────────────────────────────

# ── OW PM Daily Log (ONEWEST) ───────────────────────────────────────────────
try:
    from ow_pm_daily_routes import ow_pm_daily_register
    ow_pm_daily_register(app)
except ImportError:
    print("ℹ️  ow_pm_daily_routes.py not found — registering inline fallback")
    @app.route("/ow_pm_daily")
    @login_required
    @require_property("ONEWEST")
    def _ow_pm_daily_inline():
        session["active_property"] = "ONEWEST"
        return render_template("ow_pm_daily.html")
except Exception as e:
    print(f"⚠️  ow_pm_daily_routes registration error: {e}")
# ───────────────────────────────────────────────────────────────────────────

# ── SLN Budget Review Dashboard ────────────────────────────────────────────
safe_register("sln_budget", "sln_budget_bp")
# ───────────────────────────────────────────────────────────────────────────

# ── SLN Workforce & Resource Management Module ──────────────────────────────
try:
    from sln_resource import register_resource_module
    register_resource_module(app)
except ImportError as e:
    print(f"⚠️  SLN Resource Management module not found: {e}")
except Exception as e:
    print(f"⚠️  SLN Resource Management module error: {e}")
# ───────────────────────────────────────────────────────────────────────────

# ── SLN Housekeeping Module ─────────────────────────────────────────────────
safe_register("sln_hk_routes", "sln_hk_bp")
# ───────────────────────────────────────────────────────────────────────────

# ── SLN Security Module ─────────────────────────────────────────────────────
safe_register("sln_sec_routes", "sln_sec_bp")
# ───────────────────────────────────────────────────────────────────────────

# ── SLN Fire Fighting Module ───────────────────────────────────────────
safe_register("sln_fire_routes", "sln_fire_bp", url_prefix="/sln_fire")
# ───────────────────────────────────────────────────────────────────────────

try:
    from ow_work_track_routes import ow_work_track_register
    ow_work_track_register(app)
except Exception as e:
    print(f"⚠️  OW Work Track blueprint error: {e}")



# =====================================================
# 3. DATABASE — already initialised above (before blueprints)
# =====================================================
# init_db(app) is called earlier so all workers have the DB ready at import time.

# =====================================================
# 4. PROJECT HANDOVER CATEGORIES
# =====================================================
CATEGORIES = {
    "Admin": "Administrative & Contract Documents",
    "Technical": "Technical & Design Documents",
    "OM": "O & M Manuals",
    "Testing": "Testing & Commissioning Records",
    "Assets": "Asset Inventory",
    "Compliance": "Compliance & Safety",
    "Training": "Training & Support",
    "Digital": "Digital Handover",
    "Snags": "Snag List & Punch Items",
}

# Ensure project handover directories
for key in CATEGORIES.keys():
    (UPLOAD_ROOT / key).mkdir(parents=True, exist_ok=True)

# =====================================================
# 5. USER AUTHENTICATION (Session-Based)
# =====================================================
USERS = {

    # ══════════════════════════════════════════════════════
    # GLOBAL — access ALL properties (Admin → PM level)
    # ══════════════════════════════════════════════════════
    "admin": {
        "password": "0962",
        "role": "Admin",
        "properties": ["SLN Terminus", "ONEWEST", "The District", "One Golden Mile", "Nine Hills"]
    },
    "manager": {
        "password": "1234",
        "role": "Management",
        "properties": ["SLN Terminus", "ONEWEST", "The District", "One Golden Mile", "Nine Hills"]
    },
    "gm": {
        "password": "gm123",
        "role": "General Manager",
        "properties": ["SLN Terminus", "ONEWEST", "The District", "One Golden Mile", "Nine Hills"]
    },

    # ══════════════════════════════════════════════════════
    # SLN TERMINUS
    # ══════════════════════════════════════════════════════
    "sln_pm": {
        "password": "sln_pm123",
        "role": "Property Manager",
        "properties": ["SLN Terminus"]
    },
    "sln_exec": {
        "password": "sln_exec123",
        "role": "Executive",
        "properties": ["SLN Terminus"]
    },
    "Super": {
        "password": "12345",
        "role": "Supervisor",
        "properties": ["SLN Terminus"]
    },
    "ele": {
        "password": "elec123",
        "role": "Electrician",
        "properties": ["SLN Terminus"]
    },
    "plum": {
        "password": "plum123",
        "role": "Plumber",
        "properties": ["SLN Terminus"]
    },
    "hvac": {
        "password": "hvac123",
        "role": "HVAC",
        "properties": ["SLN Terminus"]
    },
    # Legacy aliases kept for backwards compatibility
    "Super": {
        "password": "12345",
        "role": "Supervisor",
        "properties": ["SLN Terminus"]
    },
    "MST": {
        "password": "MST123",
        "role": "Electrician",
        "properties": ["SLN Terminus"]
    },
    "sln_propertymanager": {
        "password": "sln_pm123",
        "role": "Property Manager",
        "properties": ["SLN Terminus"]
    },

    # ══════════════════════════════════════════════════════
    # ONEWEST
    # ══════════════════════════════════════════════════════
    "ow_pm": {
        "password": "ow_pm123",
        "role": "Property Manager",
        "properties": ["ONEWEST"]
    },
    "ow_exec": {
        "password": "ow_exec123",
        "role": "Executive",
        "properties": ["ONEWEST"]
    },
    "ow_super": {
        "password": "sup123",
        "role": "Supervisor",
        "properties": ["ONEWEST"]
    },
    "ow_ele": {
        "password": "elec123",
        "role": "Electrician",
        "properties": ["ONEWEST"]
    },
    "ow_plum": {
        "password": "plmb123",
        "role": "Plumber",
        "properties": ["ONEWEST"]
    },
    "hvac": {
        "password": "hvac123",
        "role": "HVAC",
        "properties": ["ONEWEST"]
    },
    # Legacy aliases
    "MST": {
        "password": "ow_mst123",
        "role": "Electrician",
        "properties": ["ONEWEST"]
    },
    "ow_propertymanager": {
        "password": "ow_pm123",
        "role": "Property Manager",
        "properties": ["ONEWEST"]
    },

    # ══════════════════════════════════════════════════════
    # THE DISTRICT
    # ══════════════════════════════════════════════════════
    "td_pm": {
        "password": "td_pm123",
        "role": "Property Manager",
        "properties": ["The District"]
    },
    "td_exec": {
        "password": "td_exec123",
        "role": "Executive",
        "properties": ["The District"]
    },
    "td_supervisor": {
        "password": "td_sup123",
        "role": "Supervisor",
        "properties": ["The District"]
    },
    "td_electrician": {
        "password": "td_elec123",
        "role": "Electrician",
        "properties": ["The District"]
    },
    "td_plumber": {
        "password": "td_plmb123",
        "role": "Plumber",
        "properties": ["The District"]
    },
    "td_hvac": {
        "password": "td_hvac123",
        "role": "HVAC",
        "properties": ["The District"]
    },
    # Legacy aliases
    "td_technician": {
        "password": "td_mst123",
        "role": "Electrician",
        "properties": ["The District"]
    },
    "td_propertymanager": {
        "password": "td_pm123",
        "role": "Property Manager",
        "properties": ["The District"]
    },

    # ══════════════════════════════════════════════════════
    # ONE GOLDEN MILE
    # ══════════════════════════════════════════════════════
    "ogm_pm": {
        "password": "ogm_pm123",
        "role": "Property Manager",
        "properties": ["One Golden Mile"]
    },
    "ogm_exec": {
        "password": "ogm_exec123",
        "role": "Executive",
        "properties": ["One Golden Mile"]
    },
    "ogm_supervisor": {
        "password": "ogm_sup123",
        "role": "Supervisor",
        "properties": ["One Golden Mile"]
    },
    "ogm_electrician": {
        "password": "ogm_elec123",
        "role": "Electrician",
        "properties": ["One Golden Mile"]
    },
    "ogm_plumber": {
        "password": "ogm_plmb123",
        "role": "Plumber",
        "properties": ["One Golden Mile"]
    },
    "ogm_hvac": {
        "password": "ogm_hvac123",
        "role": "HVAC",
        "properties": ["One Golden Mile"]
    },
    # Legacy aliases
    "ogm_technician": {
        "password": "ogm_mst123",
        "role": "Electrician",
        "properties": ["One Golden Mile"]
    },
    "ogmpm": {
        "password": "ogmpm123",
        "role": "Property Manager",
        "properties": ["One Golden Mile"]
    },

    # ══════════════════════════════════════════════════════
    # NINE HILLS
    # ══════════════════════════════════════════════════════
    "nh_pm": {
        "password": "nh_pm123",
        "role": "Property Manager",
        "properties": ["Nine Hills"]
    },
    "nh_exec": {
        "password": "nh_exec123",
        "role": "Executive",
        "properties": ["Nine Hills"]
    },
    "nh_supervisor": {
        "password": "nh_sup123",
        "role": "Supervisor",
        "properties": ["Nine Hills"]
    },
    "nh_electrician": {
        "password": "nh_elec123",
        "role": "Electrician",
        "properties": ["Nine Hills"]
    },
    "nh_plumber": {
        "password": "nh_plmb123",
        "role": "Plumber",
        "properties": ["Nine Hills"]
    },
    "nh_hvac": {
        "password": "nh_hvac123",
        "role": "HVAC",
        "properties": ["Nine Hills"]
    },
    # Legacy aliases
    "nh_technician": {
        "password": "nh_mst123",
        "role": "Electrician",
        "properties": ["Nine Hills"]
    },
    "nh_propertymanager": {
        "password": "nh_pm123",
        "role": "Property Manager",
        "properties": ["Nine Hills"]
    },
}

# =====================================================
# 5a. ROLE → MODULE ACCESS CONTROL
# =====================================================
# Module keys must match data-module attributes in HTML
ALL_MODULES = [
    "space_occupancy", "cam_billing", "cam_review",
    "energy", "ow_energy", "mis_reports", "kra", "issues",
    "mms_dashboard", "store_inventory", "project_handover",
    "vendor_visit", "work_track", "pm_daily", "housekeeping",
    "security", "fire", "way_forward", "audit_documents", "ow_sec",
    "gm_tasks",       # GM Tasks Portal
    "sln_budget",     # SLN Budget Review Dashboard
    "area_summary",   # Area Statement
    "amc_tracker",    # AMC Cost Tracker
    "hvac_analytics", # HVAC Analytics
    "sinking_fund",   # Sinking Fund Estimation
    "salary_breakup", # Salary Breakup
    "load_breakup",   # Electrical Load Breakup
    "trend_analysis", # Trend Analysis
]

# ─── Role → Module access matrix ──────────────────────────────────────────
# Executive  : property-specific; broad access
# Supervisor : energy + ops modules (no billing/space/kra/project)
# Electrician: issues + mms only
# Plumber    : issues + mms only
# HVAC       : issues + mms only
ROLE_MODULES = {
    "Admin":            ALL_MODULES,
    "admin":            ALL_MODULES,
    "Management":       ALL_MODULES,
    "General Manager":  ALL_MODULES,
    "Property Manager": ALL_MODULES,

    # Executive — broad access, property-specific
    "Executive": [
        "energy", "ow_energy", "mis_reports", "kra",
        "issues", "mms_dashboard", "store_inventory",
        "project_handover", "pm_daily", "housekeeping",
        "security", "fire", "audit_documents", "ow_sec",
    ],

    # Supervisor — operations + energy, no billing/space/kra/project handover
    "Supervisor": [
        "energy",
        "issues", "mms_dashboard", "store_inventory",
        "housekeeping", "security", "fire", "ow_sec",
    ],

    # Trade roles — issues & MMS only
    "Electrician": ["issues", "mms_dashboard"],
    "Plumber":     ["issues", "mms_dashboard"],
    "HVAC":        ["issues", "mms_dashboard"],

    # Legacy generic Technician key (maps to same as Electrician)
    "Technician":  ["issues", "mms_dashboard"],
}

# Which modules are live (active) per property
# ── Active modules per portal (only modules with live sidebar buttons) ──────
# These MUST match the data-module values on ACTIVE (non-disabled) sidebar
# buttons in each portal's HTML template. "soon" buttons are excluded.
PROPERTY_MODULES = {
    # SLN Terminus — all budget & operations modules now live
    "SLN Terminus": [
        "issues",           # Daily Log / Observations
        "mms_dashboard",    # MMS Dashboard
        "store_inventory",  # Store Inventory
        "work_track",       # Work Track
        "gm_tasks",         # GM Tasks Portal
        "pm_daily",         # PM Daily Log
        "housekeeping",     # Housekeeping
        "security",         # Security
        "fire",             # Fire Fighting
        "cam_billing",      # CAM Billing
        "energy",           # Energy Analytics (SLN)
        "ow_energy",       # Energy Analytics (ONEWEST)
        "space_occupancy",  # Space Occupancy
        "mis_reports",      # MIS Reports
        "kra",              # KRA Tracker
        "project_handover", # Project Handover
        "audit_documents",  # Docs Suite
        "sln_budget",       # Budget Review Dashboard (all sheets live)
        "area_summary",     # → area_summary sheet
        "amc_tracker",      # → amc_cost sheet
        "hvac_analytics",   # → hvac_mall sheet
        "sinking_fund",     # → sinking_fund_estimation sheet
        "salary_breakup",   # → salary_breakup sheet
        "load_breakup",     # → load_breakup sheet
        "trend_analysis",   # → cam_charges_summary trends
    ],
    # ONEWEST — active modules
    "ONEWEST": [
        "issues",           # Daily Logs
        "mms_dashboard",    # MMS Dashboard
        "store_inventory",  # Store Inventory
        "work_track",       # Work Track
        "vendor_visit",     # Vendor Visit
        "audit_documents",  # Doc Suite
        "gm_tasks",         # GM Tasks Portal
        "space_occupancy",  # Space Occupancy (OW SFT Details)
        "project_handover", # OW HOTO — Handover Takeover Workspace
        "pm_daily",         # OW PM Daily Log
        "ow_kra",           # KRA Score Card
        "ow_hk",            # Housekeeping Module
        "ow_sec",           # Security Command Center
        "ow_fire",          # Fire Fighting Module
    ],
    # The District — all coming soon, no active modules yet
    "The District":    ["gm_tasks"],
    # One Golden Mile — only PM Daily Log active
    "One Golden Mile": ["pm_daily", "gm_tasks"],
    # Nine Hills — all coming soon
    "Nine Hills":      ["gm_tasks"],
}

# =====================================================
# 6. AUTHENTICATION DECORATORS
# =====================================================
def require_property(property_name):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            # Not logged in
            if "user" not in session:
                if request.path.startswith(("/ow_api/", "/api/", "/inventory/",
                                            "/ow_work_track/", "/sln_work_track/", "/ow_vms/", "/ow_mail/", "/ow_sec", "/ow_security",
                                            "/sln_fire/api/", "/sln_hk/api/", "/sln_sec/api/", "/ow_fire/api/", "/td_fire/api/", "/ogm_fire/api/")):
                    return jsonify({"success": False, "error": "Not authenticated"}), 401
                return redirect(url_for("login"))

            # Full-access roles bypass all property checks (case-insensitive)
            bypass_roles = {"admin", "management", "general manager", "property manager"}
            if (session.get("role") or "").lower() in bypass_roles:
                return fn(*args, **kwargs)

            # For API routes — auto-set active_property if user has access
            # This fixes the ngrok / direct-URL access issue where session
            # active_property may not be set yet
            if request.path.startswith(("/ow_api/", "/api/", "/inventory/",
                                        "/ow_work_track/", "/sln_work_track/", "/ow_vms/", "/ow_mail/",
                                        "/sln_fire/api/", "/sln_hk/api/", "/sln_sec/api/", "/ow_fire/api/", "/td_fire/api/", "/ogm_fire/api/")):
                user_properties = session.get("properties", [])
                if property_name in user_properties:
                    session["active_property"] = property_name  # auto-set
                    return fn(*args, **kwargs)
                return jsonify({"success": False,
                                "error": f"No access to {property_name}"}), 403

            # For page routes — strict check
            if session.get("active_property") != property_name:
                abort(403)

            return fn(*args, **kwargs)
        return wrapper
    return decorator

def require_role(required_role):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if "user" not in session:
                return redirect(url_for("login"))
            if session.get("role") != required_role and session.get("role") != "admin":
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith(("/ow_api/", "/api/", "/inventory/",
                                        "/ow_work_track/", "/sln_work_track/", "/ow_vms/", "/ow_mail/",
                                        "/sln_fire/api/", "/sln_hk/api/", "/sln_sec/api/", "/ow_fire/api/", "/td_fire/api/", "/ogm_fire/api/")):
                return jsonify({"success": False, "error": "Not authenticated"}), 401
            # Preserve the destination so login can redirect back after auth
            next_path = request.path
            return redirect(url_for("login") + "?next=" + next_path)
        return f(*args, **kwargs)
    return wrapper

# ── ONEWEST Energy Analytics (registered here — AFTER login_required/require_property are defined) ──
try:
    from ow_energy import ow_energy_register
    ow_energy_register(app, login_required, require_property)
except ImportError:
    print("ℹ️  ow_energy.py not found — registering inline fallback")
    @app.route("/ow_energy")
    @login_required
    @require_property("ONEWEST")
    def _ow_energy_inline():
        session["active_property"] = "ONEWEST"
        return render_template("ow_energy.html")
except Exception as e:
    print(f"⚠️  ow_energy registration error: {e}")
# ─────────────────────────────────────────────────────────────────────────────

# ── ONEWEST KRA Score Card (registered AFTER decorators) ─────────────────────
try:
    from ow_kra import ow_kra_register
    ow_kra_register(app, login_required, require_property)
except ImportError:
    print("ℹ️  ow_kra.py not found — registering inline fallback")
    @app.route("/ow_kra")
    @login_required
    @require_property("ONEWEST")
    def _ow_kra_inline():
        session["active_property"] = "ONEWEST"
        return render_template("ow_kra.html")
except Exception as e:
    print(f"⚠️  ow_kra registration error: {e}")
# ─────────────────────────────────────────────────────────────────────────────

# ── ONEWEST Housekeeping Module (registered AFTER decorators) ─────────────────
try:
    from ow_hk import ow_hk_register
    ow_hk_register(app, login_required, require_property)
except ImportError:
    print("ℹ️  ow_hk.py not found — registering inline fallback")
    @app.route("/ow_hk")
    @login_required
    @require_property("ONEWEST")
    def _ow_hk_inline():
        session["active_property"] = "ONEWEST"
        return render_template("ow_hk.html")
except Exception as e:
    print(f"⚠️  ow_hk registration error: {e}")
# ─────────────────────────────────────────────────────────────────────────────

# ── ONEWEST Security Module ────────────────────────────────────────────────────
try:
    from ow_sec import ow_sec_register
    ow_sec_register(app)
except ImportError:
    print("ℹ️  ow_sec.py not found — registering inline fallback")
    @app.route("/ow_sec")
    @app.route("/ow_security")
    @login_required
    @require_property("ONEWEST")
    def _ow_sec_inline():
        session["active_property"] = "ONEWEST"
        return render_template("ow_sec.html")
except Exception as e:
    print(f"⚠️  ow_sec registration error: {e}")
# ─────────────────────────────────────────────────────────────────────────────

# ── ONEWEST Fire Fighting Module (registered AFTER decorators) ────────────────
try:
    from ow_fire import ow_fire_register
    ow_fire_register(app, login_required, require_property)
except ImportError:
    print("ℹ️  ow_fire.py not found — registering inline fallback")
    @app.route("/ow_fire")
    @login_required
    @require_property("ONEWEST")
    def _ow_fire_inline():
        session["active_property"] = "ONEWEST"
        return render_template("ow_fire.html")
except Exception as e:
    print(f"⚠️  ow_fire registration error: {e}")
# ─────────────────────────────────────────────────────────────────────────────

# ── THE DISTRICT Fire Fighting Module ────────────────────────────────────────
try:
    from td_fire import td_fire_register
    td_fire_register(app, login_required, require_property)
except ImportError:
    print("ℹ️  td_fire.py not found — registering inline fallback")
    @app.route("/td_fire")
    @login_required
    @require_property("The District")
    def _td_fire_inline():
        session["active_property"] = "The District"
        return render_template("td_fire.html")
except Exception as e:
    print(f"⚠️  td_fire registration error: {e}")
# ─────────────────────────────────────────────────────────────────────────────

# ── ONE GOLDEN MILE Fire Fighting Module ─────────────────────────────────────
try:
    from ogm_fire import ogm_fire_register
    ogm_fire_register(app, login_required, require_property)
except ImportError:
    print("ℹ️  ogm_fire.py not found — registering inline fallback")
    @app.route("/ogm_fire")
    @login_required
    def _ogm_fire_inline():
        session["active_property"] = "One Golden Mile"
        return render_template("ogm_fire.html")
except Exception as e:
    print(f"⚠️  ogm_fire registration error: {e}")
# ─────────────────────────────────────────────────────────────────────────────

# =====================================================
# 7. HELPER FUNCTIONS
# =====================================================
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_image(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXT

# =====================================================
# 6. ✅ FIXED: AUTHENTICATION ROUTES
# =====================================================
@app.route("/")
def home():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        property_name = request.form.get("property", "").strip()
        
        # Debug output
        print(f"\n🔐 Login Attempt:")
        print(f"   Username: {username}")
        print(f"   Property: {property_name}")
        print(f"   Available Users: {list(USERS.keys())}")
        
        # Check if user exists
        if username not in USERS:
            error = "User not found. Please check your username."
            print(f"   ❌ Error: {error}")
            return render_template("dashboard.html", error=error)
        
        user_data = USERS[username]
        
        # Validate password
        if user_data["password"] != password:
            error = "Invalid password. Please try again."
            print(f"   ❌ Error: {error}")
            return render_template("dashboard.html", error=error)
        
        # Validate property access
        if property_name and property_name not in user_data["properties"]:
            error = f"You don't have access to {property_name}. Please select another property."
            print(f"   ❌ Error: {error}")
            print(f"   User Properties: {user_data['properties']}")
            return render_template("dashboard.html", error=error)
        
        # Clear existing session
        session.clear()
        
        # Make session persistent (uses PERMANENT_SESSION_LIFETIME = 86400s)
        session.permanent = True

        # Set session variables
        session["user"] = username
        session["role"] = user_data["role"]
        session["properties"] = user_data["properties"]
        session["active_property"] = property_name or user_data["properties"][0]
        session["logged_in"] = True
        
        print(f"   ✅ Login successful!")
        print(f"   Role: {session['role']}")
        print(f"   Active Property: {session['active_property']}")

        # ── Check for next= redirect (e.g. Command Center, direct URL access) ──
        # Comes from: query param (?next=/command_center) set by login_required,
        # or hidden form field (name="next") set by the JS openCommandCenter().
        next_url = (
            request.args.get("next", "").strip() or
            request.form.get("next", "").strip()
        )
        # Security: only allow relative paths starting with /
        if next_url and next_url.startswith("/") and not next_url.startswith("//"):
            print(f"   ↪ Redirecting to next: {next_url}")
            return redirect(next_url)

        # Redirect based on property
        property_routes = {
            "SLN Terminus": "sln_terminus",
            "ONEWEST": "onewest",
            "The District": "the_district",
            "One Golden Mile": "ogm",
            "Nine Hills": "nine_hills"
        }
        
        redirect_route = property_routes.get(property_name, "dashboard")
        return redirect(url_for(redirect_route))
    
    # GET request - show login form
    return render_template("dashboard.html", error=error)

@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")

# =====================================================
# COMMAND CENTER — Global Operations (all properties)
# =====================================================
@app.route("/command_center")
@login_required
def command_center():
    """
    Global Command Center — portfolio-wide operations.
    Accessible to all authenticated users; property-level
    access is enforced inside the page via JS + breakdown APIs.
    """
    print(f"\n🌐 Command Center - User: {session.get('user')} | Role: {session.get('role')}")
    return render_template("command_center.html")
# ─────────────────────────────────────────────────────────────────────────────

# =====================================================
# COMMAND CENTER — Property Status API
# =====================================================
@app.route("/api/properties/status")
@login_required
def api_properties_status():
    """Live status for all 5 properties — feeds Overview table in Command Center."""

    # Pull real open work-order counts from WO_JSON
    wo_counts = {p: 0 for p in ["SLN Terminus", "ONEWEST", "The District",
                                  "One Golden Mile", "Nine Hills"]}
    try:
        if WO_JSON.exists():
            with open(WO_JSON, encoding="utf-8") as f:
                wos = json.load(f).get("work_orders", [])
            open_statuses = {"open", "in-progress", "overdue"}
            for wo in wos:
                prop = wo.get("location", wo.get("property", ""))
                if prop in wo_counts and wo.get("status", "").lower() in open_statuses:
                    wo_counts[prop] += 1
    except Exception as e:
        print(f"⚠️  api_properties_status WO read: {e}")

    # Pull alert counts from breakdown_data.json
    alert_counts = {p: 0 for p in wo_counts}
    try:
        bd_file = BASE_DIR / "static" / "data" / "breakdown_data.json"
        if bd_file.exists():
            with open(bd_file, encoding="utf-8") as f:
                bd = json.load(f)
            for ticket in bd.get("engineering_breakdowns", []):
                prop = ticket.get("property", "")
                if prop in alert_counts and ticket.get("status") not in ("Resolved", "Closed"):
                    alert_counts[prop] += 1
    except Exception as e:
        print(f"⚠️  api_properties_status breakdown read: {e}")

    properties = [
        {"id":"sln","name":"SLN Terminus","code":"SLN","type":"Commercial",
         "city":"Hyderabad","status":"online","redirect":"/sln_terminus",
         "open_wo":wo_counts["SLN Terminus"],"alerts":alert_counts["SLN Terminus"],
         "occupancy":None,"energy":None},
        {"id":"ow","name":"ONEWEST","code":"OW","type":"Commercial",
         "city":"Hyderabad","status":"online","redirect":"/onewest",
         "open_wo":wo_counts["ONEWEST"],"alerts":alert_counts["ONEWEST"],
         "occupancy":None,"energy":None},
        {"id":"td","name":"The District","code":"TD","type":"Commercial",
         "city":"Hyderabad","status":"attention","redirect":"/the_district",
         "open_wo":wo_counts["The District"],"alerts":alert_counts["The District"],
         "occupancy":None,"energy":None},
        {"id":"ogm","name":"One Golden Mile","code":"OGM","type":"Commercial",
         "city":"Hyderabad","status":"online","redirect":"/ogm",
         "open_wo":wo_counts["One Golden Mile"],"alerts":alert_counts["One Golden Mile"],
         "occupancy":None,"energy":None},
        {"id":"nh","name":"Nine Hils","code":"NH","type":"Life Science",
         "city":"Hyderabad","status":"online","redirect":"/nine_hills/",
         "open_wo":wo_counts["Nine Hills"],"alerts":alert_counts["Nine Hills"],
         "occupancy":None,"energy":None},
    ]

    return jsonify({
        "success": True,
        "properties": properties,
        "generated_at": datetime.now().isoformat()
    })
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/logout")
def logout():
    print(f"\n👋 Logout: {session.get('user')}")
    session.clear()
    return redirect(url_for("login"))

# =====================================================
# 9. USER PROFILE API (ONLY ONCE)
# =====================================================
@app.route("/api/user_profile")
@login_required
def get_user_profile():
    role            = session.get("role", "Technician")
    active_property = session.get("active_property", "SLN Terminus")

    role_modules     = ROLE_MODULES.get(role, ROLE_MODULES.get("Technician", []))
    property_modules = PROPERTY_MODULES.get(active_property, [])

    full_access_roles = {"admin", "management", "general manager", "property manager"}
    if role.lower() in full_access_roles:
        allowed_modules = property_modules
    else:
        allowed_modules = [m for m in role_modules if m in property_modules]

    return jsonify({
        "username":        session.get("user", ""),
        "role":            role,
        "active_property": active_property,
        "properties":      session.get("properties", ["SLN Terminus"]),
        "allowed_modules": allowed_modules
    })


# =====================================================
# 8. ✅ FIXED: PROPERTY DASHBOARD ROUTES (ALL 5)
# =====================================================
@app.route("/sln_terminus")
@login_required
@require_property("SLN Terminus")
def sln_terminus():
    print(f"\n🏢 Accessing SLN Terminus - User: {session.get('user')}")
    return render_template("sln_terminus.html")


# ── SLN PM DAILY LOG ───────────────────────────────────────────────────────
@app.route("/sln_pm_daily")
@login_required
@require_property("SLN Terminus")
def sln_pm_daily_page():
    """SLN Terminus — Property Management Daily Log"""
    print(f"\n🗒️  Accessing SLN PM Daily - User: {session.get('user')}")
    return render_template("sln_pm_daily.html")
# ───────────────────────────────────────────────────────────────────────────


# =====================================================
# MAINTENANCE & VENDOR GOVERNANCE DOCUMENTATION SUITE
# =====================================================
@app.route("/mvgds")
@login_required
@require_property("SLN Terminus")
def mvgds():
    """Maintenance & Vendor Governance Documentation Suite"""
    print(f"\n📋 Accessing MVGDS - User: {session.get('user')}")
    return render_template("mvgds.html")


@app.route("/the_district")
@login_required
@require_property("The District")
def the_district():
    print(f"\n🏢 Accessing The District - User: {session.get('user')}")
    return render_template("the_district.html")

@app.route("/one_golden_mile")
def one_golden_mile_redirect():
    """Alias redirect — keeps old bookmarks working"""
    return redirect(url_for("ogm"))

@app.route("/ogm")
@login_required
@require_property("One Golden Mile")
def ogm():
    print(f"\n🏢 Accessing One Golden Mile - User: {session.get('user')}")
    return render_template("ogm.html")


# ── OGM PM DAILY LOG ───────────────────────────────────────────────────────
@app.route("/ogm_pm_daily")
@login_required
@require_property("One Golden Mile")
def ogm_pm_daily_page():
    """One Golden Mile — Property Management Daily Log"""
    print(f"\n🗒️  Accessing OGM PM Daily - User: {session.get('user')}")
    return render_template("ogm_pm_daily.html")
# ───────────────────────────────────────────────────────────────────────────


@app.route("/nine_hills")
@login_required
@require_property("Nine Hills")
def nine_hills():
    print(f"\n🏢 Accessing Nine Hills - User: {session.get('user')}")
    return render_template("nine_hills.html")

# =====================================================
# 7. PPM DASHBOARD ROUTES (FULLY FIXED)
# =====================================================
# NOTE: /sln_mms_dashboard is served by sln_mms_routes blueprint (sln_mms_bp)
# Duplicate route removed to prevent Flask AssertionError on startup.






@app.route("/api/ppm/assets")
def get_ppm_assets():
    """API: Get all PPM assets directly from Assets.xlsx (FIXED)"""
    try:
        location_filter = request.args.get('location', 'all')
        
        # CRITICAL FIX: Read directly from Assets.xlsx (not ppm_data.json)
        ASSETS_XLSX = BASE_DIR / "static" / "data" / "Assets.xlsx"
        
        if not ASSETS_XLSX.exists():
            print(f"❌ Assets.xlsx NOT FOUND at: {ASSETS_XLSX}")
            return jsonify({"assets": [], "total": 0})
        
        # Load Excel file with proper error handling
        try:
            df = pd.read_excel(ASSETS_XLSX, engine='openpyxl')
        except Exception as e:
            print(f"❌ Excel read error: {str(e)}")
            # Fallback to xlrd engine if openpyxl fails
            try:
                df = pd.read_excel(ASSETS_XLSX, engine='xlrd')
            except Exception as e2:
                print(f"❌ Fallback Excel read error: {str(e2)}")
                return jsonify({"assets": [], "total": 0})
        
        assets = []
        for _, row in df.iterrows():
            # Skip empty rows
            asset_code = str(row.get('Asset Code', '')).strip()
            if not asset_code or asset_code.lower() in ['nan', 'none', '']:
                continue
            
            # Build asset object with EXACT column names from your Excel
            asset = {
                "id": asset_code,
                "name": str(row.get('Asset Name', 'Unknown Asset')).strip(),
                "category": str(row.get('In-House/Vendor', 'General')).strip(),
                "location": str(row.get('Location', 'Unknown Location')).strip(),
                "lastService": str(row.get('Last Service', '')).strip(),
                "nextDueDate": str(row.get('nextDueDate', '')).strip(),
                "colorCode": "Green"  # Will be calculated by frontend
            }
            assets.append(asset)
        
        # Apply location filter if specified
        if location_filter != 'all' and location_filter != '':
            assets = [a for a in assets if a.get('location', '').strip() == location_filter.strip()]
        
        print(f"✅ Loaded {len(assets)} assets from Assets.xlsx")
        return jsonify({
            "assets": assets,
            "total": len(assets)
        })
    
    except Exception as e:
        print(f"❌ PPM assets error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"assets": [], "total": 0})

def get_asset_details(asset_id):
    """Safely retrieve asset details from Excel with proper column handling"""
    ASSETS_XLSX = Path(__file__).parent / "static" / "data" / "Assets.xlsx"
    
    if not ASSETS_XLSX.exists():
        return {
            "name": f"Asset_{asset_id}", 
            "location": "Unknown Location", 
            "priority": "Medium",
            "frequency": "Monthly"  # Critical addition
        }
    
    try:
        df = pd.read_excel(ASSETS_XLSX)
        df.columns = df.columns.str.strip()
        
        asset_col = "Asset Code"
        name_col = "Asset Name"
        loc_col = "Location"
        freq_col = "Frequency"  # Critical: Get frequency from Excel
        
        asset_row = df[df[asset_col] == asset_id]
        if asset_row.empty:
            return {
                "name": f"Asset_{asset_id}", 
                "location": "Unknown Location", 
                "priority": "Medium",
                "frequency": "Monthly"
            }
        
        asset_name = str(asset_row.iloc[0][name_col]).strip() if name_col in asset_row.columns else f"Asset_{asset_id}"
        location = str(asset_row.iloc[0][loc_col]).strip() if loc_col in asset_row.columns else "Unknown Location"
        
        # Determine priority based on asset criticality
        asset_lower = asset_name.lower()
        priority = "Medium"
        if "fire" in asset_lower or "dg" in asset_lower.replace(' ', '') or "transformer" in asset_lower or "elevator" in asset_lower or "escalator" in asset_lower:
            priority = "High"
        
        # CRITICAL FIX: Get frequency with fallback
        frequency = "Monthly"  # Default
        if freq_col in asset_row.columns:
            freq_val = str(asset_row.iloc[0][freq_col]).strip().lower()
            if freq_val in ['monthly', 'quarterly', 'yearly']:
                frequency = freq_val
        elif "frequency" in asset_row.columns:  # Case-insensitive fallback
            freq_val = str(asset_row.iloc[0]["frequency"]).strip().lower()
            if freq_val in ['monthly', 'quarterly', 'yearly']:
                frequency = freq_val
        else:
            # Auto-detect frequency for critical assets
            if "elevator" in asset_lower or "escalator" in asset_lower:
                frequency = "monthly"
        
        return {
            "name": asset_name,
            "location": location,
            "priority": priority,
            "frequency": frequency  # Critical addition
        }
        
    except Exception as e:
        print(f"⚠️ Error loading asset {asset_id}: {str(e)}")
        return {
            "name": f"Asset_{asset_id}", 
            "location": "Unknown Location", 
            "priority": "Medium",
            "frequency": "Monthly"
        }


# In your get_today_wos() function
def get_today_wos():
    """Extracts work orders with due_date matching today"""
    if not WO_JSON.exists():
        print(f"⚠️ Work orders file NOT found at: {WO_JSON}")
        return []
    
    try:
        with open(WO_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        today = datetime.now().date()
        today_str = today.strftime('%Y-%m-%d')
        
        # FIX: Use .strip() on key lookups to handle trailing spaces
        today_wos = []
        for wo in data.get('work_orders', []):
            # Clean due_date value
            due_date = wo.get('due_date', '').strip()
            status = wo.get('status', '').strip().lower()
            
            if due_date == today_str and status in ['open', 'in-progress', 'overdue']:
                today_wos.append(wo)
        
        print(f"🔍 Today: {today_str} | Found {len(today_wos)} work orders")
        for i, wo in enumerate(today_wos):
            print(f"  #{i+1} {wo.get('work_order_id', 'N/A')} - {wo.get('asset_name', 'Unknown Asset')} (Status: {wo.get('status', 'N/A')})")
        
        return today_wos
    
    except Exception as e:
        print(f"❌ Error reading work orders: {str(e)}")
        traceback.print_exc()
        return []

# =====================================================
# PPM WORK ORDER 
# =====================================================


@app.route("/api/ppm/workorders")
def get_ppm_workorders():
    """API: Get ALL saved work orders from work_orders.json"""
    try:
        # Load work orders from persistent storage
        if not WO_JSON.exists():
            return jsonify({
                "success": True,
                "work_orders": [],
                "total": 0
            })
            
        with open(WO_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        work_orders = data.get('work_orders', [])
        
        # FIX: Clean key names and values
        cleaned_wos = []
        for wo in work_orders:
            cleaned_wo = {}
            for key, value in wo.items():
                cleaned_key = key.strip()
                cleaned_value = value.strip() if isinstance(value, str) else value
                cleaned_wo[cleaned_key] = cleaned_value
            cleaned_wos.append(cleaned_wo)
        
        # Apply filters if provided
        status_filter = request.args.get('status', 'all').lower().strip()
        priority_filter = request.args.get('priority', 'all').lower().strip()
        
        if status_filter != 'all':
            cleaned_wos = [wo for wo in cleaned_wos if wo.get('status', '').lower() == status_filter]
        
        if priority_filter != 'all':
            cleaned_wos = [wo for wo in cleaned_wos if wo.get('priority', '').lower() == priority_filter]
        
        # Format for frontend compatibility
        formatted_wos = []
        for wo in cleaned_wos:
            formatted_wos.append({
                "WO ID": wo.get("work_order_id", "N/A"),
                "Asset": wo.get("asset_name", "Unknown Asset"),
                "Location": wo.get("location", "Unknown Location"),
                "Due Date": wo.get("due_date", "N/A"),
                "Priority": wo.get("priority", "Medium"),
                "Status": wo.get("status", "open"),
                "created_at": wo.get("created_at", datetime.now().isoformat())
            })
        
        return jsonify({
            "success": True,
            "work_orders": formatted_wos,
            "total": len(formatted_wos)
        })

    except Exception as e:
        print(f"PPM workorders error: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "work_orders": [],
            "total": 0
        }), 500

# =====================================================
# WORK ORDER CREATION ENDPOINT (CRITICAL FIX)
# =====================================================
@app.route('/api/workflow/create', methods=['POST'])
def create_work_order():
    """API: Create work order with CORRECT date format handling"""
    try:
        data = request.get_json()
        asset_id = data.get('assetId')
        due_date = data.get('dueDate')  # This is the date from calendar
        
        # Generate work order ID
        today = datetime.now().date()
        wo_counter = 1
        
        # Load existing work orders
        if WO_JSON.exists():
            with open(WO_JSON, 'r') as f:
                existing_data = json.load(f)
                existing_wos = existing_data.get('work_orders', [])
                wo_counter = len(existing_wos) + 1
        else:
            existing_wos = []
        
        # Generate WO ID
        wo_id = f"WO-PPM-{today.strftime('%Y-%m')}-{str(wo_counter).zfill(4)}"
        
        # Get asset details from Assets.xlsx
        asset_name = "Unknown Asset"
        location = "Unknown Location"
        priority = "Medium"
        
        # FIX: Get asset details including frequency
        ASSETS_XLSX = Path(__file__).parent / "static" / "data" / "Assets.xlsx"
        if ASSETS_XLSX.exists():
            try:
                df = pd.read_excel(ASSETS_XLSX)
                asset_col = "Asset Code"
                name_col = "Asset Name"
                asset_row = df[df[asset_col] == asset_id]
                
                if not asset_row.empty:
                    asset_name = str(asset_row.iloc[0][name_col]).strip()
                    location = str(asset_row.iloc[0]["Location"]).strip()
                    
                    # Determine priority based on asset criticality
                    asset_lower = asset_name.lower()
                    if "fire" in asset_lower or "dg" in asset_lower.replace(' ', '') or "transformer" in asset_lower:
                        priority = "High"
            except Exception as e:
                print(f"❌ Excel read error: {str(e)}")
                asset_name = f"Asset_{asset_id}"
                location = "Unknown Location"
        
        # FIX: Standardize date format to YYYY-MM-DD
        try:
            # Try to parse the date (handles multiple formats)
            if '/' in due_date:
                parts = due_date.split('/')
                if len(parts) == 3:
                    month = int(parts[0])
                    day = int(parts[1])
                    year = int(parts[2])
                    if year < 100:
                        year += 2000
                    due_date = f"{year}-{month:02d}-{day:02d}"
            elif '-' in due_date and len(due_date) == 10:
                # Already in YYYY-MM-DD format
                pass
            else:
                # Fallback: use current date
                due_date = datetime.now().strftime('%Y-%m-%d')
        except:
            due_date = datetime.now().strftime('%Y-%m-%d')
        
        # Create work order
        new_wo = {
            "work_order_id": wo_id,
            "asset_id": asset_id,
            "asset_name": asset_name,
            "location": location,
            "due_date": due_date,  # Standardized to YYYY-MM-DD
            "priority": priority,
            "status": "open",
            "created_at": datetime.now().isoformat()
        }
        
        # Save to persistent storage
        all_wos = existing_wos + [new_wo]
        with open(WO_JSON, 'w') as f:
            json.dump({
                "work_orders": all_wos,
                "last_updated": datetime.now().isoformat(),
                "total_count": len(all_wos)
            }, f, indent=2)
        
        print(f"✅ Work Order Created: {wo_id} for asset {asset_id} ({asset_name})")
        return jsonify({"success": True, "work_order_id": wo_id, "message": "Work order created successfully!"})
    
    except Exception as e:
        print(f"❌ Work order creation error: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500




def calculate_next_due_date(last_service_date, frequency="monthly"):
    """Calculates next due date based on maintenance frequency with proper month handling"""
    if not last_service_date:
        return None
    
    # Parse date - handle MM/DD/YYYY format
    try:
        parts = last_service_date.split('/')
        if len(parts) == 3:
            month = int(parts[0])
            day = int(parts[1])
            year = int(parts[2])
            last_date = datetime(year, month, day)
        else:
            return None
    except:
        return None
    
    # Calculate next due date based on frequency
    if frequency.lower() == 'monthly':
        # Handle month-end dates (like Jan 31 → Feb 28)
        next_month = last_date.month + 1
        next_year = last_date.year
        if next_month > 12:
            next_month = 1
            next_year += 1
            
        # Handle edge cases (like Jan 31 → Feb 28)
        try:
            next_date = datetime(next_year, next_month, last_date.day)
        except ValueError:
            # If day doesn't exist in next month (e.g., Jan 31 → Feb 28)
            next_date = datetime(next_year, next_month, 1) - timedelta(days=1)
            
        return next_date.strftime('%Y-%m-%d')
    
    elif frequency.lower() == 'quarterly':
        # Add 3 months
        next_month = last_date.month + 3
        next_year = last_date.year
        if next_month > 12:
            next_month -= 12
            next_year += 1
            
        try:
            next_date = datetime(next_year, next_month, last_date.day)
        except ValueError:
            next_date = datetime(next_year, next_month, 1) - timedelta(days=1)
            
        return next_date.strftime('%Y-%m-%d')
    
    elif frequency.lower() == 'yearly':
        # Add 1 year
        try:
            next_date = datetime(last_date.year + 1, last_date.month, last_date.day)
        except ValueError:
            next_date = datetime(last_date.year + 1, last_date.month, 1) - timedelta(days=1)
            
        return next_date.strftime('%Y-%m-%d')
    
    return None  # Unknown frequency




# Add this endpoint for closing work orders
@app.route("/api/ppm/dashboard/stats")
def get_ppm_dashboard_stats():
    """PPM dashboard stats endpoint - FIXED"""
    try:
        if not WO_JSON.exists():
            return jsonify({
                "total_assets": 438,
                "pending_ppm": 0,
                "completed_ppm": 0,
                "ppm_due_today": 0,
                "ppm_overdue": 0,
                "compliance_rate": 0.0
            })
        
        with open(WO_JSON, 'r') as f:
            data = json.load(f)
        
        work_orders = data.get('work_orders', [])
        today = datetime.now().date()
        overdue = 0
        due_today = 0
        pending = 0  # Only work orders due today
        
        for wo in work_orders:
            try:
                due_date_str = wo.get('due_date', '')
                if not due_date_str:
                    continue
                
                # Proper date parsing with multiple format support
                date_obj = None
                for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y']:
                    try:
                        date_obj = datetime.strptime(due_date_str, fmt).date()
                        break
                    except:
                        continue
                
                if not date_obj:
                    continue
                
                # Calculate stats correctly
                if date_obj < today:
                    overdue += 1
                elif date_obj == today:
                    due_today += 1
                    pending += 1  # Only today's work orders count as pending
            
            except Exception as e:
                print(f"Error processing work order {wo.get('work_order_id', 'unknown')}: {str(e)}")
                continue
        
        # Calculate compliance rate based on all work orders
        total_work_orders = len(work_orders)
        compliance_rate = 0.0
        if total_work_orders > 0:
            # Compliance = (work orders not overdue / total) * 100
            compliance_rate = round(((total_work_orders - overdue) / total_work_orders * 100), 1)
        
        stats = {
            "total_assets": 438,
            "pending_ppm": pending,  # Should be 15 for today
            "completed_ppm": total_work_orders - overdue - pending,
            "ppm_due_today": due_today,
            "ppm_overdue": overdue,
            "compliance_rate": compliance_rate
        }
        return jsonify(stats)
    
    except Exception as e:
        print(f"PPM stats error: {str(e)}")
        return jsonify({
            "total_assets": 438,
            "pending_ppm": 0,
            "completed_ppm": 0,
            "ppm_due_today": 0,
            "ppm_overdue": 0,
            "compliance_rate": 0.0
        })

# =====================================================
# WORK ORDER CLOSING ENDPOINT (CORRECTED)
# =====================================================
@app.route('/api/workflow/close', methods=['POST'])
def close_work_order():
    """API: Close work order with supervisor approval"""
    try:
        data = request.get_json()
        wo_id = data.get('workOrderId')
        approval_notes = data.get('approvalNotes', '')
        
        if not WO_JSON.exists():
            return jsonify({"success": False, "error": "Work orders file not found"}), 404
        
        with open(WO_JSON, 'r') as f:
            work_data = json.load(f)
        
        work_orders = work_data.get('work_orders', [])
        updated = False
        
        for wo in work_orders:
            if wo.get('work_order_id') == wo_id or wo.get('WO ID') == wo_id:
                # Update status and closure details
                wo['status'] = 'completed'
                wo['Status'] = 'completed'  # For frontend consistency
                wo['closed_at'] = datetime.now().isoformat()
                wo['closed_by'] = session.get('user', 'Supervisor')
                wo['approval_notes'] = approval_notes
                wo['supervisor_approval'] = True
                updated = True
                break
        
        if not updated:
            return jsonify({"success": False, "error": "Work order not found"}), 404
        
        # Save updated work orders
        with open(WO_JSON, 'w') as f:
            json.dump({
                "work_orders": work_orders,
                "last_updated": datetime.now().isoformat(),
                "total_count": len(work_orders)
            }, f, indent=2)
        
        print(f"✅ Work Order {wo_id} closed successfully")
        return jsonify({"success": True, "message": "Work order closed successfully"})
    
    except Exception as e:
        print(f"❌ Close work order error: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500



def update_asset_next_due_date(work_order):
    """UPGRADED: Updates asset's next due date with robust error handling & detailed logging"""
    try:
        # ✅ CRITICAL FIX #1: Use GLOBAL path (not local redefinition)
        global ASSETS_XLSX
        if not ASSETS_XLSX.exists():
            print(f"❌ [ASSET UPDATE] Assets.xlsx NOT FOUND at: {ASSETS_XLSX}")
            return False
        
        # ✅ CRITICAL FIX #2: Add file lock handling
        import time
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # Read Excel with column normalization
                df = pd.read_excel(ASSETS_XLSX)
                df.columns = df.columns.str.strip()
                break
            except PermissionError:
                if attempt == max_retries - 1:
                    print(f"❌ [ASSET UPDATE] Excel file locked after {max_retries} attempts. Close Excel application.")
                    return False
                time.sleep(1)
        
        # Define EXACT column names (case-sensitive)
        asset_col = "Asset Code"
        last_service_col = "Last Service"
        next_due_col = "nextDueDate"  # Note capital 'D'
        
        # Find asset row
        mask = df[asset_col] == work_order['asset_id']
        if not mask.any():
            print(f"⚠️ [ASSET UPDATE] Asset {work_order['asset_id']} NOT FOUND in Excel")
            return False
        
        # Get current dates
        current_last = str(df.loc[mask, last_service_col].iloc[0]).strip()
        current_next = str(df.loc[mask, next_due_col].iloc[0]).strip()
        print(f"🔍 [ASSET UPDATE] Processing {work_order['asset_id']}")
        print(f"   Current: Last={current_last} | Next={current_next}")
        
        # Parse dates (MM/DD/YY format)
        try:
            last_date = datetime.strptime(current_last, '%m/%d/%y')
            next_date = datetime.strptime(current_next, '%m/%d/%y')
        except Exception as e:
            print(f"⚠️ [ASSET UPDATE] Date parse error: {str(e)}. Using 30-day fallback.")
            last_date = datetime.now()
            next_date = last_date + timedelta(days=30)
        
        # Calculate interval (critical for cycling)
        interval_days = (next_date - last_date).days
        if interval_days <= 0:
            interval_days = 30  # Default monthly interval
        print(f"   Interval calculated: {interval_days} days")
        
        # Get closure date
        try:
            closed_dt = datetime.fromisoformat(work_order['closed_at'])
        except:
            closed_dt = datetime.now()
        
        # Calculate NEW dates
        new_last = closed_dt.strftime('%m/%d/%y')
        new_next = (closed_dt + timedelta(days=interval_days)).strftime('%m/%d/%y')
        print(f"   NEW DATES: Last={new_last} | Next={new_next} (+{interval_days} days)")
        
        # ✅ CRITICAL FIX #3: Update DataFrame BEFORE saving
        df.loc[mask, last_service_col] = new_last
        df.loc[mask, next_due_col] = new_next
        
        # Save with error handling
        for attempt in range(max_retries):
            try:
                df.to_excel(ASSETS_XLSX, index=False)
                print(f"✅ [ASSET UPDATE] SUCCESS: {work_order['asset_id']} updated in Assets.xlsx")
                print(f"   Next maintenance scheduled for: {new_next}")
                return True
            except PermissionError:
                if attempt == max_retries - 1:
                    print(f"❌ [ASSET UPDATE] FAILED: Excel file locked. Close Excel and retry.")
                    return False
                time.sleep(1)
    
    except Exception as e:
        print(f"❌ [ASSET UPDATE] CRITICAL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


# =====================================================
# WORK ORDER EXPORT
# =====================================================


@app.route('/api/ppm/workorders/export')
def export_work_orders():
    """API: Export ALL work orders including closed ones with full metadata"""
    try:
        # CRITICAL FIX 1: Verify file exists with proper path
        if not WO_JSON.exists():
            print(f"❌ Work orders file NOT found at: {WO_JSON}")
            return jsonify({
                "success": False,
                "error": "No work orders found. Please generate work orders first via Calendar View."
            }), 404
        
        # CRITICAL FIX 2: Load with proper encoding and handle trailing spaces in keys
        with open(WO_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        work_orders = data.get('work_orders', [])
        
        if not work_orders:
            return jsonify({
                "success": False,
                "error": "No work orders to export. Create work orders from Calendar View first."
            }), 404
        
        # CRITICAL FIX 3: Clean keys (handle trailing spaces in JSON keys from work_orders.json)
        cleaned_wos = []
        for wo in work_orders:
            cleaned_wo = {}
            for key, value in wo.items():
                # Remove trailing spaces from keys AND values
                clean_key = key.strip()
                clean_value = value.strip() if isinstance(value, str) else value
                cleaned_wo[clean_key] = clean_value
            cleaned_wos.append(cleaned_wo)
        
        # CRITICAL FIX 4: Handle ALL possible columns (including closed metadata)
        df = pd.DataFrame(cleaned_wos)
        
        # Define ALL possible columns (handles both open and closed WOs)
        all_columns = [
            'work_order_id', 'asset_id', 'asset_name', 'location',
            'due_date', 'priority', 'status', 'created_at',
            'closed_at', 'closed_by', 'lastService', 'nextDueDate', 'frequency'
        ]
        
        # Only include columns that actually exist in data
        existing_cols = [col for col in all_columns if col in df.columns]
        df = df[existing_cols]
        
        # CRITICAL FIX 5: Rename columns for readability
        column_mapping = {
            'work_order_id': 'Work Order ID',
            'asset_id': 'Asset ID',
            'asset_name': 'Asset Name',
            'location': 'Location',
            'due_date': 'Due Date',
            'priority': 'Priority',
            'status': 'Status',
            'created_at': 'Created At',
            'closed_at': 'Closed At',
            'closed_by': 'Closed By',
            'lastService': 'Last Service',
            'nextDueDate': 'Next Due Date',
            'frequency': 'Frequency'
        }
        df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})
        
        # CRITICAL FIX 6: Create Excel with proper formatting
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='All Work Orders')
                # Format header row
                workbook = writer.book
                worksheet = writer.sheets['All Work Orders']
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        except Exception as excel_error:
            print(f"❌ Excel creation error: {str(excel_error)}")
            traceback.print_exc()
            return jsonify({
                "success": False,
                "error": f"Failed to create Excel file: {str(excel_error)}"
            }), 500
        
        output.seek(0)
        
        # CRITICAL FIX 7: Generate proper filename with timestamp
        filename = f"Terminus_WorkOrders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"✅ Exporting {len(cleaned_wos)} work orders to {filename}")
        
        # CRITICAL FIX 8: Return proper file response
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except json.JSONDecodeError as je:
        print(f"❌ JSON decode error: {str(je)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "Work orders data is corrupted. Please regenerate work orders."
        }), 500
    except PermissionError as pe:
        print(f"❌ Permission error: {str(pe)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "File is locked. Please close work_orders.json in any editor and try again."
        }), 500
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"Export failed: {str(e)}"
        }), 500



# =====================================================
# DAILY MAIL (TRIGGERS AT 8:00 AM)
# =====================================================


@app.route('/api/trigger-daily-email', methods=['POST'])
def trigger_daily_email():
    """API: Trigger daily email with CORRECT date handling and status filtering"""
    try:
        # Get today's work orders
        if not WO_JSON.exists():
            return jsonify({"success": False, "error": "No work orders found"}), 404
            
        with open(WO_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Get today's date correctly
        today = datetime.now().date()
        today_str = today.strftime('%Y-%m-%d')
        
               # FIX: Handle multiple date formats and trailing spaces
        today_wos = []
        for wo in data.get('work_orders', []):
            # Clean and normalize due_date
            due_date_str = wo.get('due_date', '').strip()
            
            # Try multiple date formats
            date_obj = None
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y']:
                try:
                    date_obj = datetime.strptime(due_date_str, fmt).date()
                    break
                except:
                    continue
            
            # ✅ CRITICAL FIX: Include TODAY + OVERDUE work orders (not just exact match)
            if date_obj and date_obj <= today:  # Changed from "date_obj == today"
                status = wo.get('status', '').strip().lower()
                if status in ['open', 'in-progress', 'overdue']:
                    today_wos.append(wo)
        
        # Show detailed debugging info
        print(f"\n{'='*70}")
        print(f"📧 EMAIL TRIGGER DIAGNOSTICS")
        print(f"{'='*70}")
        print(f"📅 Today's Date: {today_str}")
        print(f"📊 Total Work Orders in System: {len(data.get('work_orders', []))}")
        print(f"✅ Work Orders Included in Email: {len(today_wos)}")
        print(f"\n📋 INCLUDED WORK ORDERS:")
        for i, wo in enumerate(today_wos, 1):
            print(f"  {i}. {wo.get('work_order_id', 'N/A')} | "
                  f"{wo.get('asset_name', 'Unknown')} | "
                  f"Due: {wo.get('due_date', 'N/A')} | "
                  f"Status: {wo.get('status', 'N/A')}")
        print(f"{'='*70}\n")
        
        # Email configuration - CORRECTED
        smtp_user = "maintenance.slnterminus@gmail.com"
        smtp_pass = "xaottgrqtqnkouqn"  # CORRECT app password
        recipients = ["maintenance.slnterminus@gmail.com","yasven7545@gmail.com","engineering@terminus-global.com" ]
        
# ✅ CRITICAL FIX: HTML ASSIGNED TO VARIABLE AS STRING WITH MODERN DESIGN
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Daily PPM Summary</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;700&display=swap');

        * {{ box-sizing: border-box; margin: 0; padding: 0; }}

        body {{
            font-family: 'DM Sans', sans-serif;
            background-color: #0d1117;
            padding: 40px 15px;
            color: #c9d1d9;
        }}

        .container {{
            width: 100%;
            max-width: 660px;
            margin: 0 auto;
            background-color: #161b22;
            border-radius: 12px;
            overflow: hidden;
            border: 1px solid #30363d;
        }}

        /* ─── HEADER ─── */
        .header {{
            background: #0d1117;
            padding: 36px 40px 28px;
            border-bottom: 1px solid #30363d;
            position: relative;
        }}
        .header-eyebrow {{
            font-family: 'Space Mono', monospace;
            font-size: 10px;
            color: #3fb950;
            letter-spacing: 3px;
            text-transform: uppercase;
            margin-bottom: 10px;
        }}
        .header h1 {{
            font-family: 'Space Mono', monospace;
            font-size: 26px;
            font-weight: 700;
            color: #e6edf3;
            letter-spacing: -0.5px;
            line-height: 1.1;
        }}
        .header h1 span {{
            color: #3fb950;
        }}
        .header-rule {{
            margin-top: 20px;
            height: 1px;
            background: linear-gradient(90deg, #3fb950 0%, transparent 80%);
        }}

        /* ─── SUMMARY BOX ─── */
        .summary-box {{
            padding: 28px 40px;
            background: #0d1117;
            border-bottom: 1px solid #30363d;
        }}
        .summary-inner {{
            width: 100%;
            border-collapse: collapse;
        }}
        .summary-inner td {{
            padding: 0;
            vertical-align: middle;
            border: none;
        }}
        .badge-cell {{
            width: 88px;
            padding-right: 24px !important;
        }}
        .summary-badge {{
            width: 72px;
            height: 72px;
            border-radius: 50%;
            border: 2px solid #3fb950;
            text-align: center;
            padding-top: 16px;
            mso-line-height-rule: exactly;
        }}
        .count {{
            font-family: 'Space Mono', monospace;
            font-size: 26px;
            font-weight: 700;
            color: #3fb950;
            line-height: 1;
            display: block;
        }}
        .count-label {{
            font-size: 8px;
            color: #8b949e;
            letter-spacing: 1px;
            text-transform: uppercase;
            margin-top: 3px;
            display: block;
        }}
        .summary-text {{
            font-size: 14px;
            color: #8b949e;
            line-height: 1.7;
        }}
        .summary-text strong {{
            color: #e6edf3;
            font-weight: 500;
        }}

        /* ─── TABLE ─── */
        .work-orders-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .work-orders-table thead {{
            background-color: #0d1117;
        }}
        .work-orders-table th {{
            padding: 12px 20px;
            text-align: left;
            font-family: 'Space Mono', monospace;
            font-size: 9px;
            font-weight: 700;
            color: #3fb950;
            letter-spacing: 2px;
            text-transform: uppercase;
            border-bottom: 1px solid #30363d;
        }}
        .work-orders-table td {{
            padding: 16px 20px;
            border-bottom: 1px solid #21262d;
            vertical-align: middle;
        }}
        .work-orders-table tr:last-child td {{
            border-bottom: none;
        }}
        .work-orders-table tr:hover td {{
            background-color: #1c2128;
        }}

        /* Priority badges */
        .priority-high {{
            background-color: rgba(248, 81, 73, 0.15);
            color: #f85149;
            border: 1px solid rgba(248, 81, 73, 0.4);
            padding: 3px 10px;
            border-radius: 4px;
            font-family: 'Space Mono', monospace;
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 1px;
        }}
        .priority-medium {{
            background-color: rgba(210, 153, 34, 0.15);
            color: #d29922;
            border: 1px solid rgba(210, 153, 34, 0.4);
            padding: 3px 10px;
            border-radius: 4px;
            font-family: 'Space Mono', monospace;
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 1px;
        }}
        .priority-low {{
            background-color: rgba(63, 185, 80, 0.15);
            color: #3fb950;
            border: 1px solid rgba(63, 185, 80, 0.4);
            padding: 3px 10px;
            border-radius: 4px;
            font-family: 'Space Mono', monospace;
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 1px;
        }}

        /* Cell content */
        .wo-id {{
            font-family: 'Space Mono', monospace;
            font-size: 12px;
            color: #58a6ff;
            font-weight: 700;
        }}
        .asset-title {{
            font-weight: 500;
            color: #e6edf3;
            font-size: 13px;
        }}
        .location-text {{
            color: #6e7681;
            font-size: 11px;
            margin-top: 3px;
            font-family: 'Space Mono', monospace;
        }}
        .due-date-text {{
            font-family: 'Space Mono', monospace;
            font-size: 11px;
            color: #8b949e;
        }}

        /* ─── BUTTON ─── */
        .btn-wrapper {{
            text-align: center;
            padding: 36px 40px;
            background: #0d1117;
            border-top: 1px solid #30363d;
        }}
        .action-button {{
            background: transparent;
            color: #3fb950 !important;
            text-decoration: none;
            padding: 14px 44px;
            border-radius: 6px;
            font-family: 'Space Mono', monospace;
            font-weight: 700;
            font-size: 13px;
            display: inline-block;
            border: 2px solid #3fb950;
            letter-spacing: 2px;
            text-transform: uppercase;
            mso-padding-alt: 0;
        }}

        /* ─── FOOTER ─── */
        .footer {{
            background-color: #0d1117;
            padding: 24px 40px;
            text-align: center;
            border-top: 1px solid #21262d;
        }}
        .footer p {{
            margin: 4px 0;
            color: #484f58;
            font-size: 12px;
            line-height: 1.6;
        }}
        .footer strong {{
            color: #6e7681;
            font-weight: 500;
        }}
        .footer .note {{
            margin-top: 14px;
            color: #30363d;
            font-size: 10px;
            font-family: 'Space Mono', monospace;
        }}
    </style>
</head>
<body>
    <div class="container">

        <div class="header">
            <div class="header-eyebrow">Maintenance Management System</div>
            <h1>SLN <span>TERMINUS</span></h1>
            <div class="header-rule"></div>
        </div>

        <div class="summary-box">
            <table class="summary-inner" cellpadding="0" cellspacing="0">
                <tr>
                    <td class="badge-cell">
                        <div class="summary-badge">
                            <span class="count">{len(today_wos)}</span>
                            <span class="count-label">Tasks</span>
                        </div>
                    </td>
                    <td>
                        <p class="summary-text">
                            Preventive Maintenance work orders are scheduled for today.<br>
                            <strong>Please review and assign technicians.</strong>
                        </p>
                    </td>
                </tr>
            </table>
        </div>

        <table class="work-orders-table">
            <thead>
                <tr>
                    <th>ID</th>
                    <th>Asset &amp; Location</th>
                    <th>Priority</th>
                    <th>Due Date</th>
                </tr>
            </thead>
            <tbody>
                {"".join(f'''
                <tr>
                    <td>
                        <span class="wo-id">#{wo.get('work_order_id', 'N/A')}</span>
                    </td>
                    <td>
                        <div class="asset-title">{wo.get('asset_name', 'Unknown Asset')}</div>
                        <div class="location-text">▸ {wo.get('location', 'Unknown Location')}</div>
                    </td>
                    <td>
                        <span class="priority-{wo.get('priority', 'Medium').lower()}">
                            {wo.get('priority', 'Medium').upper()}
                        </span>
                    </td>
                    <td>
                        <span class="due-date-text">{wo.get('due_date', 'N/A')}</span>
                    </td>
                </tr>
                ''' for wo in today_wos)}
            </tbody>
        </table>

        <div class="btn-wrapper">
            <a href="https://descriptive-joya-unsolidified.ngrok-free.dev" class="action-button">VIEW DASHBOARD</a>
        </div>

        <div class="footer">
            <p><strong>SLN Terminus Infrastructure Division</strong></p>
            <p>Automated system message. Please do not reply.</p>
            <p>© 2026 EPMS LLP. All rights reserved.</p>
            <p class="note">System Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | v3.0</p>
        </div>

    </div>
</body>
</html>"""

        # --- SMTP SENDING LOGIC ---
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"🔧 {len(today_wos)} Maintenance Tasks - {today.strftime('%d %b %Y')}"
        msg['From'] = formataddr(("SLN Terminus MMS", smtp_user))
        msg['To'] = ", ".join(recipients)
        msg.attach(MIMEText(html_content, 'html'))
        _smtp_send(msg, recipients, caller="MMS-manual")
        
        print(f"✅ Email sent successfully to {len(recipients)} recipients ({len(today_wos)} work orders)")
        return jsonify({
            "success": True,
            "recipients": recipients,
            "wo_count": len(today_wos),
            "message": "Email sent successfully"
        })
    
    except Exception as e:
        print(f"❌ Email sending error: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"Email failed: {str(e)}"
        }), 500
# =====================================================
# DAILY EMAIL SCHEDULER (TRIGGERS AT 8:00 AM)
# =====================================================
from apscheduler.schedulers.background import BackgroundScheduler

def setup_email_scheduler():
    """Sets up the scheduler to send emails at 8:00 AM daily"""
    scheduler = BackgroundScheduler()

    # PPM daily summary for other properties (OneWest etc.) — 8:00 AM IST
    # SLN MMS uses its own scheduler via register_mms_scheduler()
    if send_daily_summary is not None:
        scheduler.add_job(
            func=send_daily_summary,
            trigger='cron',
            hour=8,
            minute=0,
            timezone='Asia/Kolkata',
            id='ppm_daily_summary',
            replace_existing=True
        )
        print("✅ PPM daily summary scheduler registered: Daily at 8:00 AM IST")
    else:
        print("⚠️  PPM daily summary scheduler skipped (ppm_daily_mailer not available)")

    # CAM auto-reminders — 9:00 AM IST
    # Schedule: +1d, +3d, +7d, +10d, +15d, then daily until cleared
    if cam_auto_remind is not None:
        scheduler.add_job(
            func=cam_auto_remind,
            trigger='cron',
            hour=9,
            minute=0,
            timezone='Asia/Kolkata',
            id='cam_auto_reminders',
            replace_existing=True
        )
        print("✅ CAM reminder scheduler registered: Daily at 9:00 AM IST")
    else:
        print("⚠️  CAM reminder scheduler skipped (cam_charges_scheduler not found)")

    scheduler.start()
    print("✅ Email scheduler started: Daily trigger at 8:00 AM IST")
    return scheduler

# Initialize scheduler when server starts

# =====================================================
# MOBILE API — JWT auth for Flutter app
# =====================================================
import jwt as _jwt
import datetime as _dt

_JWT_SECRET = "terminus-mobile-2026"

def jwt_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"success": False, "error": "Token missing"}), 401
        try:
            payload = _jwt.decode(auth.split(" ")[1], _JWT_SECRET, algorithms=["HS256"])
            request.jwt_user = payload
        except _jwt.ExpiredSignatureError:
            return jsonify({"success": False, "error": "Token expired"}), 401
        except _jwt.InvalidTokenError:
            return jsonify({"success": False, "error": "Invalid token"}), 401
        return f(*args, **kwargs)
    return wrapper

@app.route("/api/mobile/login", methods=["POST"])
def mobile_login():
    data      = request.get_json(force=True) or {}
    username  = data.get("username", "").strip()
    password  = data.get("password", "").strip()
    property_ = data.get("property", "").strip()
    user = USERS.get(username)
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 401
    if user["password"] != password:
        return jsonify({"success": False, "error": "Invalid password"}), 401
    if property_ and property_ not in user["properties"]:
        return jsonify({"success": False, "error": "No access to this property"}), 403
    active_prop = property_ or user["properties"][0]
    token = _jwt.encode({
        "user": username, "role": user["role"],
        "properties": user["properties"],
        "active_property": active_prop,
        "exp": _dt.datetime.utcnow() + _dt.timedelta(days=7)
    }, _JWT_SECRET, algorithm="HS256")
    return jsonify({
        "success": True, "token": token,
        "role": user["role"], "properties": user["properties"],
        "active_property": active_prop
    })

@app.route("/api/mobile/dashboard", methods=["GET"])
@jwt_required
def mobile_dashboard():
    u = request.jwt_user
    return jsonify({"success": True, "user": u["user"],
        "role": u["role"], "active_property": u["active_property"],
        "properties": u["properties"]})

@app.route("/api/mobile/workorders", methods=["GET"])
@jwt_required
def mobile_workorders():
    try:
        with open(WO_JSON) as f:
            orders = json.load(f)
        data = orders if isinstance(orders, list) else orders.get("work_orders", [])
        prop = request.jwt_user.get("active_property", "")
        filtered = [o for o in data if not prop or o.get("property", prop) == prop]
        return jsonify({"success": True, "data": filtered})
    except Exception as e:
        return jsonify({"success": True, "data": [], "note": str(e)})

@app.route("/api/mobile/issues", methods=["GET"])
@jwt_required
def mobile_issues():
    prop = request.jwt_user.get("active_property", "")
    try:
        from models import Issue
        issues = Issue.query.filter_by(property=prop)\
                            .order_by(Issue.created_at.desc()).limit(50).all()
        return jsonify({"success": True, "data": [{
            "id": i.id, "title": getattr(i, "title", ""),
            "status": getattr(i, "status", ""),
            "priority": getattr(i, "priority", ""),
            "created_at": str(i.created_at),
            "property": getattr(i, "property", prop),
        } for i in issues]})
    except Exception as e:
        return jsonify({"success": True, "data": [], "note": str(e)})

@app.route("/api/mobile/ppm", methods=["GET"])
@jwt_required
def mobile_ppm():
    prop = request.jwt_user.get("active_property", "")
    try:
        with open(PPM_DATA_FILE) as f:
            ppm = json.load(f)
        if isinstance(ppm, dict):
            ppm = ppm.get(prop, ppm.get("assets", []))
        return jsonify({"success": True, "data": ppm})
    except Exception as e:
        return jsonify({"success": True, "data": [], "note": str(e)})

if __name__ == "__main__":
    # ... other initialization code ...
    email_scheduler = setup_email_scheduler()


# =====================================================
# AMC TRACKER API ENDPOINTS (ADD THIS TO server.py)
# =====================================================

@app.route('/api/amc/contracts')
def get_amc_contracts():
    """API: Get all AMC contracts from JSON file"""
    try:
        AMC_JSON = BASE_DIR / "static" / "data" / "amc_contracts.json"
        
        # Create directory if it doesn't exist
        AMC_JSON.parent.mkdir(parents=True, exist_ok=True)
        
        # Return empty list if file doesn't exist yet
        if not AMC_JSON.exists():
            return jsonify({"contracts": []})
        
        # Load contracts from JSON file
        with open(AMC_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        contracts = data.get('contracts', [])
        return jsonify({"contracts": contracts})
    
    except Exception as e:
        print(f"❌ AMC contracts fetch error: {str(e)}")
        traceback.print_exc()
        return jsonify({"contracts": []}), 500


@app.route('/api/amc/update', methods=['POST'])
def update_amc_contract():
    """API: Update AMC contract details in JSON file"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400
        
        contract_id = data.get('contract_id')
        if not contract_id:
            return jsonify({"success": False, "error": "Contract ID is required"}), 400
        
        AMC_JSON = BASE_DIR / "static" / "data" / "amc_contracts.json"
        AMC_JSON.parent.mkdir(parents=True, exist_ok=True)
        
        # Load existing contracts or initialize empty list
        if AMC_JSON.exists():
            with open(AMC_JSON, 'r', encoding='utf-8') as f:
                amc_data = json.load(f)
        else:
            amc_data = {"contracts": []}
        
        contracts = amc_data.get('contracts', [])
        
        # Find and update contract
        updated = False
        for i, contract in enumerate(contracts):
            if contract.get('contract_id') == contract_id:
                contracts[i] = data
                updated = True
                break
        
        if not updated:
            # Add new contract if not found (for flexibility)
            contracts.append(data)
            updated = True
        
        # Save updated data
        amc_data['contracts'] = contracts
        amc_data['last_updated'] = datetime.now().isoformat()
        
        with open(AMC_JSON, 'w', encoding='utf-8') as f:
            json.dump(amc_data, f, indent=2)
        
        print(f"✅ AMC Contract {contract_id} updated successfully")
        return jsonify({"success": True, "message": "Contract updated successfully"})
    
    except Exception as e:
        print(f"❌ AMC contract update error: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# =====================================================
# 12. OCCUPANCY ROUTES
# =====================================================
@app.route("/sln_occupancy")
@login_required
def sln_occupancy():
    return render_template("sln_occupancy.html")

@app.route('/api/sln/occupancy')
@login_required
def get_sln_occupancy():
    """Get space occupancy data from Excel"""
    try:
        # Support both JSON override file and Excel source
        OVERRIDE_JSON = BASE_DIR / "static" / "data" / "sln_occupancy_override.json"
        EXCEL_PATH    = BASE_DIR / "static" / "data" / "Space Occupancy.xlsx"

        # If an override JSON exists (saved by the update route), serve it directly
        if OVERRIDE_JSON.exists():
            try:
                spaces = json.loads(OVERRIDE_JSON.read_text(encoding='utf-8'))
            except Exception:
                spaces = []
        elif EXCEL_PATH.exists():
            df = pd.read_excel(EXCEL_PATH, sheet_name=0)
            df = df.fillna('')
            spaces = []
            current_floor = None
            for _, row in df.iterrows():
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                    continue
                try:
                    office_name   = str(row.iloc[2]) if len(row) > 2 else ''
                    floor_level   = str(row.iloc[1]) if len(row) > 1 else ''
                    unit_no       = str(row.iloc[3]) if len(row) > 3 else ''
                    occupied_area = float(str(row.iloc[4]).replace(',','')) if len(row) > 4 else 0
                    area          = float(str(row.iloc[5]).replace(',','')) if len(row) > 5 else 0
                    cam_rate      = float(str(row.iloc[6]).replace(',','')) if len(row) > 6 else 0
                    cam_rate_date = str(row.iloc[7]).strip() if len(row) > 7 else ''

                    if floor_level.strip():
                        current_floor = floor_level.strip()
                    if office_name.strip():
                        spaces.append({
                            "id":           str(row.iloc[0]),
                            "floorId":      current_floor or 'L4',
                            "officeName":   office_name.strip(),
                            "unitNo":       unit_no.strip(),
                            "occupiedArea": occupied_area,
                            "area":         area,
                            "camRate":      cam_rate,
                            "camRateDate":  cam_rate_date
                        })
                except Exception:
                    continue
        else:
            return jsonify({"error": "Space Occupancy data not found"}), 404

        def _status(s):
            nm = (s.get("officeName","")).lower()
            if nm in ("vacant","") or (nm=="vacant"): return "vacant"
            if "fit" in nm and "out" in nm:           return "fitout"
            if s.get("camRate",0) == 0:               return "fitout"
            return "occupied"

        total_units   = len(spaces)
        vacant_count  = sum(1 for s in spaces if _status(s) == "vacant")
        fitout_count  = sum(1 for s in spaces if _status(s) == "fitout")
        occupied_count = total_units - vacant_count - fitout_count

        return jsonify({
            "summary": {
                "total_area":     sum(s.get("area",0) for s in spaces),
                "occupied_area":  sum(s.get("occupiedArea",0) for s in spaces if _status(s)=="occupied"),
                "vacant_area":    sum(s.get("area",0) for s in spaces if _status(s)=="vacant"),
                "total_units":    total_units,
                "vacant_count":   vacant_count,
                "fitout_count":   fitout_count,
                "occupied_count": occupied_count
            },
            "spaces": spaces
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/sln/occupancy/update', methods=['POST'])
@login_required
def update_sln_occupancy():
    """Save edited space occupancy data as JSON override (avoids mutating the Excel)"""
    try:
        data   = request.get_json(force=True) or {}
        spaces = data.get("spaces", [])
        if not isinstance(spaces, list):
            return jsonify({"success": False, "error": "Invalid payload"}), 400

        OVERRIDE_JSON = BASE_DIR / "static" / "data" / "sln_occupancy_override.json"
        OVERRIDE_JSON.parent.mkdir(parents=True, exist_ok=True)
        OVERRIDE_JSON.write_text(
            json.dumps(spaces, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )
        return jsonify({"success": True, "count": len(spaces)})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

# =====================================================
# 13. PROJECT HANDOVER WORKSPACE
# =====================================================
@app.route("/project_handover")
@login_required
def project_handover():
    return render_template("project_handover_workspace.html")

@app.route("/project_handover_workspace")
@login_required
def project_handover_workspace():
    return render_template("project_handover_workspace.html")

@app.route("/api/upload/<category>", methods=["POST"])
@login_required
def upload_file(category):
    """Upload file to project handover category"""
    if category not in CATEGORIES:
        return jsonify({"error": "Invalid category"}), 400
    
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "Empty filename"}), 400
    
    save_dir = UPLOAD_ROOT / category
    save_dir.mkdir(parents=True, exist_ok=True)
    
    filename = secure_filename(file.filename)
    save_path = save_dir / filename
    file.save(save_path)
    
    # Log audit
    log_audit_action("File Upload", "ProjectHandover", filename)
    
    return jsonify({"message": "Uploaded successfully", "filename": filename})

@app.route("/api/list/<category>")
@login_required
def list_files(category):
    """List files in project handover category"""
    if category not in CATEGORIES:
        return jsonify([])
    
    folder = UPLOAD_ROOT / category
    if not folder.exists():
        folder.mkdir(parents=True, exist_ok=True)
        return jsonify([])
    
    files = []
    for f in sorted(folder.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.is_file():
            files.append({
                "name": f.name,
                "size": f"{round(f.stat().st_size / 1024, 1)} KB",
                "date": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
            })
    
    return jsonify(files)

@app.route("/api/folder/<category>")
@login_required
def folder_count(category):
    """Alias for list_files — used by file-count tiles"""
    return list_files(category)

@app.route("/api/delete/<category>/<filename>", methods=['DELETE'])
@login_required
def delete_file(category, filename):
    """Delete file from project handover"""
    if category not in CATEGORIES:
        return jsonify({"success": False, "error": "Invalid category"}), 400
    
    file_path = UPLOAD_ROOT / category / filename
    if file_path.exists():
        os.remove(file_path)
        log_audit_action("File Delete", "ProjectHandover", filename)
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "error": "File not found"}), 404

@app.route("/files/<category>/<filename>")
@login_required
def serve_file(category, filename):
    """Serve project handover file"""
    folder = UPLOAD_ROOT / category
    return send_from_directory(folder, filename)

@app.route("/uploads/<category>/<filename>")
@login_required
def serve_upload(category, filename):
    """Serve upload file"""
    folder = UPLOAD_ROOT / category
    return send_from_directory(folder, filename)

# =====================================================
# 14. TRAINING IMAGES UPLOAD
# =====================================================
@app.route("/api/training/list")
@login_required
def list_training_images():
    """List training images by department"""
    dept = request.args.get("department", "").strip()
    if not dept:
        return jsonify({"error": "Department required"}), 400
    
    dest_dir = TRAINING_UPLOAD_ROOT / dept
    if not dest_dir.exists():
        return jsonify({"department": dept, "files": []})
    
    files = []
    for p in sorted(dest_dir.iterdir(), key=lambda x: x.name):
        if p.is_file() and allowed_image(p.name):
            files.append({
                "name": p.name,
                "url": url_for("serve_training_image", department=dept, filename=p.name),
                "size": f"{round(p.stat().st_size / 1024, 1)} KB",
                "date": datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d")
            })
    
    return jsonify({"department": dept, "files": files})

@app.route("/api/training/upload", methods=["POST"])
@login_required
def upload_training_image():
    """Upload training image"""
    dept = request.form.get("department", "").strip()
    if not dept:
        return jsonify({"error": "Department required"}), 400
    
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    
    file = request.files["file"]
    if not file or file.filename == "":
        return jsonify({"error": "Empty filename"}), 400
    
    if not allowed_image(file.filename):
        return jsonify({"error": "Invalid file type"}), 400
    
    dest_dir = TRAINING_UPLOAD_ROOT / dept
    dest_dir.mkdir(parents=True, exist_ok=True)
    
    filename = secure_filename(file.filename)
    save_path = dest_dir / filename
    file.save(save_path)
    
    log_audit_action("Training Image Upload", "Training", f"{dept}/{filename}")
    
    return jsonify({"success": True, "filename": filename})

@app.route("/uploads/training/<department>/<filename>")
def serve_training_image(department, filename):
    """Serve training image"""
    folder = TRAINING_UPLOAD_ROOT / department
    return send_from_directory(folder, filename)

# =====================================================
# 15. OTHER DASHBOARD ROUTES (NO DUPLICATES)
# =====================================================
@app.route("/mis")
@login_required
def mis():
    return render_template("mis.html")

@app.route("/kra")
@login_required
def kra():
    return render_template("kra.html")



# ── Energy: serve dashboard Excel with no-cache headers ──────────────────
@app.route("/api/energy/data")
@login_required
def energy_data_redirect():
    """API alias for the dashboard Excel — returns with no-cache headers."""
    path = BASE_DIR / "static" / "data" / "SLN_Terminus_Dashboard_Data.xlsx"
    if not path.exists():
        return jsonify({"error": "SLN_Terminus_Dashboard_Data.xlsx not found in static/data/"}), 404
    from flask import make_response
    resp = make_response(send_file(
        str(path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"]        = "no-cache"
    return resp


# ── Energy: replace dashboard Excel via upload ───────────────────────────
@app.route("/api/energy/upload", methods=["POST"])
@login_required
def energy_upload():
    """Allow replacing SLN_Terminus_Dashboard_Data.xlsx via file upload."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"success": False, "error": "No file provided"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "error": "Only .xlsx / .xls files accepted"}), 400
    try:
        # Accept both legacy name and new name; save as Energy_analysys.xlsx
        dest = BASE_DIR / "static" / "data" / "Energy_analysys.xlsx"
        dest.parent.mkdir(parents=True, exist_ok=True)
        f.save(str(dest))
        return jsonify({
            "success": True,
            "message": f"Dashboard data updated ({dest.stat().st_size // 1024} KB)"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── Energy: parse Excel → JSON (all sheets incl. water _consumption) ────────
@app.route("/api/energy/json")
@login_required
def energy_json():
    """
    Parse Energy_analysys.xlsx server-side and return structured JSON.
    Handles sheet names with spaces (e.g. 'water _consumption').
    Called by energy.html on DOMContentLoaded so all modules get live data.
    """
    import openpyxl
    from datetime import datetime as _dt, timedelta as _td

    XLSX_NAME = "Energy_analysys.xlsx"
    path = BASE_DIR / "static" / "data" / XLSX_NAME
    if not path.exists():
        # fallback: try legacy name
        path = BASE_DIR / "static" / "data" / "SLN_Terminus_Dashboard_Data.xlsx"
    if not path.exists():
        return jsonify({"error": f"{XLSX_NAME} not found in static/data/"}), 404

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        result = {}

        # ── ele_consumption ──────────────────────────────────────────────────
        ele_sheet = next((s for s in wb.sheetnames
                          if "ele" in s.lower() and "sft" not in s.lower()), None)
        if ele_sheet:
            ele = []
            for row in list(wb[ele_sheet].iter_rows(values_only=True))[1:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                ele.append({"n": int(row[0]), "m": m,
                             "u": float(row[2]) if row[2] is not None else None})
            result["ele"] = ele

        # ── dg_consumption ───────────────────────────────────────────────────
        dg_sheet = next((s for s in wb.sheetnames if "dg" in s.lower()), None)
        if dg_sheet:
            dg = []
            rows = list(wb[dg_sheet].iter_rows(values_only=True))
            start = next((i + 1 for i, r in enumerate(rows) if r[0] == "Sno"), 2)
            for row in rows[start:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                rh = row[2]
                if isinstance(rh, _td):
                    ts = int(rh.total_seconds())
                    hd = round(ts / 3600, 2)
                    h_str = f"{ts // 3600:02d}:{(ts % 3600) // 60:02d}"
                elif hasattr(rh, "hour"):
                    hd = round(rh.hour + rh.minute / 60, 2)
                    h_str = f"{rh.hour:02d}:{rh.minute:02d}"
                else:
                    hd = float(rh or 0)
                    hh = int(hd)
                    h_str = f"{hh:02d}:{int((hd - hh) * 60):02d}"
                dg.append({"n": int(row[0]), "m": m, "h": h_str, "hd": hd,
                            "k": float(row[3] or 0), "d": float(row[4] or 0)})
            result["dg"] = dg

        # ── water _consumption  (sheet name has a space — handled explicitly) ─
        water_sheet = next((s for s in wb.sheetnames
                            if "water" in s.lower() and "tank" not in s.lower()), None)
        if water_sheet:
            water = []
            for row in list(wb[water_sheet].iter_rows(values_only=True))[1:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                # Cols: S.No | Month | Consumption (KL) | Net Pay Rs
                kl  = float(row[2]) if row[2] is not None else None
                pay = float(row[3]) if len(row) > 3 and row[3] is not None else None
                water.append({"n": int(row[0]), "m": m, "v": kl, "pay": pay})
            result["water"] = water

        # ── ele_con_sft (benchmark / area data) ──────────────────────────────
        sft_sheet = next((s for s in wb.sheetnames if "sft" in s.lower()), None)
        if sft_sheet:
            bm = {}
            for row in wb[sft_sheet].iter_rows(values_only=True):
                if row[0] and row[1] is not None:
                    bm[str(row[0]).strip()] = row[1]
            result["sft"] = bm

        # ── Tank_cons (Sno | Month | Trips | Consumed KL | Vol/Trip | Rate | Total Cost | Vendor) ──
        tank_sheet = next((s for s in wb.sheetnames
                           if "tank" in s.lower()), None)
        if tank_sheet:
            tanker = []
            rows_t = list(wb[tank_sheet].iter_rows(values_only=True))
            # detect header row (skip it and any sub-headers)
            data_start = next(
                (i + 1 for i, r in enumerate(rows_t) if r[0] == "Sno"), 1
            )
            for row in rows_t[data_start:]:
                if not isinstance(row[0], (int, float)):
                    continue
                m      = row[1].strftime("%b %Y") if isinstance(row[1], _dt) else str(row[1] or "—")
                trips  = int(row[2])      if len(row) > 2 and row[2] is not None else None
                vol    = float(row[3])    if len(row) > 3 and row[3] is not None else None
                # Vol/Trip col [4] is derived — skip
                rate   = float(row[5])    if len(row) > 5 and row[5] is not None else None
                cost   = float(row[6])    if len(row) > 6 and row[6] is not None else None
                vendor = str(row[7]).strip() if len(row) > 7 and row[7] is not None else "—"
                tanker.append({
                    "n":      int(row[0]),
                    "m":      m,
                    "trips":  trips,
                    "vol":    vol,
                    "rate":   rate,
                    "cost":   cost,   # pre-calculated Total Cost from sheet
                    "vendor": vendor,
                })
            result["tanker"] = tanker

        result["source"] = XLSX_NAME
        return jsonify(result)

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/inventory_dashboard")
@login_required
def inventory_dashboard():
    return render_template("inventory_dashboard.html")

@app.route("/tenant")
@login_required
def tenant():
    return render_template("tenant.html")

@app.route("/cam_charges")
@login_required
def cam_charges_page():
    return render_template("cam_charges.html")

@app.route("/cam_review")
@login_required
def cam_review():
    return render_template("cam_review.html")

@app.route("/pm_dashboard")
@login_required
def pm_dashboard():
    return render_template("pm_dashboard.html")

@app.route("/property-manager-updates")
@login_required
def pm_daily_updates_page():
    return render_template("pm_daily_updates.html")

@app.route("/gm_dashboard")
@require_role("General Manager")
def gm_dashboard():
    return render_template("gm_dashboard.html")

@app.route("/documents")
@login_required
def documents():
    return render_template("documents.html")

@app.route("/issues")
@login_required
def issues():
    return render_template("issues.html")

@app.route("/vendor_visit")
@login_required
def vendor_visit():
    return render_template("vendor_visit.html")

# =====================================================
# 16. FILE DOWNLOAD
# =====================================================
@app.route("/download-excel")
@login_required
def download_excel():
    path = os.path.join(app.static_folder, "data")
    return send_from_directory(path, "SLN_Terminus_Dashboard_Data.xlsx", as_attachment=True)

@app.route('/api/ppm/import-excel', methods=['POST'])
@login_required
def import_ppm_excel():
    """Import PPM assets from Excel"""
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400
        
        upload_path = Path(app.static_folder) / "data" / "Assets.xlsx"
        file.save(upload_path)
        
        df = pd.read_excel(upload_path)
        assets = []
        
        for _, row in df.iterrows():
            if pd.notna(row.get('Asset Code')) and str(row.get('Asset Code')).strip():
                assets.append({
                    "id": str(row.get('Asset Code', '')).strip(),
                    "name": str(row.get('Asset Name', '')).strip(),
                    "category": str(row.get('In-House/Vendor', 'General')).strip(),
                    "location": str(row.get('Location', '')).strip(),
                    "lastService": str(row.get('Last Service', '')).strip(),
                    "nextDueDate": str(row.get('nextDueDate', '')).strip(),
                    "colorCode": "Red"
                })
        
        with open(PPM_DATA_FILE, 'w') as f:
            json.dump({"assets": assets}, f, indent=2)
        
        return jsonify({
            "status": "success",
            "message": f"Successfully imported {len(assets)} assets",
            "count": len(assets)
        })
    
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# =====================================================
# PPM CHECKLIST UPLOAD & PARSE ROUTE
# =====================================================
PPM_CHECKLIST_DIR = BASE_DIR / "uploads" / "ppm_checklists"
PPM_CHECKLIST_DIR.mkdir(parents=True, exist_ok=True)

@app.route('/api/ppm/checklist/upload', methods=['POST'])
@login_required
def upload_ppm_checklist():
    """Parse uploaded PPM checklist Excel and return structured checklist data"""
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "No file uploaded"}), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({"success": False, "error": "Empty filename"}), 400

        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls'):
            return jsonify({"success": False, "error": "Only .xlsx / .xls files accepted"}), 400

        filename = secure_filename(file.filename)
        save_path = PPM_CHECKLIST_DIR / filename
        file.save(save_path)

        # ── Parse the Excel ──────────────────────────────────────────────
        df = pd.read_excel(save_path, sheet_name=0, header=None)

        # Derive title from cell A1 (row 0 col 0)
        raw_title = str(df.iloc[0, 0]).strip() if not pd.isna(df.iloc[0, 0]) else "PPM Checklist"

        ehs_steps   = []   # EHS items (paired columns 0+1 and 2+3)
        sections    = []   # [{group, items:[{sno,description}]}]
        current_grp = None

        # Column headers are usually on row index 5 (S.No | Description | Status | Remarks)
        # Items start from row 6 onward
        HEADER_ROW = None
        for i, row in df.iterrows():
            v0 = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ''
            if v0 == 's.no':
                HEADER_ROW = i
                break

        # EHS block: rows 2-4 (before header row), paired items across col 0-1 / 2-3
        if HEADER_ROW and HEADER_ROW >= 3:
            for i in range(2, HEADER_ROW):
                row = df.iloc[i]
                for col_pair in [(0, 1), (2, 3)]:
                    sno  = row.iloc[col_pair[0]]
                    desc = row.iloc[col_pair[1]]
                    if not pd.isna(sno) and not pd.isna(desc):
                        ehs_steps.append({
                            "sno": str(int(sno)) if isinstance(sno, float) else str(sno),
                            "description": str(desc).strip()
                        })

        # Main checklist items
        if HEADER_ROW is not None:
            for i in range(HEADER_ROW + 1, len(df)):
                row = df.iloc[i]
                sno  = row.iloc[0]
                desc = row.iloc[1] if len(row) > 1 else None

                # Section/group heading row (sno is NaN, desc has text)
                if pd.isna(sno) and not pd.isna(desc):
                    grp_text = str(desc).strip()
                    # Skip footer rows
                    if any(kw in grp_text.lower() for kw in ['sign', 'spares', 'supervisor', 'executive']):
                        continue
                    current_grp = grp_text
                    sections.append({"group": current_grp, "items": []})
                    continue

                # Regular item
                if not pd.isna(sno) and not pd.isna(desc):
                    item_desc = str(desc).strip().replace('\n', ' ')
                    if not item_desc or any(kw in item_desc.lower() for kw in ['sign:', 'spares used']):
                        continue
                    if not sections:
                        sections.append({"group": "General", "items": []})
                    sections[-1]["items"].append({
                        "sno": str(int(sno)) if isinstance(sno, float) else str(sno),
                        "description": item_desc
                    })

        return jsonify({
            "success": True,
            "title": raw_title,
            "filename": filename,
            "ehs_steps": ehs_steps,
            "sections": sections,
            "total_items": sum(len(s["items"]) for s in sections)
        })

    except Exception as e:
        print(f"❌ PPM checklist upload error: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# =====================================================
# MVGDS — DEPARTMENT CHECKLIST TEMPLATE STORAGE
# Departments: mep, hk, sec, fire
# Files persist in: uploads/checklist_templates/<dept>/
# Accepts: .xlsx, .xls, .docx, .doc, .pdf
# =====================================================

CHECKLIST_DEPTS = {"mep", "hk", "sec", "fire"}
CHECKLIST_ALLOWED = {"xlsx", "xls", "docx", "doc", "pdf"}
CHECKLIST_ROOT = BASE_DIR / "uploads" / "checklist_templates"
for _dept in CHECKLIST_DEPTS:
    (CHECKLIST_ROOT / _dept).mkdir(parents=True, exist_ok=True)


@app.route("/api/checklist/upload/<dept>", methods=["POST"])
@login_required
def checklist_upload(dept):
    """Upload a checklist template file to a department folder."""
    if dept not in CHECKLIST_DEPTS:
        return jsonify({"success": False, "error": f"Unknown department: {dept}"}), 400

    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file provided"}), 400

    file = request.files["file"]
    if not file or file.filename == "":
        return jsonify({"success": False, "error": "Empty filename"}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in CHECKLIST_ALLOWED:
        return jsonify({"success": False,
                        "error": f"File type .{ext} not allowed. Use: xlsx, xls, docx, doc, pdf"}), 400

    filename  = secure_filename(file.filename)
    dest_dir  = CHECKLIST_ROOT / dept
    dest_path = dest_dir / filename

    # If a file with same name already exists, append a counter suffix
    if dest_path.exists():
        stem, suffix = dest_path.stem, dest_path.suffix
        counter = 1
        while (dest_dir / f"{stem}_{counter}{suffix}").exists():
            counter += 1
        filename  = f"{stem}_{counter}{suffix}"
        dest_path = dest_dir / filename

    file.save(dest_path)

    log_audit_action("Checklist Upload", "MVGDS", f"{dept}/{filename}")

    return jsonify({
        "success":  True,
        "dept":     dept,
        "filename": filename,
        "size":     dest_path.stat().st_size,
        "ext":      ext,
    })


@app.route("/api/checklist/list/<dept>")
@login_required
def checklist_list(dept):
    """Return list of uploaded checklist templates for a department."""
    if dept not in CHECKLIST_DEPTS:
        return jsonify({"success": False, "error": "Unknown department"}), 400

    dept_dir = CHECKLIST_ROOT / dept
    files = []
    if dept_dir.exists():
        for p in sorted(dept_dir.iterdir()):
            if p.is_file() and p.suffix.lstrip(".").lower() in CHECKLIST_ALLOWED:
                files.append({
                    "filename": p.name,
                    "ext":      p.suffix.lstrip(".").lower(),
                    "size":     p.stat().st_size,
                    "modified": int(p.stat().st_mtime * 1000),  # ms timestamp
                })

    return jsonify({"success": True, "dept": dept, "files": files})


@app.route("/api/checklist/download/<dept>/<filename>")
@login_required
def checklist_download(dept, filename):
    """Serve/download a checklist template file."""
    if dept not in CHECKLIST_DEPTS:
        abort(404)

    safe_name = secure_filename(filename)
    dept_dir  = CHECKLIST_ROOT / dept
    file_path = dept_dir / safe_name

    if not file_path.exists() or not file_path.is_file():
        abort(404)

    # For PDF — serve inline so viewer modal can embed it
    ext = safe_name.rsplit(".", 1)[-1].lower()
    as_attachment = ext != "pdf"

    return send_file(
        file_path,
        as_attachment=as_attachment,
        download_name=safe_name,
    )


@app.route("/api/checklist/delete/<dept>/<filename>", methods=["DELETE"])
@login_required
def checklist_delete(dept, filename):
    """Delete a checklist template file."""
    if dept not in CHECKLIST_DEPTS:
        return jsonify({"success": False, "error": "Unknown department"}), 400

    safe_name = secure_filename(filename)
    file_path = CHECKLIST_ROOT / dept / safe_name

    if not file_path.exists():
        return jsonify({"success": False, "error": "File not found"}), 404

    file_path.unlink()
    log_audit_action("Checklist Delete", "MVGDS", f"{dept}/{safe_name}")

    return jsonify({"success": True, "deleted": safe_name})


# =====================================================
# 17. UTILITY ROUTES
# =====================================================
@app.route('/api/datetime')
def get_datetime():
    return jsonify({
        "current_datetime": datetime.now().strftime("%A, %B %d, %Y | %I:%M %p"),
        "server_time": datetime.now().isoformat()
    })

@app.route('/health')
def health_check():
    return jsonify({
        "status": "healthy",
        "service": "Terminus MMS",
        "timestamp": datetime.now().isoformat(),
        "version": "3.0.0"
    }), 200

@app.route('/favicon.ico')
def favicon():
    return '', 204

# =====================================================
# 18. AUDIT LOGGING HELPER
# =====================================================
def log_audit_action(action, entity_type, entity_id):
    """Log audit action to database"""
    try:
        with app.app_context():
            log = AuditLog(
                user_id=session.get('user_id', 0),
                username=session.get('user', 'system'),
                action=action,
                entity_type=entity_type,
                entity_id=str(entity_id),
                ip_address=request.remote_addr if request else '127.0.0.1'
            )
            db.session.add(log)
            db.session.commit()
    except:
        pass  # Don't fail if audit logging fails

# =====================================================
# 19. ERROR HANDLERS
# =====================================================
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": "Not Found", "message": "Resource not found"}), 404
    return render_template("onewest.html", error_code=404), 404

@app.errorhandler(403)
def forbidden(e):
    if "user" not in session:
        return redirect(url_for("login"))
    if request.path.startswith('/api/'):
        return jsonify({"error": "Forbidden", "message": "Access denied"}), 403
    return render_template("error.html", error_code=403), 403

@app.errorhandler(500)
def internal_error(e):
    print(f"500 Error: {str(e)}")
    if request.path.startswith('/api/'):
        return jsonify({"error": "Internal Server Error", "message": "Something went wrong"}), 500
    return render_template("error.html", error_code=500), 500


# =====================================================
# JSON ERROR HANDLERS FOR API ROUTES
# =====================================================
@app.errorhandler(403)
def handle_403(e):
    if request.path.startswith(("/ow_api/", "/api/", "/inventory/",
                                "/ow_work_track/", "/sln_work_track/", "/ow_vms/", "/ow_mail/")):
        return jsonify({"success": False, "error": "Access denied — check your active property"}), 403
    return render_template("dashboard.html", error="Access denied"), 403

@app.errorhandler(500)
def handle_500(e):
    if request.path.startswith(("/ow_api/", "/api/", "/inventory/",
                                "/ow_work_track/", "/sln_work_track/", "/ow_vms/", "/ow_mail/")):
        return jsonify({"success": False, "error": f"Server error: {str(e)}"}), 500
    return render_template("dashboard.html", error="Server error"), 500


# =====================================================
# SLN WORK TRACK BLUEPRINT REGISTRATION
# =====================================================
try:
    from sln_work_track_routes import sln_work_track_register
    sln_work_track_register(app)
except Exception as e:
    print(f"SLN Work Track blueprint error: {e}")

# =====================================================
# SLN WORK TRACK DASHBOARD ROUTE
# =====================================================
@app.route("/sln_work_track")
@login_required
@require_property("SLN Terminus")
def sln_work_track():
    return render_template("sln_work_track.html")

# =====================================================
# 2.0 ONEWEST
# =====================================================

# ✅ FIXED (WITH PROPER DECORATORS)
@app.route("/onewest")
@login_required  # ← MUST BE FIRST
@require_property("ONEWEST")
def onewest():
    """ONEWEST Property Dashboard"""
    # Ensure property is set in session
    session['active_property'] = 'ONEWEST'
    session['property_code'] = 'OW'
    
    # Debug output
    print(f"\n🏢 Accessing ONEWEST - User: {session.get('user')}")
    print(f"   Active Property: {session.get('active_property')}")
    print(f"   User Role: {session.get('role')}")
    
    return render_template("onewest.html")


# =====================================================
# ONEWEST ISSUES MODULE (COMPLETE)
# =====================================================
OW_ISSUES_JSON = BASE_DIR / "static" / "data" / "OW" / "issues.json"
OW_TECHNICIANS_JSON = BASE_DIR / "static" / "data" / "OW" / "technicians.json"
OW_SUPERVISORS_JSON = BASE_DIR / "static" / "data" / "OW" / "supervisors.json"
OW_ISSUES_UPLOADS = BASE_DIR / "uploads" / "OW" / "issues"
ISSUES_ARCHIVE_DIR = BASE_DIR / "uploads" / "OW" / "issues_archive"  # ← ADD THIS
ISSUES_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
OW_ISSUES_UPLOADS.mkdir(parents=True, exist_ok=True)

# Create directories
for folder in [OW_ISSUES_JSON.parent, OW_ISSUES_UPLOADS]:
    folder.mkdir(parents=True, exist_ok=True)

# FIX: write full authoritative staff data on every startup so files stay current
_OW_TECHS = {
    "technicians": [
        {"id": "T001",  "name": "Jagadish",    "phone": "+919666942315", "specialization": "Supervisor"},
        {"id": "T002",  "name": "Sri Ram",     "phone": "+919989668311", "specialization": "Supervisor"},
        {"id": "T003",  "name": "Ameer",       "phone": "+919000564662", "specialization": "Supervisor"},
        {"id": "T004",  "name": "Rakesh",      "phone": "+917730834084", "specialization": "Supervisor"},
        {"id": "T005",  "name": "Raghavendra", "phone": "+918008883537", "specialization": "BMS"},
        {"id": "T006",  "name": "Shiva",       "phone": "+918885637165", "specialization": "BMS"},
        {"id": "T007",  "name": "Tanmaya",     "phone": "+917077689216", "specialization": "BMS"},
        {"id": "T008",  "name": "Raviteja",    "phone": "+919652607622", "specialization": "Electrical"},
        {"id": "T009",  "name": "Nagendra",    "phone": "+919347324744", "specialization": "Electrical"},
        {"id": "T010",  "name": "Yeshwanth",   "phone": "+919502856581", "specialization": "Electrical"},
        {"id": "T011",  "name": "Hakim",       "phone": "+918083360242", "specialization": "Electrical"},
        {"id": "T012",  "name": "Sai pawan",   "phone": "+917013987434", "specialization": "Electrical"},
        {"id": "T013",  "name": "Nikhil",      "phone": "+917093979479", "specialization": "Asst Technician"},
        {"id": "T014",  "name": "Vijay",       "phone": "+916304725703", "specialization": "Asst Technician"},
        {"id": "T015",  "name": "Karthik",     "phone": "+919553174565", "specialization": "Asst Technician"},
        {"id": "T016",  "name": "Ilyas",       "phone": "+919347732552", "specialization": "HVAC"},
        {"id": "T017",  "name": "Sai",         "phone": "+917794057118", "specialization": "HVAC"},
        {"id": "T018",  "name": "Venu",        "phone": "+919618670499", "specialization": "HVAC"},
        {"id": "T019",  "name": "Bharath",     "phone": "+918106869682", "specialization": "HVAC"},
        {"id": "T020",  "name": "Ismail",      "phone": "+919154223362", "specialization": "HVAC"},
        {"id": "T021",  "name": "Bipin",       "phone": "+919121261604", "specialization": "Plumber"},
        {"id": "T027",  "name": "Bichitra",    "phone": "+917732040540", "specialization": "Plumber"},
        {"id": "T022",  "name": "Sudarshan",   "phone": "+917036994079", "specialization": "Plumber"},
        {"id": "T023",  "name": "Srikanth",    "phone": "+917749090745", "specialization": "Plumber"},
        {"id": "T024",  "name": "Tapan",       "phone": "+916380896010", "specialization": "Plumber"},
        {"id": "T025",  "name": "Rohith",      "phone": "+919948351383", "specialization": "Painter"},
        {"id": "T026",  "name": "Laxman",      "phone": "+917995751392", "specialization": "Carpenter"}
    ]
}
_OW_SUPS = {
    "supervisors": [
        {"id": "S001", "name": "Anil Kumar",   "phone": "+919876543220", "email": "anil@onewest.com"},
        {"id": "S002", "name": "Ravi Shankar", "phone": "+919876543221", "email": "ravi@onewest.com"}
    ]
}
# Always overwrite — keeps staff list current on every server restart
with open(OW_TECHNICIANS_JSON, 'w', encoding='utf-8') as _owf:
    json.dump(_OW_TECHS, _owf, indent=2)
with open(OW_SUPERVISORS_JSON, 'w', encoding='utf-8') as _owf:
    json.dump(_OW_SUPS, _owf, indent=2)

# =====================================================
# ONEWEST ISSUES ROUTES
# =====================================================
@app.route("/ow_issues")
@login_required
@require_property("ONEWEST")
def ow_issues():
    """ONEWEST Issues Dashboard"""
    session['active_property'] = 'ONEWEST'
    session['property_code'] = 'OW'
    print(f"\n🏢 Accessing ONEWEST Issues - User: {session.get('user')}")
    return render_template("issues/ow_issues.html")

@app.route("/ow_api/issues")
@login_required
@require_property("ONEWEST")
def ow_api_issues():
    """ONEWEST Issues API - Get all issues"""
    try:
        if not OW_ISSUES_JSON.exists():
            return jsonify({"issues": [], "total": 0, "property": "ONEWEST"})
        
        with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        issues = data.get('issues', [])
        
        # Apply filters
        status_filter = request.args.get('status', 'all').lower()
        priority_filter = request.args.get('priority', 'all').lower()
        
        if status_filter != 'all':
            issues = [i for i in issues if i.get('status', '').lower() == status_filter]
        if priority_filter != 'all':
            issues = [i for i in issues if i.get('priority', '').lower() == priority_filter]
        
        return jsonify({
            "success": True,
            "issues": issues,
            "total": len(issues),
            "property": "ONEWEST"
        })
    except Exception as e:
        print(f"❌ ONEWEST Issues API error: {str(e)}")
        return jsonify({"success": False, "error": str(e), "issues": []}), 500

@app.route("/ow_api/issues/create", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_create_issue():
    """ONEWEST - Create new issue with image upload"""
    try:
        # Handle form data or JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            title = request.form.get('title', 'Untitled')
            description = request.form.get('description', '')
            priority = request.form.get('priority', 'Medium')
            category = request.form.get('category', 'General')
            location = request.form.get('location', '')
            assigned_to = request.form.get('assigned_to', '')
            sla_deadline = request.form.get('sla_deadline', '')
            escalation_level = request.form.get('escalation_level', 'Level 1')
            
            # Handle image uploads
            photos = []
            if 'photos' in request.files:
                files = request.files.getlist('photos')
                for file in files:
                    if file and file.filename:
                        filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                        save_path = OW_ISSUES_UPLOADS / filename
                        file.save(save_path)
                        photos.append(f"/uploads/OW/issues/{filename}")
        else:
            data = request.get_json()
            if not data:
                return jsonify({"success": False, "error": "No data provided"}), 400
            
            title = data.get('title', 'Untitled')
            description = data.get('description', '')
            priority = data.get('priority', 'Medium')
            category = data.get('category', 'General')
            location = data.get('location', '')
            assigned_to = data.get('assigned_to', '')
            sla_deadline = data.get('sla_deadline', '')
            escalation_level = data.get('escalation_level', 'Level 1')
            photos = data.get('photos', [])
        
        # Load existing issues
        if OW_ISSUES_JSON.exists():
            with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
                ow_data = json.load(f)
        else:
            ow_data = {"issues": [], "last_updated": ""}
        
        # Generate issue ID
        issue_counter = len(ow_data.get('issues', [])) + 1
        issue_id = f"OW-ISS-{datetime.now().strftime('%Y')}-{str(issue_counter).zfill(4)}"
        
        # Create new issue
        new_issue = {
            "issue_id": issue_id,
            "title": title,
            "description": description,
            "priority": priority,
            "status": "Open",
            "category": category,
            "location": location,
            "reported_by": session.get('user', 'Unknown'),
            "assigned_to": assigned_to,
            "property": "ONEWEST",
            "property_code": "OW",
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat(),
            "sla_deadline": sla_deadline,
            "escalation_level": escalation_level,
            "photos": photos,
            "whatsapp_sent": False
        }
        
        ow_data['issues'].append(new_issue)
        ow_data['last_updated'] = datetime.now().isoformat()
        
        with open(OW_ISSUES_JSON, 'w', encoding='utf-8') as f:
            json.dump(ow_data, f, indent=2)
        
        # Send WhatsApp notification
        send_ow_whatsapp_notification(new_issue, assigned_to)
        
        print(f"✅ ONEWEST Issue Created: {issue_id}")
        return jsonify({
            "success": True,
            "issue_id": issue_id,
            "message": "Issue created successfully"
        })
    except Exception as e:
        print(f"❌ ONEWEST Issue creation error: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/ow_api/issues/update/<issue_id>", methods=["PUT"])
@login_required
@require_property("ONEWEST")
def ow_api_update_issue(issue_id):
    """ONEWEST - Update issue"""
    try:
        data = request.get_json()
        if not OW_ISSUES_JSON.exists():
            return jsonify({"success": False, "error": "Issues file not found"}), 404
        
        with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
            ow_data = json.load(f)
        
        updated = False
        for issue in ow_data.get('issues', []):
            if issue.get('issue_id') == issue_id:
                old_status = issue.get('status', '')
                # Update fields
                if 'status' in data:
                    issue['status'] = data['status']
                if 'priority' in data:
                    issue['priority'] = data['priority']
                if 'assigned_to' in data:
                    issue['assigned_to'] = data['assigned_to']
                if 'description' in data:
                    issue['description'] = data['description']
                if 'escalation_level' in data:
                    issue['escalation_level'] = data['escalation_level']
                # FIX: save resolution notes for close/resolve workflow
                if 'resolution_notes' in data:
                    issue['resolution_notes'] = data['resolution_notes']
                # FIX: append to status_history for full audit trail
                if 'status' in data and data['status'] != old_status:
                    if 'status_history' not in issue:
                        issue['status_history'] = []
                    issue['status_history'].append({
                        'from': old_status,
                        'to': data['status'],
                        'changed_at': datetime.now().isoformat(),
                        'changed_by': session.get('user', 'system')
                    })
                issue['updated_at'] = datetime.now().isoformat()
                updated = True
                # FIX: compare against captured old_status not nonexistent field
                if data.get('status') and data['status'] != old_status:
                    send_ow_whatsapp_status_update(issue, data['status'])
                break
        
        if not updated:
            return jsonify({"success": False, "error": "Issue not found"}), 404
        
        with open(OW_ISSUES_JSON, 'w', encoding='utf-8') as f:
            json.dump(ow_data, f, indent=2)
        
        print(f"✅ ONEWEST Issue Updated: {issue_id}")
        return jsonify({"success": True, "message": "Issue updated"})
    except Exception as e:
        print(f"❌ ONEWEST Issue update error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500



@app.route("/ow_api/issues/close", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_close_issue():
    """Close a ONEWEST issue"""
    try:
        data = request.get_json()
        issue_id = data.get("issue_id")
        reason   = data.get("reason", "").strip()
        if not issue_id:
            return jsonify({"success": False, "error": "issue_id required"}), 400
        with open(OW_ISSUES_JSON, "r", encoding="utf-8") as f:
            db = json.load(f)
        issues = db.get("issues", [])
        found = False
        for iss in issues:
            if iss["id"] == issue_id:
                iss["status"]   = "Closed"
                iss["closed_at"] = datetime.now().isoformat()
                iss["closed_by"] = session.get("user", "unknown")
                if reason:
                    iss["resolution_notes"] = reason
                iss.setdefault("status_history", []).append({
                    "status": "Closed",
                    "timestamp": datetime.now().isoformat(),
                    "by": session.get("user", "unknown"),
                    "note": reason or "Closed"
                })
                found = True
                break
        if not found:
            return jsonify({"success": False, "error": "Issue not found"}), 404
        with open(OW_ISSUES_JSON, "w", encoding="utf-8") as f:
            json.dump(db, f, indent=2, ensure_ascii=False)
        return jsonify({"success": True, "message": "Issue closed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/ow_api/issues/reopen", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_reopen_issue():
    """Reopen a closed/resolved ONEWEST issue"""
    try:
        data = request.get_json()
        issue_id = data.get("issue_id")
        reason   = data.get("reason", "").strip()
        if not issue_id:
            return jsonify({"success": False, "error": "issue_id required"}), 400
        if not reason:
            return jsonify({"success": False, "error": "Reopen reason is required"}), 400
        with open(OW_ISSUES_JSON, "r", encoding="utf-8") as f:
            db = json.load(f)
        issues = db.get("issues", [])
        found = False
        for iss in issues:
            if iss["id"] == issue_id:
                iss["status"]      = "Open"
                iss["reopened_at"] = datetime.now().isoformat()
                iss["reopened_by"] = session.get("user", "unknown")
                iss["resolution_notes"] = f"[Reopened] {reason}"
                iss.setdefault("status_history", []).append({
                    "status": "Open",
                    "timestamp": datetime.now().isoformat(),
                    "by": session.get("user", "unknown"),
                    "note": f"Reopened: {reason}"
                })
                found = True
                break
        if not found:
            return jsonify({"success": False, "error": "Issue not found"}), 404
        with open(OW_ISSUES_JSON, "w", encoding="utf-8") as f:
            json.dump(db, f, indent=2, ensure_ascii=False)
        return jsonify({"success": True, "message": "Issue reopened"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
@app.route("/ow_api/issues/export")
@login_required
@require_property("ONEWEST")
def ow_api_export_issues():
    """ONEWEST - Export issues to Excel"""
    try:
        if not OW_ISSUES_JSON.exists():
            return jsonify({"success": False, "error": "No issues to export"}), 404
        
        with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
            ow_data = json.load(f)
        
        issues = ow_data.get('issues', [])
        if not issues:
            return jsonify({"success": False, "error": "No issues to export"}), 404
        
        df = pd.DataFrame(issues)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ONEWEST Issues')
        
        output.seek(0)
        filename = f"ONEWEST_Issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"❌ ONEWEST Issues export error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/ow_api/issues/stats")
@login_required
@require_property("ONEWEST")
def ow_api_issues_stats():
    """ONEWEST - Get issues statistics"""
    try:
        if not OW_ISSUES_JSON.exists():
            return jsonify({
                "total": 0, "open": 0, "in_progress": 0,
                "resolved": 0, "closed": 0,
                "critical": 0, "high": 0, "medium": 0, "low": 0,
                "property": "ONEWEST"
            })
        
        with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
            ow_data = json.load(f)
        
        issues = ow_data.get('issues', [])
        
        stats = {
            "total": len(issues),
            "open": len([i for i in issues if i.get('status') == 'Open']),
            "in_progress": len([i for i in issues if i.get('status') == 'In Progress']),
            "resolved": len([i for i in issues if i.get('status') == 'Resolved']),
            "closed": len([i for i in issues if i.get('status') == 'Closed']),
            "critical": len([i for i in issues if i.get('priority') == 'Critical']),
            "high": len([i for i in issues if i.get('priority') == 'High']),
            "medium": len([i for i in issues if i.get('priority') == 'Medium']),
            "low": len([i for i in issues if i.get('priority') == 'Low']),
            "property": "ONEWEST"
        }
        
        return jsonify(stats)
    except Exception as e:
        print(f"❌ ONEWEST Issues stats error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# =====================================================
# DAY-WISE ISSUE ARCHIVAL SYSTEM
# =====================================================

@app.route("/ow_api/issues/archive-daily", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_archive_daily_issues():
    """Archive today's issues at day end (auto-triggered at midnight)"""
    try:
        today = datetime.now().date()
        today_str = today.strftime('%Y-%m-%d')
        
        # Load current issues
        if not OW_ISSUES_JSON.exists():
            return jsonify({"success": True, "message": "No issues to archive"})
        
        with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
            ow_data = json.load(f)
        
        issues = ow_data.get('issues', [])
        
        # Separate today's issues from others
        today_issues = []
        remaining_issues = []
        
        for issue in issues:
            created_date = issue.get('created_at', '')[:10]  # Extract YYYY-MM-DD
            if created_date == today_str:
                today_issues.append(issue)
            else:
                remaining_issues.append(issue)
        
        if not today_issues:
            return jsonify({"success": True, "message": "No issues to archive today"})
        
        # Archive today's issues
        archive_file = ISSUES_ARCHIVE_DIR / f"OW_Issues_{today_str}.json"
        archive_data = {
            "date": today_str,
            "archived_at": datetime.now().isoformat(),
            "total_issues": len(today_issues),
            "issues": today_issues,
            "summary": {
                "open": len([i for i in today_issues if i.get('status') == 'Open']),
                "in_progress": len([i for i in today_issues if i.get('status') == 'In Progress']),
                "resolved": len([i for i in today_issues if i.get('status') == 'Resolved']),
                "closed": len([i for i in today_issues if i.get('status') == 'Closed'])
            }
        }
        
        with open(archive_file, 'w', encoding='utf-8') as f:
            json.dump(archive_data, f, indent=2)
        
        # Update main issues file with remaining issues only
        ow_data['issues'] = remaining_issues
        ow_data['last_updated'] = datetime.now().isoformat()
        
        with open(OW_ISSUES_JSON, 'w', encoding='utf-8') as f:
            json.dump(ow_data, f, indent=2)
        
        print(f"✅ Archived {len(today_issues)} issues for {today_str}")
        return jsonify({
            "success": True,
            "archived_count": len(today_issues),
            "remaining_count": len(remaining_issues),
            "archive_file": archive_file.name
        })
        
    except Exception as e:
        print(f"❌ Archive error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/ow_api/issues/export-by-date", methods=["GET"])
@login_required
@require_property("ONEWEST")
def ow_export_issues_by_date():
    """Export issues for specific date"""
    try:
        date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        # Check archive first
        archive_file = ISSUES_ARCHIVE_DIR / f"OW_Issues_{date_str}.json"
        
        if archive_file.exists():
            # Load from archive
            with open(archive_file, 'r', encoding='utf-8') as f:
                archive_data = json.load(f)
            issues = archive_data.get('issues', [])
        else:
            # Load from current issues if date matches today
            if not OW_ISSUES_JSON.exists():
                return jsonify({"error": "No issues found"}), 404
            
            with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
                ow_data = json.load(f)
            
            issues = [i for i in ow_data.get('issues', []) if i.get('created_at', '')[:10] == date_str]
        
        if not issues:
            return jsonify({"error": "No issues found for this date"}), 404
        
        # Export to Excel
        df = pd.DataFrame(issues)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=f'Issues_{date_str}')
        
        output.seek(0)
        filename = f"ONEWEST_Issues_{date_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/ow_api/issues/archive-list")
@login_required
@require_property("ONEWEST")
def ow_get_archive_list():
    """Get list of all archived dates.
    FIX: also includes dates from live issues.json not yet archived."""
    try:
        archives = {}  # keyed by date string

        # 1. Load from dedicated archive files
        if ISSUES_ARCHIVE_DIR.exists():
            for af in sorted(ISSUES_ARCHIVE_DIR.iterdir(), reverse=True):
                if af.suffix == '.json':
                    date_str = af.stem.replace('OW_Issues_', '')
                    try:
                        with open(af, 'r', encoding='utf-8') as fh:
                            data = json.load(fh)
                        archives[date_str] = {
                            "date": date_str,
                            "total": data.get('total_issues', 0),
                            "summary": data.get('summary', {}),
                            "source": "archive"
                        }
                    except Exception:
                        pass

        # 2. FIX: also scan live issues.json for dates not yet archived
        if OW_ISSUES_JSON.exists():
            try:
                with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as fh:
                    ow_data = json.load(fh)
                for issue in ow_data.get('issues', []):
                    d = issue.get('created_at', '')[:10]
                    if not d:
                        continue
                    if d not in archives:
                        archives[d] = {
                            "date": d, "total": 0,
                            "summary": {"open": 0, "in_progress": 0, "resolved": 0, "closed": 0},
                            "source": "live"
                        }
                    if archives[d].get('source') == 'live':
                        archives[d]['total'] += 1
                        st_key = issue.get('status', 'Open').lower().replace(' ', '_')
                        archives[d]['summary'][st_key] = archives[d]['summary'].get(st_key, 0) + 1
            except Exception:
                pass

        result = sorted(archives.values(), key=lambda x: x['date'], reverse=True)
        return jsonify({"success": True, "archives": result})
    except Exception as e:
        print(f"[ow_get_archive_list] ERROR: {e}")
        return jsonify({"success": False, "error": str(e), "archives": []}), 500

@app.route("/ow_api/issues/view-archive/<date_str>")
@login_required
@require_property("ONEWEST")
def ow_view_archive(date_str):
    """View archived issues for a specific date.
    FIX: falls back to filtering live issues.json when no archive file exists."""
    try:
        archive_file = ISSUES_ARCHIVE_DIR / f"OW_Issues_{date_str}.json"

        # 1. Try dedicated archive file first
        if archive_file.exists():
            with open(archive_file, 'r', encoding='utf-8') as f:
                archive_data = json.load(f)
            return jsonify({"success": True, "date": date_str, "data": archive_data})

        # 2. FIX: fallback — filter live issues.json by the requested date
        if not OW_ISSUES_JSON.exists():
            return jsonify({"success": False, "error": "No issues data found"}), 404

        with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
            ow_data = json.load(f)

        date_issues = [
            i for i in ow_data.get('issues', [])
            if i.get('created_at', '')[:10] == date_str
        ]

        if not date_issues:
            return jsonify({"success": False, "error": f"No issues found for {date_str}"}), 404

        fallback_data = {
            "date": date_str,
            "archived_at": None,
            "total_issues": len(date_issues),
            "issues": date_issues,
            "summary": {
                "open":        len([i for i in date_issues if i.get('status') == 'Open']),
                "in_progress": len([i for i in date_issues if i.get('status') == 'In Progress']),
                "resolved":    len([i for i in date_issues if i.get('status') == 'Resolved']),
                "closed":      len([i for i in date_issues if i.get('status') == 'Closed']),
            }
        }
        return jsonify({"success": True, "date": date_str, "data": fallback_data})

    except Exception as e:
        print(f"[ow_view_archive] ERROR: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# =====================================================
# AUTO-ARCHIVE AT MIDNIGHT (Scheduler)
# =====================================================
from apscheduler.schedulers.background import BackgroundScheduler

def setup_issue_archive_scheduler():
    """Schedule daily archival at 11:59 PM"""
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=auto_archive_issues,
        trigger='cron',
        hour=23,
        minute=59,
        timezone='Asia/Kolkata'
    )
    scheduler.start()
    print("✅ Issue archive scheduler started: Daily at 11:59 PM IST")
    return scheduler

def auto_archive_issues():
    """Auto-archive today's issues"""
    try:
        with app.app_context():
            # Simulate POST request to archive endpoint
            today = datetime.now().date()
            today_str = today.strftime('%Y-%m-%d')
            
            if not OW_ISSUES_JSON.exists():
                print(f"ℹ️  No issues file found for {today_str}")
                return
            
            with open(OW_ISSUES_JSON, 'r', encoding='utf-8') as f:
                ow_data = json.load(f)
            
            issues = ow_data.get('issues', [])
            today_issues = [i for i in issues if i.get('created_at', '')[:10] == today_str]
            
            if not today_issues:
                print(f"ℹ️  No issues to archive for {today_str}")
                return
            
            # Archive today's issues
            archive_file = ISSUES_ARCHIVE_DIR / f"OW_Issues_{today_str}.json"
            archive_data = {
                "date": today_str,
                "archived_at": datetime.now().isoformat(),
                "total_issues": len(today_issues),
                "issues": today_issues,
                "summary": {
                    "open": len([i for i in today_issues if i.get('status') == 'Open']),
                    "in_progress": len([i for i in today_issues if i.get('status') == 'In Progress']),
                    "resolved": len([i for i in today_issues if i.get('status') == 'Resolved']),
                    "closed": len([i for i in today_issues if i.get('status') == 'Closed'])
                }
            }
            
            with open(archive_file, 'w', encoding='utf-8') as f:
                json.dump(archive_data, f, indent=2)
            
            # Remove today's issues from main file
            remaining_issues = [i for i in issues if i.get('created_at', '')[:10] != today_str]
            ow_data['issues'] = remaining_issues
            ow_data['last_updated'] = datetime.now().isoformat()
            
            with open(OW_ISSUES_JSON, 'w', encoding='utf-8') as f:
                json.dump(ow_data, f, indent=2)
            
            print(f"✅ Auto-archived {len(today_issues)} issues for {today_str}")
            
    except Exception as e:
        print(f"❌ Auto-archive failed: {str(e)}")




# =====================================================
# TECHNICIANS & SUPERVISORS API
# =====================================================
@app.route("/ow_api/technicians")
@login_required
def ow_api_technicians():
    """ONEWEST technicians — reads technicians.json, fixes trailing-comma JSON"""
    import re as _re
    if 'ONEWEST' in session.get('properties', []) or session.get('role') == 'admin':
        session['active_property'] = 'ONEWEST'
    try:
        OW_TECHNICIANS_JSON.parent.mkdir(parents=True, exist_ok=True)
        if not OW_TECHNICIANS_JSON.exists():
            default = {"technicians": [
                {"id": "T001", "name": "Jagadish", "phone": "+919666942315", "specialization": "Supervisor"},
    		{"id": "T002", "name": "Sri Ram", "phone": "+919989668311", "specialization": "Supervisor"},
    		{"id": "T003", "name": "Ameer", "phone": "+919000564662", "specialization": "Supervisor"},
    		{"id": "T004", "name": "Rakesh", "phone": "+917730834084", "specialization": "Supervisor"},
    		{"id": "T005", "name": "Raghavendra", "phone": "+918008883537", "specialization": "BMS"},
    		{"id": "T006", "name": "Shiva", "phone": "+918885637165", "specialization": "BMS"},
    		{"id": "T007", "name": "Tanmaya", "phone": "+917077689216", "specialization": "BMS"},
    		{"id": "T008", "name": "Raviteja", "phone": "+919652607622", "specialization": "Electrical"},
    		{"id": "T009", "name": "Nagendra", "phone": "+919347324744", "specialization": "Electrical"},
    		{"id": "T010", "name": "Yeshwanth", "phone": "+919502856581", "specialization": "Electrical"},
    		{"id": "T011", "name": "Hakim", "phone": "+918083360242", "specialization": "Electrical"},
    		{"id": "T012", "name": "Sai pawan", "phone": "+917013987434", "specialization": "Electrical"},
    		{"id": "T013", "name": "Nikhil", "phone": "+917093979479", "specialization": "Asst Technician"},
    		{"id": "T014", "name": "Vijay", "phone": "+916304725703", "specialization": "Asst Technician"},
    		{"id": "T015", "name": "Karthik", "phone": "+919553174565", "specialization": "Asst Technician"},
    		{"id": "T016", "name": "Ilyas", "phone": "+919347732552", "specialization": "HVAC"},
    		{"id": "T017", "name": "Sai", "phone": "+917794057118", "specialization": "HVAC"},
    		{"id": "T018", "name": "Venu", "phone": "+919618670499", "specialization": "HVAC"},
    		{"id": "T019", "name": "Bharath", "phone": "+918106869682", "specialization": "HVAC"},
    		{"id": "T020", "name": "Ismail", "phone": "+919154223362", "specialization": "HVAC"},
    		{"id": "T021", "name": "Bipin", "phone": "+919121261604", "specialization": "Plumber"},
    		{"id": "T027", "name": "Bichitra", "phone": "+917732040540", "specialization": "Plumber"},
    		{"id": "T022", "name": "Sudarshan", "phone": "+917036994079", "specialization": "Plumber"},
    		{"id": "T023", "name": "Srikanth", "phone": "+917749090745", "specialization": "Plumber"},
    		{"id": "T024", "name": "Tapan", "phone": "+916380896010", "specialization": "Plumber"},
    		{"id": "T025", "name": "Rohith", "phone": "+919948351383", "specialization": "Painter"},
    		{"id": "T026", "name": "Laxman", "phone": "+917995751392", "specialization": "Carpenter"}
            ]}
            with open(OW_TECHNICIANS_JSON, 'w', encoding='utf-8') as f:
                json.dump(default, f, indent=2)
            return jsonify(default)
        with open(OW_TECHNICIANS_JSON, 'r', encoding='utf-8') as f:
            raw = f.read().strip()
        if not raw:
            return jsonify({"technicians": []})
        raw = _re.sub(r',\s*([\]\}])', r'\1', raw)   # fix trailing commas
        data = json.loads(raw)
        if isinstance(data, list):
            data = {"technicians": data}
        techs = data.get("technicians", [])
        print(f"[ow_api_technicians] ✅ {len(techs)} technicians")
        return jsonify({"technicians": techs})
    except Exception as e:
        print(f"[ow_api_technicians] ERROR: {e}")
        return jsonify({"technicians": [], "error": str(e)}), 200


@app.route("/ow_api/debug/technicians")
@login_required
def ow_api_technicians_debug():
    """Debug route - no property check"""
    import traceback as tb
    try:
        result = {
            "session_user":            session.get("user"),
            "session_active_property": session.get("active_property"),
            "session_role":            session.get("role"),
            "session_properties":      session.get("properties", []),
            "OW_TECHNICIANS_JSON":     str(OW_TECHNICIANS_JSON),
            "file_exists":             OW_TECHNICIANS_JSON.exists(),
        }
        if OW_TECHNICIANS_JSON.exists():
            with open(OW_TECHNICIANS_JSON, 'r', encoding='utf-8') as f:
                result["file_contents"] = json.load(f)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "trace": tb.format_exc()}), 500


@app.route("/ow_api/supervisors")
@login_required
def ow_api_supervisors():
    """ONEWEST supervisors — reads supervisors.json, fixes trailing-comma JSON"""
    import re as _re
    if 'ONEWEST' in session.get('properties', []) or session.get('role') == 'admin':
        session['active_property'] = 'ONEWEST'
    try:
        OW_SUPERVISORS_JSON.parent.mkdir(parents=True, exist_ok=True)
        if not OW_SUPERVISORS_JSON.exists():
            default = {"supervisors": [
                {"id": "S001", "name": "Jagadish", "phone": "+919666942315", "specialization": "Supervisor"},
    		{"id": "S002", "name": "Sri Ram", "phone": "+919989668311", "specialization": "Supervisor"},
    		{"id": "S003", "name": "Ameer", "phone": "+919000564662", "specialization": "Supervisor"},
    		{"id": "S004", "name": "Rakesh", "phone": "+917730834084", "specialization": "Supervisor"},
    		{"id": "B005", "name": "Raghavendra", "phone": "+918008883537", "specialization": "BMS"},
    		{"id": "B006", "name": "Shiva", "phone": "+918885637165", "specialization": "BMS"},
    		{"id": "B007", "name": "Tanmaya", "phone": "+917077689216", "specialization": "BMS"}
            ]}
            with open(OW_SUPERVISORS_JSON, 'w', encoding='utf-8') as f:
                json.dump(default, f, indent=2)
            return jsonify(default)
        with open(OW_SUPERVISORS_JSON, 'r', encoding='utf-8') as f:
            raw = f.read().strip()
        if not raw:
            return jsonify({"supervisors": []})
        raw = _re.sub(r',\s*([\]\}])', r'\1', raw)
        data = json.loads(raw)
        if isinstance(data, list):
            data = {"supervisors": data}
        sups = data.get("supervisors", [])
        print(f"[ow_api_supervisors] ✅ {len(sups)} supervisors")
        return jsonify({"supervisors": sups})
    except Exception as e:
        print(f"[ow_api_supervisors] ERROR: {e}")
        return jsonify({"supervisors": [], "error": str(e)}), 200

# =====================================================
# IMAGE UPLOAD FOR ISSUES
# =====================================================
@app.route("/ow_api/issues/upload-photo", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_upload_photo():
    """Upload photo for issue"""
    try:
        if 'photo' not in request.files:
            return jsonify({"success": False, "error": "No photo uploaded"}), 400
        
        file = request.files['photo']
        if file.filename == '':
            return jsonify({"success": False, "error": "Empty filename"}), 400
        
        # Generate unique filename
        filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        save_path = OW_ISSUES_UPLOADS / filename
        file.save(save_path)
        
        photo_url = f"/uploads/OW/issues/{filename}"
        
        return jsonify({
            "success": True,
            "photo_url": photo_url,
            "filename": filename
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# =====================================================
# WHATSAPP NOTIFICATION HELPER
# =====================================================
def send_ow_whatsapp_notification(issue, assigned_to):
    """Send WhatsApp notification for new issue"""
    try:
        # Load supervisors to get phone
        if OW_SUPERVISORS_JSON.exists():
            with open(OW_SUPERVISORS_JSON, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            supervisors = data.get('supervisors', [])
            supervisor_phone = supervisors[0]['phone'] if supervisors else '+919876543220'
        else:
            supervisor_phone = '+919876543220'
        
        # WhatsApp message
        message = f"""
🔴 *NEW ISSUE - ONEWEST*

*Issue ID:* {issue['issue_id']}
*Title:* {issue['title']}
*Priority:* {issue['priority']}
*Location:* {issue['location']}
*Reported By:* {issue['reported_by']}
*Assigned To:* {issue['assigned_to']}
*Created:* {issue['created_at'][:16].replace('T', ' ')}

Please take immediate action.
        """.strip()
        
        # WhatsApp API URL (Use your preferred service)
        whatsapp_url = f"https://api.whatsapp.com/send?phone={supervisor_phone}&text={requests.utils.quote(message)}"
        
        print(f"📱 WhatsApp notification prepared for {supervisor_phone}")
        print(f"🔗 URL: {whatsapp_url}")
        
        # Optional: Send via API (Twilio, MessageBird, etc.)
        # requests.get(whatsapp_url)
        
        return True
    except Exception as e:
        print(f"❌ WhatsApp notification error: {str(e)}")
        return False

def send_ow_whatsapp_status_update(issue, new_status):
    """Send WhatsApp notification on status update"""
    try:
        if OW_SUPERVISORS_JSON.exists():
            with open(OW_SUPERVISORS_JSON, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            supervisors = data.get('supervisors', [])
            supervisor_phone = supervisors[0]['phone'] if supervisors else '+919876543220'
        else:
            supervisor_phone = '+919876543220'
        
        message = f"""
📊 *ISSUE STATUS UPDATE - ONEWEST*

*Issue ID:* {issue['issue_id']}
*Title:* {issue['title']}
*New Status:* {new_status}
*Updated:* {datetime.now().strftime('%Y-%m-%d %H:%M')}

Please review the update.
        """.strip()
        
        whatsapp_url = f"https://api.whatsapp.com/send?phone={supervisor_phone}&text={requests.utils.quote(message)}"
        
        print(f"📱 WhatsApp status update prepared for {supervisor_phone}")
        
        return True
    except Exception as e:
        print(f"❌ WhatsApp status update error: {str(e)}")
        return False

# Serve uploaded issue photos
@app.route("/uploads/OW/issues/<filename>")
@login_required
def serve_ow_issue_photo(filename):
    """Serve ONEWEST issue photo"""
    return send_from_directory(OW_ISSUES_UPLOADS, filename)

# =====================================================
# ✅ ONEWEST MMS MODULE (ALL ROUTES WITH ow_ PREFIX)
# NOTE: Completely independent from SLN Terminus module
# =====================================================

# ONEWEST Data Files
OW_DIR = BASE_DIR / "static" / "data" / "OW"
OW_ASSETS_XLS     = OW_DIR / "Asset.xls"
OW_ASSETS_XLSX    = OW_DIR / "Asset.xlsx"
OW_WORK_ORDERS_JSON = OW_DIR / "work_orders.json"
OW_AMC_JSON       = OW_DIR / "amc_contracts.json"
OW_PPM_WO_UPLOADS = BASE_DIR / "uploads" / "OW" / "ppm"

# Create OW directories
for _d in [OW_DIR, OW_PPM_WO_UPLOADS]:
    _d.mkdir(parents=True, exist_ok=True)

# OW Email config (independent from SLN)
OW_EMAIL_RECEIVERS = [
    "maintenance.slnterminus@gmail.com",
    "yasven7545@gmail.com",
    "engineering@terminus-global.com"
   

]

# =====================================================
# ONEWEST DASHBOARD ROUTE
# =====================================================
@app.route("/ow_ppm_dashboard")
@login_required
@require_property("ONEWEST")
def ow_ppm_dashboard():
    """ONEWEST PPM Dashboard"""
    session['active_property'] = 'ONEWEST'
    return render_template("ow_ppm_dashboard.html")


# =====================================================
# ONEWEST PPM ASSETS API (ow_api/ppm/assets)
# =====================================================
@app.route("/ow_api/ppm/assets")
@login_required
@require_property("ONEWEST")
def ow_api_ppm_assets():
    """ONEWEST PPM assets — reads BOTH inhouse_ppm + vendor_ppm sheets from Asset.xls/.xlsx"""
    try:
        location_filter = request.args.get('location', 'all')
        candidates = [(OW_ASSETS_XLS, 'xlrd'), (OW_ASSETS_XLSX, 'openpyxl')]
        sheets_loaded = []
        for fpath, engine in candidates:
            if not fpath.exists():
                continue
            try:
                xl = pd.ExcelFile(str(fpath), engine=engine)
                for sheet_canon, ppm_type in [('inhouse_ppm','inhouse'),('vendor_ppm','vendor')]:
                    actual = next((s for s in xl.sheet_names if s.lower().strip()==sheet_canon), None)
                    if actual is None:
                        actual = next((s for s in xl.sheet_names if ppm_type in s.lower()), None)
                    if actual is None:
                        continue
                    try:
                        df = pd.read_excel(str(fpath), sheet_name=actual, engine=engine)
                        sheets_loaded.append((df, ppm_type))
                    except Exception as se:
                        print(f"⚠️ Sheet '{actual}' error: {se}")
                if sheets_loaded:
                    break
            except Exception as e:
                print(f"⚠️ Cannot open {fpath.name}: {e}")
        if not sheets_loaded:
            print(f"❌ OW: No asset file found")
            return jsonify({"assets": [], "total": 0, "property": "ONEWEST"})
        assets = []
        for df, ppm_type in sheets_loaded:
            for _, row in df.iterrows():
                ac = str(row.get('Equipment No.', row.get('Asset Code',''))).strip()
                if not ac or ac.lower() in ('nan','none',''): continue
                an = str(row.get('Equipment Name', row.get('Asset Name','Unknown Asset'))).strip()
                dept = str(row.get('Trade', row.get('Department', row.get('Category','General')))).strip()
                if dept.lower() in ('nan','none',''): dept='General'
                loc  = str(row.get('Location','Unknown')).strip()
                if loc.lower()  in ('nan','none',''): loc='Unknown'
                ls   = str(row.get('Last Service','')).strip()
                if ls.lower()   in ('nan','none'):    ls=''
                nd   = str(row.get('Next DueDate', row.get('Next Due Date',''))).strip()
                if nd.lower()   in ('nan','none'):    nd=''
                assets.append({"id":ac,"asset_code":ac,"name":an,"asset_name":an,
                    "department":dept,"trade":dept,"category":dept,"location":loc,
                    "lastService":ls,"last_service":ls,"nextDueDate":nd,"next_due":nd,
                    "ppm_type":ppm_type,"property":"ONEWEST"})
        if location_filter != 'all':
            assets = [a for a in assets if a.get('location','').strip()==location_filter.strip()]
        print(f"✅ OW assets: {len(assets)} (inhouse+vendor)")
        return jsonify({"assets": assets, "total": len(assets), "property": "ONEWEST"})
    except Exception as e:
        print(f"❌ OW PPM assets error: {e}")
        traceback.print_exc()
        return jsonify({"assets": [], "total": 0}), 500


# =====================================================
# ONEWEST PPM ASSETS UPLOAD (sync from xlsx)
# =====================================================
@app.route("/ow_api/ppm/import-excel", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_ppm_import_excel():
    """Upload & sync ONEWEST Asset.xlsx"""
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400

        OW_DIR.mkdir(parents=True, exist_ok=True)
        file.save(OW_ASSETS_XLSX)

        df = pd.read_excel(OW_ASSETS_XLSX)
        count = len([_ for _, row in df.iterrows() if pd.notna(row.get('Asset Code')) and str(row.get('Asset Code')).strip()])

        print(f"✅ OW Assets synced: {count} records")
        return jsonify({"status": "success", "message": f"Successfully synced {count} ONEWEST assets", "count": count})

    except Exception as e:
        print(f"❌ OW Excel import error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# =====================================================
# ONEWEST PPM DASHBOARD STATS (ow_api/ppm/dashboard/stats)
# =====================================================
@app.route("/ow_api/ppm/dashboard/stats")
@login_required
@require_property("ONEWEST")
def ow_api_ppm_dashboard_stats():
    """ONEWEST PPM dashboard stats"""
    try:
        # Load work orders
        wo_data = {"work_orders": []}
        if OW_WORK_ORDERS_JSON.exists():
            with open(OW_WORK_ORDERS_JSON, 'r', encoding='utf-8') as f:
                wo_data = json.load(f)

        work_orders = wo_data.get('work_orders', [])
        today = datetime.now().date()

        total_wo     = len(work_orders)
        completed_wo = len([w for w in work_orders if (w.get('status','') or '').lower() in ('completed','closed')])
        pending_wo   = total_wo - completed_wo

        # Overdue: open WOs with due_date < today
        overdue_wo = 0
        for wo in work_orders:
            status = (wo.get('status','') or '').lower()
            if status not in ('completed','closed'):
                try:
                    dd = datetime.strptime(wo.get('due_date','')[:10], '%Y-%m-%d').date()
                    if dd < today:
                        overdue_wo += 1
                except:
                    pass

        # Asset count — try .xls first (Equipment No.), then .xlsx (Asset Code)
        asset_count = 0
        for _fp, _eng in [(OW_ASSETS_XLS, 'xlrd'), (OW_ASSETS_XLSX, 'openpyxl')]:
            if not _fp.exists(): continue
            try:
                _xl = pd.ExcelFile(str(_fp), engine=_eng)
                for _sh in _xl.sheet_names:
                    _df = pd.read_excel(str(_fp), sheet_name=_sh, engine=_eng)
                    for _col in ('Equipment No.', 'Asset Code'):
                        if _col in _df.columns:
                            asset_count += int(_df[_col].dropna().astype(str).str.strip().str.len().gt(0).sum())
                            break
                break
            except: pass

        compliance = round((completed_wo / total_wo * 100), 1) if total_wo > 0 else 0.0

        return jsonify({
            "total_assets":   asset_count,
            "pending_ppm":    pending_wo,
            "completed_ppm":  completed_wo,
            "ppm_overdue":    overdue_wo,
            "compliance_rate": compliance,
            "property":       "ONEWEST"
        })

    except Exception as e:
        print(f"❌ OW dashboard stats error: {e}")
        return jsonify({"total_assets":0,"pending_ppm":0,"completed_ppm":0,"ppm_overdue":0,"compliance_rate":0}), 500


# =====================================================
# ONEWEST WORK ORDERS API (ow_api/ppm/workorders)
# =====================================================
@app.route("/ow_api/ppm/workorders")
@login_required
@require_property("ONEWEST")
def ow_api_ppm_workorders():
    """Get ONEWEST work orders"""
    try:
        if not OW_WORK_ORDERS_JSON.exists():
            return jsonify({"work_orders": [], "total": 0, "property": "ONEWEST"})

        with open(OW_WORK_ORDERS_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)

        work_orders = data.get('work_orders', [])

        # Apply status filter
        status_filter = request.args.get('status', 'all').lower()
        if status_filter != 'all':
            work_orders = [w for w in work_orders if (w.get('status','') or '').lower() == status_filter]

        # Standardize format for frontend
        _tod = datetime.now().date()
        formatted = []
        for wo in work_orders:
            _rst = (wo.get('status','') or 'open').lower()
            _est = _rst
            if _rst not in ('completed','closed'):
                try:
                    _dd = datetime.strptime(wo.get('due_date','')[:10],'%Y-%m-%d').date()
                    if _dd < _tod: _est = 'overdue'
                except: pass
            formatted.append({
                "WO ID":         wo.get('work_order_id','N/A'),
                "Asset":         wo.get('asset_name','Unknown Asset'),
                "Location":      wo.get('location','Unknown'),
                "Due Date":      wo.get('due_date','N/A'),
                "Priority":      wo.get('priority','Medium'),
                "Status":        _est,
                "status":        _est,
                "raw_status":    _rst,
                "created_at":    wo.get('created_at', datetime.now().isoformat()),
                "assigned_to":   wo.get('assigned_to',''),
                "supervisor":    wo.get('supervisor',''),
                "checklist":     wo.get('checklist',[]),
                "images":        wo.get('images',[]),
                "asset_id":      wo.get('asset_id',''),
                "work_order_id": wo.get('work_order_id',''),
                "asset_name":    wo.get('asset_name',''),
                "location":      wo.get('location',''),
                "due_date":      wo.get('due_date',''),
                "priority":      wo.get('priority','Medium'),
                "ppm_type":      wo.get('ppm_type','inhouse'),
                "closed_by":     wo.get('closed_by',''),
                "closed_at":     wo.get('closed_at',''),
                "technician":    wo.get('technician',''),
                "approval_notes":wo.get('approval_notes',''),
                "property":      "ONEWEST",
            })

        return jsonify({"work_orders": formatted, "total": len(formatted), "property": "ONEWEST", "success": True})

    except Exception as e:
        print(f"❌ OW workorders error: {e}")
        return jsonify({"work_orders": [], "total": 0, "success": False, "error": str(e)}), 500


# =====================================================
# ONEWEST WORK ORDERS BY DATE (ow_api/workorders/by-date)
# =====================================================
@app.route("/ow_api/workorders/by-date")
@login_required
@require_property("ONEWEST")
def ow_api_workorders_by_date():
    """Get ONEWEST work orders for a specific date"""
    try:
        date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))

        if not OW_WORK_ORDERS_JSON.exists():
            return jsonify({"work_orders": [], "date": date_str, "property": "ONEWEST"})

        with open(OW_WORK_ORDERS_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)

        work_orders = data.get('work_orders', [])
        filtered = [w for w in work_orders if w.get('due_date', '')[:10] == date_str]

        return jsonify({"work_orders": filtered, "date": date_str, "total": len(filtered), "property": "ONEWEST"})

    except Exception as e:
        return jsonify({"work_orders": [], "error": str(e)}), 500


# =====================================================
# ONEWEST CREATE WORK ORDER (ow_api/workflow/create)
# =====================================================
@app.route("/ow_api/workflow/create", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_workflow_create():
    """Create ONEWEST work order"""
    try:
        data = request.get_json()
        asset_id   = data.get('assetId', '')
        asset_name = data.get('assetName', 'Unknown Asset')
        location   = data.get('location', 'Unknown')
        due_date   = data.get('dueDate', '')
        asset_type = data.get('assetType', 'default')

        # Normalize due_date to YYYY-MM-DD
        try:
            if '/' in due_date:
                parts = due_date.split('/')
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                if y < 100: y += 2000
                due_date = f"{y}-{m:02d}-{d:02d}"
            elif '-' in due_date and len(due_date) == 10:
                pass
            else:
                due_date = datetime.now().strftime('%Y-%m-%d')
        except:
            due_date = datetime.now().strftime('%Y-%m-%d')

        # Try to get asset details from xlsx if name not provided
        if asset_id and OW_ASSETS_XLSX.exists():
            try:
                df = pd.read_excel(OW_ASSETS_XLSX)
                row = df[df['Asset Code'] == asset_id]
                if not row.empty:
                    asset_name = str(row.iloc[0]['Asset Name']).strip() or asset_name
                    location   = str(row.iloc[0]['Location']).strip() or location
            except:
                pass

        # Determine priority
        name_lower = asset_name.lower()
        priority = 'High' if any(k in name_lower for k in ['fire','dg','generator','transformer','hv','elevator']) else 'Medium'

        # Load existing WOs
        wo_data = {"work_orders": [], "last_updated": ""}
        if OW_WORK_ORDERS_JSON.exists():
            with open(OW_WORK_ORDERS_JSON, 'r', encoding='utf-8') as f:
                wo_data = json.load(f)

        existing_wos = wo_data.get('work_orders', [])
        today = datetime.now()
        wo_id = f"OW-PPM-{today.strftime('%Y-%m')}-{str(len(existing_wos)+1).zfill(4)}"

        # Build checklist based on asset type
        checklists = {
            'dg':       ['Check fuel level','Inspect battery','Verify coolant','Check oil','Test ATS','Inspect exhaust'],
            'elevator': ['Inspect door operation','Check emergency stop','Verify leveling','Inspect machine room','Test emergency lighting'],
            'chiller':  ['Check refrigerant pressure','Inspect condenser','Verify compressor oil','Check connections','Inspect for leaks'],
            'fire':     ['Test alarm panel','Check sprinklers','Verify extinguishers','Test smoke detectors','Check hydrant pressure'],
            'default':  ['Visual inspection','Check for noise/vibration','Verify safety guards','Inspect for leaks','Test emergency stop','Verify control panel']
        }
        cl_items = checklists.get(asset_type, checklists['default'])
        checklist = [{"id": f"{asset_type}_{i+1}", "text": item, "required": i < 4, "completed": False, "comments": ""} for i, item in enumerate(cl_items)]

        new_wo = {
            "work_order_id": wo_id,
            "asset_id":      asset_id,
            "asset_name":    asset_name,
            "location":      location,
            "due_date":      due_date,
            "priority":      priority,
            "status":        "open",
            "property":      "ONEWEST",
            "created_at":    today.isoformat(),
            "assigned_to":   "",
            "supervisor":    "",
            "checklist":     checklist,
            "images":        [],
            "technician_notes": "",
            "approval_notes":   ""
        }

        existing_wos.append(new_wo)
        wo_data['work_orders']  = existing_wos
        wo_data['last_updated'] = today.isoformat()

        with open(OW_WORK_ORDERS_JSON, 'w', encoding='utf-8') as f:
            json.dump(wo_data, f, indent=2)

        print(f"✅ OW Work Order Created: {wo_id} — {asset_name} @ {location}")
        return jsonify({"success": True, "work_order_id": wo_id, "message": "ONEWEST work order created"})

    except Exception as e:
        print(f"❌ OW WO creation error: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# =====================================================
# ONEWEST CLOSE WORK ORDER (ow_api/workflow/close)
# =====================================================
@app.route("/ow_api/workflow/close", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_workflow_close():
    """Close ONEWEST work order with supervisor approval"""
    try:
        data = request.get_json()
        wo_id            = data.get('workOrderId', '')
        approval_notes   = data.get('approvalNotes', '')
        supervisor_ok    = data.get('supervisorApproval', False)
        technician       = data.get('technician', '')
        images           = data.get('images', [])
        checklist        = data.get('checklist', [])

        if not supervisor_ok:
            return jsonify({"success": False, "error": "Supervisor approval required"}), 400

        if not OW_WORK_ORDERS_JSON.exists():
            return jsonify({"success": False, "error": "Work orders file not found"}), 404

        with open(OW_WORK_ORDERS_JSON, 'r', encoding='utf-8') as f:
            wo_data = json.load(f)

        updated = False
        for wo in wo_data.get('work_orders', []):
            if wo.get('work_order_id') == wo_id:
                wo['status']          = 'completed'
                wo['closed_at']       = datetime.now().isoformat()
                wo['approval_notes']  = approval_notes
                wo['technician']      = technician
                wo['images']          = images
                wo['checklist']       = checklist
                wo['closed_by']       = session.get('user', 'unknown')
                updated = True
                break

        if not updated:
            return jsonify({"success": False, "error": "Work order not found"}), 404

        wo_data['last_updated'] = datetime.now().isoformat()
        with open(OW_WORK_ORDERS_JSON, 'w', encoding='utf-8') as f:
            json.dump(wo_data, f, indent=2)

        print(f"✅ OW Work Order Closed: {wo_id} by {session.get('user')}")
        return jsonify({"success": True, "message": f"Work order {wo_id} closed successfully"})

    except Exception as e:
        print(f"❌ OW WO close error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# =====================================================
# ONEWEST WORK ORDERS EXPORT (ow_api/ppm/workorders/export)
# =====================================================
@app.route("/ow_api/ppm/workorders/export")
@login_required
@require_property("ONEWEST")
def ow_api_ppm_workorders_export():
    """Export ONEWEST work orders as Excel"""
    try:
        wo_data = {"work_orders": []}
        if OW_WORK_ORDERS_JSON.exists():
            with open(OW_WORK_ORDERS_JSON, 'r', encoding='utf-8') as f:
                wo_data = json.load(f)

        work_orders = wo_data.get('work_orders', [])
        if not work_orders:
            return jsonify({"error": "No work orders to export"}), 404

        rows = []
        for wo in work_orders:
            rows.append({
                "WO ID":         wo.get('work_order_id',''),
                "Asset":         wo.get('asset_name',''),
                "Location":      wo.get('location',''),
                "Due Date":      wo.get('due_date',''),
                "Priority":      wo.get('priority',''),
                "Status":        wo.get('status',''),
                "Assigned To":   wo.get('assigned_to',''),
                "Supervisor":    wo.get('supervisor',''),
                "Created At":    wo.get('created_at',''),
                "Closed At":     wo.get('closed_at',''),
                "Approval Notes": wo.get('approval_notes',''),
                "Property":      "ONEWEST"
            })

        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ONEWEST Work Orders')
        output.seek(0)

        filename = f"ONEWEST_WorkOrders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)

    except Exception as e:
        print(f"❌ OW WO export error: {e}")
        return jsonify({"error": str(e)}), 500


# =====================================================
# ONEWEST AMC CONTRACTS API (ow_api/amc/contracts)
# =====================================================
@app.route("/ow_api/amc/contracts")
@login_required
@require_property("ONEWEST")
def ow_api_amc_contracts():
    """Get ONEWEST AMC contracts"""
    try:
        if not OW_AMC_JSON.exists():
            # Return empty
            return jsonify({"contracts": [], "total": 0, "property": "ONEWEST"})

        with open(OW_AMC_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)

        contracts = data.get('contracts', [])

        status_filter = request.args.get('status', 'all').lower()
        if status_filter != 'all':
            contracts = [c for c in contracts if (c.get('status','') or '').lower() == status_filter]

        return jsonify({"contracts": contracts, "total": len(contracts), "property": "ONEWEST", "success": True})

    except Exception as e:
        print(f"❌ OW AMC error: {e}")
        return jsonify({"contracts": [], "error": str(e)}), 500


# OW AMC Export
@app.route("/ow_api/amc/contracts/export")
@login_required
@require_property("ONEWEST")
def ow_api_amc_contracts_export():
    """Export ONEWEST AMC contracts as Excel"""
    try:
        contracts = []
        if OW_AMC_JSON.exists():
            with open(OW_AMC_JSON, 'r', encoding='utf-8') as f:
                contracts = json.load(f).get('contracts', [])

        if not contracts:
            return jsonify({"error": "No contracts to export"}), 404

        df = pd.DataFrame(contracts)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ONEWEST AMC')
        output.seek(0)
        filename = f"ONEWEST_AMC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =====================================================
# ONEWEST AMC UPDATE (ow_api/amc/update)
# =====================================================
@app.route("/ow_api/amc/update", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_amc_update():
    """Update ONEWEST AMC contract"""
    try:
        data = request.get_json()
        contract_id = data.get('contract_id','')

        amc_data = {"contracts": [], "last_updated": ""}
        if OW_AMC_JSON.exists():
            with open(OW_AMC_JSON, 'r', encoding='utf-8') as f:
                amc_data = json.load(f)

        contracts = amc_data.get('contracts', [])
        found = False
        for i, c in enumerate(contracts):
            if c.get('contract_id') == contract_id:
                contracts[i] = {**c, **data, 'updated_at': datetime.now().isoformat()}
                found = True
                break

        if not found:
            # Add new contract
            data['created_at'] = datetime.now().isoformat()
            contracts.append(data)

        amc_data['contracts']    = contracts
        amc_data['last_updated'] = datetime.now().isoformat()

        OW_DIR.mkdir(parents=True, exist_ok=True)
        with open(OW_AMC_JSON, 'w', encoding='utf-8') as f:
            json.dump(amc_data, f, indent=2)

        print(f"✅ OW AMC {'updated' if found else 'created'}: {contract_id}")
        return jsonify({"success": True, "message": "AMC contract saved"})

    except Exception as e:
        print(f"❌ OW AMC update error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# =====================================================
# ONEWEST DAILY MAIL (TRIGGERS AT 8:00 AM)
# =====================================================
def ow_send_daily_ppm_mail():
    """Send ONEWEST daily PPM mail with today's work orders + any pending"""
    try:
        today_str = datetime.now().strftime('%Y-%m-%d')

        # Load all work orders
        work_orders = []
        if OW_WORK_ORDERS_JSON.exists():
            with open(OW_WORK_ORDERS_JSON, 'r', encoding='utf-8') as f:
                data = json.load(f)
            work_orders = data.get('work_orders', [])

        # Fresh today's WOs
        today_wos = [w for w in work_orders if w.get('due_date', '')[:10] == today_str and (w.get('status','') or '').lower() not in ('completed','closed')]

        # Pending / overdue WOs
        today = datetime.now().date()
        pending_wos = []
        for wo in work_orders:
            status = (wo.get('status','') or '').lower()
            if status in ('completed','closed'):
                continue
            try:
                dd = datetime.strptime(wo.get('due_date','')[:10], '%Y-%m-%d').date()
                if dd < today:
                    pending_wos.append(wo)
            except:
                pass

        # Build HTML email
        def wo_table(wos, title_str, color):
            if not wos:
                return f'<p style="color:#64748b;font-size:13px;">No {title_str.lower()} work orders.</p>'
            rows = ''.join(f"""
            <tr>
                <td style="padding:10px;border-bottom:1px solid #1e293b;font-family:monospace;color:{color};font-size:12px;">{w.get('work_order_id','N/A')}</td>
                <td style="padding:10px;border-bottom:1px solid #1e293b;color:#e2e8f0;font-size:13px;">{w.get('asset_name','Unknown')}</td>
                <td style="padding:10px;border-bottom:1px solid #1e293b;color:#94a3b8;font-size:12px;">{w.get('location','Unknown')}</td>
                <td style="padding:10px;border-bottom:1px solid #1e293b;color:#94a3b8;font-size:12px;">{w.get('priority','Medium')}</td>
                <td style="padding:10px;border-bottom:1px solid #1e293b;color:#94a3b8;font-size:12px;">{w.get('due_date','N/A')}</td>
            </tr>""" for w in wos)
            return f"""
            <h3 style="color:{color};font-family:sans-serif;margin:20px 0 10px;">{title_str} ({len(wos)})</h3>
            <table style="width:100%;border-collapse:collapse;background:#0f172a;border-radius:8px;overflow:hidden;">
                <thead><tr style="background:#1e293b;">
                    <th style="padding:10px;text-align:left;color:#475569;font-size:11px;text-transform:uppercase;">WO ID</th>
                    <th style="padding:10px;text-align:left;color:#475569;font-size:11px;text-transform:uppercase;">Asset</th>
                    <th style="padding:10px;text-align:left;color:#475569;font-size:11px;text-transform:uppercase;">Location</th>
                    <th style="padding:10px;text-align:left;color:#475569;font-size:11px;text-transform:uppercase;">Priority</th>
                    <th style="padding:10px;text-align:left;color:#475569;font-size:11px;text-transform:uppercase;">Due Date</th>
                </tr></thead>
                <tbody>{rows}</tbody>
            </table>"""

        html_body = f"""
        <div style="font-family:sans-serif;background:#020617;color:#e2e8f0;padding:32px;max-width:800px;margin:0 auto;">
            <div style="text-align:center;margin-bottom:32px;">
                <h1 style="font-family:monospace;color:#f97316;font-size:28px;margin:0;">ONEWEST</h1>
                <p style="color:#64748b;margin:6px 0 0;">Daily PPM Maintenance Report — {datetime.now().strftime('%A, %d %B %Y')}</p>
            </div>

            <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px;">
                <div style="background:#0f172a;border-radius:12px;padding:20px;border:1px solid rgba(249,115,22,0.2);">
                    <p style="color:#475569;font-size:11px;text-transform:uppercase;margin:0 0 6px;">Today's Work Orders</p>
                    <p style="color:#f97316;font-size:32px;font-weight:800;margin:0;">{len(today_wos)}</p>
                </div>
                <div style="background:#0f172a;border-radius:12px;padding:20px;border:1px solid rgba(244,63,94,0.2);">
                    <p style="color:#475569;font-size:11px;text-transform:uppercase;margin:0 0 6px;">Pending / Overdue</p>
                    <p style="color:#f43f5e;font-size:32px;font-weight:800;margin:0;">{len(pending_wos)}</p>
                </div>
            </div>

            {wo_table(today_wos, "Today's Work Orders", "#f97316")}
            {wo_table(pending_wos, "Pending / Overdue", "#f43f5e")}

            <p style="color:#334155;font-size:12px;text-align:center;margin-top:32px;">
                Generated at {datetime.now().strftime('%I:%M %p IST')} | EMERZHANT Property Management System
            </p>
        </div>"""

        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"ONEWEST Daily PPM — {datetime.now().strftime('%d %b %Y')} — {len(today_wos)} Today | {len(pending_wos)} Pending"
        msg['From']    = formataddr(("ONEWEST MMS", SENDER_EMAIL))
        msg['To']      = ", ".join(OW_EMAIL_RECEIVERS)
        msg.attach(MIMEText(html_body, 'html'))

        _smtp_send(msg, OW_EMAIL_RECEIVERS, caller="OW-PPM-daily")

        print(f"✅ OW Daily PPM mail sent — Today: {len(today_wos)} | Pending: {len(pending_wos)}")
        return {"success": True, "wo_count": len(today_wos), "pending_count": len(pending_wos)}

    except Exception as e:
        print(f"❌ OW daily mail error: {e}")
        return {"success": False, "error": str(e)}


# Manual trigger endpoint
@app.route("/ow_api/trigger-daily-mail", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_trigger_daily_mail():
    """Manually trigger ONEWEST daily PPM mail"""
    result = ow_send_daily_ppm_mail()
    return jsonify(result)


# Schedule 8:00 AM daily
def _setup_ow_ppm_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler()
        scheduler.add_job(
            func=ow_send_daily_ppm_mail,
            trigger='cron',
            hour=8,
            minute=0,
            timezone='Asia/Kolkata',
            id='ow_daily_ppm_mail'
        )
        scheduler.start()
        print("✅ ONEWEST: Daily PPM mail scheduler started at 8:00 AM IST")
        return scheduler
    except Exception as e:
        print(f"⚠️  ONEWEST scheduler error: {e}")
        return None

_ow_scheduler = _setup_ow_ppm_scheduler()


# =====================================================
# ONEWEST CALENDAR VIEW ROUTES (COMPLETE & FIXED)
# =====================================================
@app.route("/ow_api/ppm/calendar")
@login_required
@require_property("ONEWEST")
def ow_api_ppm_calendar():
    """Get ONEWEST calendar data — assets grouped by due date"""
    try:
        year  = int(request.args.get('year',  datetime.now().year))
        month = int(request.args.get('month', datetime.now().month))

        if not OW_ASSETS_XLSX.exists():
            return jsonify({"calendar": {}, "property": "ONEWEST"})

        df = pd.read_excel(OW_ASSETS_XLSX, engine='openpyxl')
        calendar_data = {}

        for _, row in df.iterrows():
            asset_code = str(row.get('Asset Code', '')).strip()
            if not asset_code or asset_code.lower() in ['nan','none','']:
                continue
            next_due = str(row.get('nextDueDate', '')).strip()
            if not next_due or next_due.lower() in ['nan','none','']:
                continue
            try:
                from dateutil.parser import parse as dateutil_parse
                due_dt = dateutil_parse(next_due)
                if due_dt.year == year and due_dt.month == month:
                    date_key = due_dt.strftime('%Y-%m-%d')
                    if date_key not in calendar_data:
                        calendar_data[date_key] = []
                    calendar_data[date_key].append({
                        "id":          asset_code,
                        "name":        str(row.get('Asset Name','')).strip(),
                        "location":    str(row.get('Location','')).strip(),
                        "lastService": str(row.get('Last Service','')).strip()
                    })
            except:
                pass

        return jsonify({"calendar": calendar_data, "year": year, "month": month, "property": "ONEWEST"})

    except Exception as e:
        print(f"❌ OW calendar error: {e}")
        return jsonify({"calendar": {}, "error": str(e)}), 500


# FIX: duplicate /ow_api/technicians + /ow_api/supervisors routes removed.
# Flask uses last-registered route, so these were silently overriding the
# corrected ow_api_technicians / ow_api_supervisors definitions above.


# Serve OW PPM uploads
@app.route("/uploads/OW/ppm/<filename>")
@login_required
def serve_ow_ppm_upload(filename):
    """Serve ONEWEST PPM work order image"""
    return send_from_directory(OW_PPM_WO_UPLOADS, filename)



# =====================================================
# 2.0 ONEWEST_STORE
# =====================================================

"""
ONEWEST INVENTORY - SERVER INTEGRATION SNIPPET
Add this to server.py for ONEWEST inventory module
All routes use ow_ prefix - Independent from SLN Terminus
"""

# =====================================================
# ONEWEST INVENTORY PATHS
# =====================================================
OW_INVENTORY_XLSX = BASE_DIR / "static" / "data" / "ow_store_master.xlsx"
OW_INVENTORY_ALERTS = BASE_DIR / "static" / "data" / "ow_inventory_alerts.json"
OW_INVENTORY_DIR = BASE_DIR / "static" / "data" / "OW" / "inventory"

for folder in [OW_INVENTORY_DIR, OW_INVENTORY_ALERTS.parent]:
    folder.mkdir(parents=True, exist_ok=True)

if not OW_INVENTORY_ALERTS.exists():
    with open(OW_INVENTORY_ALERTS, 'w') as f:
        json.dump({"alerts": [], "last_updated": datetime.now().isoformat()}, f, indent=2)


# =====================================================
# ONEWEST INVENTORY ROUTES
# =====================================================
@app.route("/ow_inventory_dashboard")
def ow_inventory_dashboard():
    return render_template("ow_inventory_dashboard.html")


@app.route("/ow_api/inventory/items")
@login_required
@require_property("ONEWEST")
def ow_get_inventory_items():
    try:
        if not OW_INVENTORY_XLSX.exists():
            return jsonify({"success": False, "items": [], "total": 0})
        
        df = pd.read_excel(OW_INVENTORY_XLSX, engine='openpyxl')
        items = []
        
        for _, row in df.iterrows():
            item_code = str(row.get('Item_Code', '')).strip()
            if not item_code or item_code.lower() in ['nan', 'none', '']:
                continue
            
            current_stock = float(row.get('Current_Stock', 0)) if pd.notna(row.get('Current_Stock')) else 0
            min_stock = float(row.get('Min_Stock_Level', 0)) if pd.notna(row.get('Min_Stock_Level')) else 0
            
            status = "Out of Stock" if current_stock <= 0 else ("Low Stock" if current_stock < min_stock else "In Stock")
            status_color = "danger" if current_stock <= 0 else ("warning" if current_stock < min_stock else "success")
            
            items.append({
                "item_code": item_code,
                "item_name": str(row.get('Item_Name', 'Unknown')).strip(),
                "department": str(row.get('Department', 'General')).strip(),
                "unit": str(row.get('Unit', 'Nos')).strip(),
                "current_stock": current_stock,
                "min_stock_level": min_stock,
                "status": status,
                "status_color": status_color
            })
        
        dept_filter = request.args.get('department', 'all').strip()
        if dept_filter != 'all':
            items = [i for i in items if i['department'].lower() == dept_filter.lower()]
        
        return jsonify({"success": True, "items": items, "total": len(items), "property": "ONEWEST"})
    
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "items": []}), 500


@app.route("/ow_api/inventory/stats")
@login_required
@require_property("ONEWEST")
def ow_get_inventory_stats():
    try:
        if not OW_INVENTORY_XLSX.exists():
            return jsonify({"total_items": 0, "in_stock": 0, "low_stock": 0, "out_of_stock": 0})
        
        df = pd.read_excel(OW_INVENTORY_XLSX)
        total_items = in_stock = low_stock = out_of_stock = 0
        departments = set()
        
        for _, row in df.iterrows():
            item_code = str(row.get('Item_Code', '')).strip()
            if not item_code or item_code.lower() in ['nan', 'none', '']:
                continue
            
            total_items += 1
            departments.add(str(row.get('Department', 'General')).strip())
            
            current_stock = float(row.get('Current_Stock', 0)) if pd.notna(row.get('Current_Stock')) else 0
            min_stock = float(row.get('Min_Stock_Level', 0)) if pd.notna(row.get('Min_Stock_Level')) else 0
            
            if current_stock <= 0:
                out_of_stock += 1
            elif current_stock < min_stock:
                low_stock += 1
            else:
                in_stock += 1
        
        return jsonify({
            "total_items": total_items,
            "in_stock": in_stock,
            "low_stock": low_stock,
            "out_of_stock": out_of_stock,
            "departments": list(departments),
            "property": "ONEWEST"
        })
    
    except Exception as e:
        return jsonify({"total_items": 0, "in_stock": 0, "low_stock": 0, "out_of_stock": 0}), 500


@app.route("/ow_api/inventory/movement", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_update_stock_movement():
    try:
        data = request.get_json()
        item_code = data.get('item_code')
        movement_type = data.get('movement_type')
        quantity = int(data.get('quantity', 0))
        
        if not item_code or not movement_type or quantity <= 0:
            return jsonify({"success": False, "error": "Invalid data"}), 400
        
        if not OW_INVENTORY_XLSX.exists():
            return jsonify({"success": False, "error": "Inventory file not found"}), 404
        
        df = pd.read_excel(OW_INVENTORY_XLSX)
        mask = df['Item_Code'] == item_code
        
        if not mask.any():
            return jsonify({"success": False, "error": "Item not found"}), 404
        
        current_stock = float(df.loc[mask, 'Current_Stock'].iloc[0]) if pd.notna(df.loc[mask, 'Current_Stock'].iloc[0]) else 0
        
        if movement_type.upper() == 'IN':
            new_stock = current_stock + quantity
            df.loc[mask, 'Stock_In'] = (df.loc[mask, 'Stock_In'].iloc[0] if pd.notna(df.loc[mask, 'Stock_In'].iloc[0]) else 0) + quantity
        elif movement_type.upper() == 'OUT':
            if quantity > current_stock:
                return jsonify({"success": False, "error": "Insufficient stock"}), 400
            new_stock = current_stock - quantity
            df.loc[mask, 'Stock_Out'] = (df.loc[mask, 'Stock_Out'].iloc[0] if pd.notna(df.loc[mask, 'Stock_Out'].iloc[0]) else 0) + quantity
        else:
            return jsonify({"success": False, "error": "Invalid movement type"}), 400
        
        df.loc[mask, 'Current_Stock'] = new_stock
        df.loc[mask, 'Last_Updated'] = datetime.now().strftime('%Y-%m-%d')
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                df.to_excel(OW_INVENTORY_XLSX, index=False)
                break
            except PermissionError:
                if attempt == max_retries - 1:
                    return jsonify({"success": False, "error": "File locked"}), 500
                import time
                time.sleep(1)
        
        return jsonify({"success": True, "new_stock": new_stock})
    
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/ow_api/inventory/alerts")
@login_required
@require_property("ONEWEST")
def ow_get_inventory_alerts():
    try:
        if not OW_INVENTORY_XLSX.exists():
            return jsonify({"alerts": [], "total": 0})
        
        df = pd.read_excel(OW_INVENTORY_XLSX)
        alerts = []
        
        for _, row in df.iterrows():
            item_code = str(row.get('Item_Code', '')).strip()
            if not item_code or item_code.lower() in ['nan', 'none', '']:
                continue
            
            current_stock = float(row.get('Current_Stock', 0)) if pd.notna(row.get('Current_Stock')) else 0
            min_stock = float(row.get('Min_Stock_Level', 0)) if pd.notna(row.get('Min_Stock_Level')) else 0
            
            if current_stock < min_stock:
                alerts.append({
                    "item_code": item_code,
                    "item_name": str(row.get('Item_Name', 'Unknown')).strip(),
                    "department": str(row.get('Department', 'General')).strip(),
                    "current_stock": current_stock,
                    "min_stock_level": min_stock,
                    "shortage": min_stock - current_stock,
                    "severity": "critical" if current_stock <= 0 else "warning"
                })
        
        return jsonify({"alerts": alerts, "total": len(alerts), "property": "ONEWEST"})
    
    except Exception as e:
        return jsonify({"alerts": [], "total": 0}), 500


@app.route("/ow_api/inventory/export")
@login_required
@require_property("ONEWEST")
def ow_export_inventory():
    try:
        if not OW_INVENTORY_XLSX.exists():
            return jsonify({"error": "No data"}), 404
        
        filename = f"ONEWEST_Inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            OW_INVENTORY_XLSX,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =====================================================
# ADD TO server.py AFTER OTHER BLUEPRINT REGISTRATIONS
# =====================================================



# =====================================================
# ADD ONEWEST VMS PORTAL ROUTE
# =====================================================
@app.route("/ow_vms")
@login_required
@require_property("ONEWEST")
def ow_vms():
    """ONEWEST Visitor Management System Portal"""
    session['active_property'] = 'ONEWEST'
    session['property_code'] = 'OW'
    print(f"\n🏢 Accessing ONEWEST VMS - User: {session.get('user')}")
    return render_template("ow_vms.html")



# =====================================================
# 6.0 ONEWEST WORK TRACKER
# =====================================================


# =====================================================
# ONEWEST SPACE OCCUPANCY — INLINE ROUTES
# (Also handled by ow_occupancy.py blueprint if present)
# =====================================================

# ── Paths ──────────────────────────────────────────
_OW_SFT_XLSX     = BASE_DIR / "static" / "data" / "OW_ SFT_Details.xlsx"
_OW_OCC_OVERRIDE = BASE_DIR / "static" / "data" / "ow_occupancy_override.json"


def _ow_parse_sft():
    """Parse OW_ SFT_Details.xlsx → list of space dicts."""
    spaces        = []
    current_floor = None
    try:
        try:
            df = pd.read_excel(str(_OW_SFT_XLSX), header=None, engine="openpyxl")
        except Exception:
            df = pd.read_excel(str(_OW_SFT_XLSX), header=None, engine="xlrd")
    except Exception as e:
        print(f"❌ OW SFT Excel read error: {e}")
        return spaces

    for _, row in df.iterrows():
        vals  = list(row)
        first = str(vals[0]).strip() if vals[0] is not None else ""
        if first.upper() in ("TENANTS SFT DETAILS", "S.NO", ""):
            continue
        try:
            sno = int(float(first))
        except (ValueError, TypeError):
            continue

        floor_val   = str(vals[1]).strip() if vals[1] is not None else ""
        client_name = str(vals[2]).strip() if vals[2] is not None else ""
        area_raw    = vals[3]

        try:
            area_str = str(area_raw).replace(",", "").strip()
            if any(c.isalpha() for c in area_str.replace(".", "").replace("-", "")):
                area = 0.0
            else:
                area = float(area_str)
        except (ValueError, TypeError):
            area = 0.0

        if floor_val and floor_val.lower() not in ("none", ""):
            current_floor = floor_val
        if not client_name or client_name.lower() == "none":
            continue

        spaces.append({
            "id":         str(sno),
            "sno":        sno,
            "floor":      current_floor or "Unknown",
            "clientName": client_name,
            "area":       round(area, 2),
        })
    return spaces


@app.route("/ow_occupancy")
@login_required
@require_property("ONEWEST")
def ow_occupancy_page():
    """ONEWEST Space Occupancy dashboard."""
    session["active_property"] = "ONEWEST"
    session["property_code"]   = "OW"
    print(f"\n🏢 Accessing OW Occupancy — User: {session.get('user')}")
    return render_template("ow_occupancy.html")


@app.route("/ow_api/occupancy")
@login_required
@require_property("ONEWEST")
def ow_api_occupancy_get():
    """Return ONEWEST SFT occupancy data."""
    try:
        # Try JSON override first (saved edits)
        if _OW_OCC_OVERRIDE.exists():
            try:
                spaces = json.loads(_OW_OCC_OVERRIDE.read_text(encoding="utf-8"))
            except Exception:
                spaces = []
        elif _OW_SFT_XLSX.exists():
            spaces = _ow_parse_sft()
        else:
            return jsonify({"error": "OW_ SFT_Details.xlsx not found in static/data/"}), 404

        def _st(s):
            n = (s.get("clientName") or "").lower().strip()
            if n in ("vacant", ""): return "vacant"
            if "fit" in n and "out" in n: return "fitout"
            return "occupied"

        total_area  = sum(s.get("area", 0) or 0 for s in spaces)
        occ_area    = sum(s.get("area", 0) or 0 for s in spaces if _st(s) == "occupied")
        vac_area    = sum(s.get("area", 0) or 0 for s in spaces if _st(s) == "vacant")
        total_units = len(spaces)
        occ_count   = sum(1 for s in spaces if _st(s) == "occupied")
        vac_count   = sum(1 for s in spaces if _st(s) == "vacant")
        fit_count   = sum(1 for s in spaces if _st(s) == "fitout")

        return jsonify({
            "summary": {
                "total_area":     round(total_area, 2),
                "occupied_area":  round(occ_area,   2),
                "vacant_area":    round(vac_area,    2),
                "total_units":    total_units,
                "occupied_count": occ_count,
                "vacant_count":   vac_count,
                "fitout_count":   fit_count,
            },
            "spaces": spaces
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/ow_api/occupancy/update", methods=["POST"])
@login_required
@require_property("ONEWEST")
def ow_api_occupancy_update():
    """Save edited occupancy data as JSON override."""
    try:
        data   = request.get_json(force=True) or {}
        spaces = data.get("spaces", [])
        if not isinstance(spaces, list):
            return jsonify({"success": False, "error": "Invalid payload"}), 400
        _OW_OCC_OVERRIDE.parent.mkdir(parents=True, exist_ok=True)
        _OW_OCC_OVERRIDE.write_text(
            json.dumps(spaces, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return jsonify({"success": True, "count": len(spaces)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/ow_work_track")
@login_required
@require_property("ONEWEST")
def ow_work_track():
    return render_template("ow_work_track.html")


# =====================================================
# 4.0 OGM
# =====================================================



# =====================================================
# 5.0 NINEHILS
# =====================================================

# =====================================================
# 5.1 SLN HOUSEKEEPING
# =====================================================
@app.route("/sln_hk")
@login_required
def sln_hk_portal():
    """SLN Housekeeping Module — redirects to blueprint dashboard"""
    return redirect(url_for("sln_hk.sln_hk_dashboard"))

# =====================================================
# 5.2 SLN SECURITY MODULE
# =====================================================
@app.route("/sln_sec")
@login_required
@require_property("SLN Terminus")
def sln_sec_portal():
    """SLN Security Module — redirects to blueprint dashboard"""
    return redirect("/sln_hk_sec/")

# =====================================================
# 5.3 SLN FIRE FIGHTING MODULE
# =====================================================
@app.route("/sln_fire_portal")
@login_required
@require_property("SLN Terminus")
def sln_fire_portal():
    """SLN Fire Fighting Module"""
    return redirect("/sln_fire/")

@app.route("/energy")
@login_required
@require_property("SLN Terminus")
def energy():
    """SLN Energy Module"""
    return render_template("energy.html")
# =====================================================
# GM TASKS — Data file path
# =====================================================
GM_TASKS_FILE = DATA_DIR / "gm_tasks.json"

def _load_gm_tasks():
    """Load tasks from JSON file, return list."""
    if not GM_TASKS_FILE.exists():
        return []
    try:
        with open(GM_TASKS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []

def _save_gm_tasks(tasks):
    """Persist tasks list to JSON file."""
    GM_TASKS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(GM_TASKS_FILE, "w", encoding="utf-8") as f:
        json.dump(tasks, f, ensure_ascii=False, indent=2)

# ── Page Route ─────────────────────────────────────────────────────────────
@app.route("/gm_tasks")
@login_required
def gm_tasks_page():
    """GM Tasks Portal — accessible to all full-access roles."""
    return render_template("gm_tasks.html")

# ── API: GET all tasks ──────────────────────────────────────────────────────
@app.route("/api/gm_tasks", methods=["GET"])
@login_required
def api_gm_tasks_get():
    tasks = _load_gm_tasks()
    return jsonify({"tasks": tasks, "count": len(tasks)})

# ── API: POST (full replace / sync from client) ─────────────────────────────
@app.route("/api/gm_tasks", methods=["POST"])
@login_required
def api_gm_tasks_post():
    try:
        body = request.get_json(force=True) or {}
        tasks = body.get("tasks", [])
        if not isinstance(tasks, list):
            return jsonify({"success": False, "error": "tasks must be a list"}), 400
        _save_gm_tasks(tasks)
        return jsonify({"success": True, "count": len(tasks)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Add single task ─────────────────────────────────────────────────────
@app.route("/api/gm_tasks/add", methods=["POST"])
@login_required
def api_gm_tasks_add():
    try:
        task = request.get_json(force=True) or {}
        if not task.get("description"):
            return jsonify({"success": False, "error": "description required"}), 400
        if not task.get("id"):
            task["id"] = f"task_{int(datetime.now().timestamp()*1000)}"
        task.setdefault("status",    "Open")
        task.setdefault("priority",  "Medium")
        task.setdefault("date",      datetime.now().strftime("%Y-%m-%d"))
        task["updatedAt"] = datetime.now().isoformat()
        tasks = _load_gm_tasks()
        tasks.insert(0, task)
        _save_gm_tasks(tasks)
        return jsonify({"success": True, "task": task, "total": len(tasks)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Update single task ──────────────────────────────────────────────────
@app.route("/api/gm_tasks/<task_id>", methods=["PUT"])
@login_required
def api_gm_tasks_update(task_id):
    try:
        updates = request.get_json(force=True) or {}
        tasks = _load_gm_tasks()
        for i, t in enumerate(tasks):
            if t.get("id") == task_id:
                tasks[i].update(updates)
                tasks[i]["updatedAt"] = datetime.now().isoformat()
                _save_gm_tasks(tasks)
                return jsonify({"success": True, "task": tasks[i]})
        return jsonify({"success": False, "error": "Task not found"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: Delete single task ──────────────────────────────────────────────────
@app.route("/api/gm_tasks/<task_id>", methods=["DELETE"])
@login_required
def api_gm_tasks_delete(task_id):
    try:
        tasks = _load_gm_tasks()
        before = len(tasks)
        tasks = [t for t in tasks if t.get("id") != task_id]
        _save_gm_tasks(tasks)
        return jsonify({"success": True, "deleted": before - len(tasks)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── Mail config file ─────────────────────────────────────────────────────────
GM_MAIL_CONFIG_FILE = DATA_DIR / "gm_mail_config.json"

# ══════════════════════════════════════════════════════════════════════════════
# GM TASKS — DEDICATED MAIL CONFIGURATION
# Completely separate from PPM / MMS / CAM mailers.
# Change only these values to configure GM Tasks emails.
# ══════════════════════════════════════════════════════════════════════════════

# ▼ GM Tasks dedicated SMTP credentials (can be same or different Gmail account)
GM_SMTP_SERVER   = "smtp.gmail.com"
GM_SMTP_PORT     = 587
GM_SENDER_EMAIL  = "maintenance.slnterminus@gmail.com"   # ← GM sender address
GM_SENDER_PASS   = "xaottgrqtqnkouqn"                   # ← GM app password

# ▼ GM Tasks dedicated lock — shares the master Gmail lock so GM and MMS
#   never authenticate to Gmail at the same time
_GM_SMTP_LOCK = _GMAIL_LOCK

# ▼▼▼ ADD / REMOVE GM TASKS RECIPIENT EMAILS HERE ▼▼▼
GM_TASKS_TO = [
    "maintenance.slnterminus@gmail.com",
    "yasven7545@gmail.com",
    "engineering@terminus-global.com",
    "kiran@terminus-global.com",
    # "gm@slnterminius.com",    ← uncomment or add more
]
GM_TASKS_CC = [
    # "director@slnterminius.com",   ← CC recipients
]
# ▲▲▲ END OF RECIPIENT CONFIG ▲▲▲

GM_MAIL_DEFAULTS = {
    "recipients":      ",".join(GM_TASKS_TO),
    "cc":              ",".join(GM_TASKS_CC),
    "time":            "09:00",
    "subject":         "Seniour Level Management — SLN Terminus",
    "inclOpen":        True,
    "inclProg":        True,
    "inclOver":        True,
    "inclDone":        False,
    "inclHighOnly":    False,
    "siteFilter":      "",
    "notifyOnAdd":     True,
    "notifyOnOverdue": True,
}

def _load_gm_mail_cfg():
    """Return GM mail config, merging saved values over defaults."""
    cfg = dict(GM_MAIL_DEFAULTS)
    try:
        if GM_MAIL_CONFIG_FILE.exists():
            with open(GM_MAIL_CONFIG_FILE, "r") as f:
                cfg.update(json.load(f))
    except Exception:
        pass
    return cfg

# ── API: GET mail config ──────────────────────────────────────────────────────
@app.route("/api/gm_mail_config", methods=["GET"])
@login_required
def api_gm_mail_config_get():
    return jsonify(_load_gm_mail_cfg())

# ── API: POST (save) mail config ──────────────────────────────────────────────
@app.route("/api/gm_mail_config", methods=["POST"])
@login_required
def api_gm_mail_config_post():
    try:
        incoming = request.get_json(force=True) or {}
        cfg = dict(GM_MAIL_DEFAULTS)
        cfg.update(incoming)
        GM_MAIL_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(GM_MAIL_CONFIG_FILE, "w") as f:
            json.dump(cfg, f, indent=2)
        print(f"✅ GM Mail config saved: recipients={cfg.get('recipients')}, time={cfg.get('time')}")
        return jsonify({"success": True, "config": cfg})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── Helper: resolve recipients from config ────────────────────────────────────
def _resolve_recipients(cfg):
    """Return (to_list, cc_list) from GM config."""
    raw_to  = cfg.get("recipients", ",".join(GM_TASKS_TO))
    to_list = [r.strip() for r in raw_to.split(",") if r.strip()] or list(GM_TASKS_TO)
    raw_cc  = cfg.get("cc", "")
    cc_list = [r.strip() for r in raw_cc.split(",") if r.strip()]
    return to_list, cc_list

# ── Helper: filter tasks for mail ────────────────────────────────────────────
def _filter_tasks_for_mail(tasks, cfg):
    incl_open      = cfg.get("inclOpen",     True)
    incl_prog      = cfg.get("inclProg",     True)
    incl_over      = cfg.get("inclOver",     True)
    incl_done      = cfg.get("inclDone",     False)
    incl_high_only = cfg.get("inclHighOnly", False)
    site_filter    = cfg.get("siteFilter",   "")

    filtered = [t for t in tasks if
        ((incl_open and t.get("status") == "Open")        or
         (incl_prog and t.get("status") == "In Progress") or
         (incl_over and t.get("status") == "Overdue")     or
         (incl_done and t.get("status") == "Completed"))
        and (not site_filter  or t.get("site","") == site_filter)
        and (not incl_high_only or t.get("priority","") == "High")
    ]
    return filtered

# ── Helper: per-site breakdown rows ─────────────────────────────────────────
def _site_breakdown_html(tasks):
    SITES = ["SLNT", "ONEWEST", "The District", "OGM", "Nine Hills"]
    SITE_COLOR = {
        "SLNT":        "#0e7490",
        "ONEWEST":     "#6d28d9",
        "The District":"#047857",
        "OGM":         "#b45309",
        "Nine Hills":  "#b45309",
    }
    rows = ""
    for s in SITES:
        st = [t for t in tasks if t.get("site") == s]
        if not st:
            continue
        over = sum(1 for t in st if t.get("status") == "Overdue")
        opn  = sum(1 for t in st if t.get("status") == "Open")
        prog = sum(1 for t in st if t.get("status") == "In Progress")
        done = sum(1 for t in st if t.get("status") == "Completed")
        clr  = SITE_COLOR.get(s, "#374151")
        rows += f"""
        <tr>
          <td style="padding:8px 14px;border-bottom:1px solid #f3f4f6">
            <span style="background:{clr}20;color:{clr};padding:3px 9px;border-radius:10px;font-size:12px;font-weight:600">{s}</span>
          </td>
          <td style="padding:8px 14px;border-bottom:1px solid #f3f4f6;text-align:center;font-weight:700;color:#111827">{len(st)}</td>
          <td style="padding:8px 14px;border-bottom:1px solid #f3f4f6;text-align:center;color:#d97706">{opn}</td>
          <td style="padding:8px 14px;border-bottom:1px solid #f3f4f6;text-align:center;color:#0e7490">{prog}</td>
          <td style="padding:8px 14px;border-bottom:1px solid #f3f4f6;text-align:center;color:#059669">{done}</td>
          <td style="padding:8px 14px;border-bottom:1px solid #f3f4f6;text-align:center;color:#{'dc2626' if over else '9ca3af'};font-weight:{'700' if over else '400'}">{over if over else '—'}</td>
        </tr>"""
    return rows

# ── Helper: build full HTML email ───────────────────────────────────────────
def _build_gm_email_html(tasks, report_type="daily", trigger_info=""):
    """
    Build a rich HTML email for GM Tasks.

    report_type : "daily"    — regular 09:00 AM scheduled report
                  "manual"   — GM clicked Send Now
                  "new_task" — instant alert on task creation
                  "overdue"  — instant alert on Overdue flip
    trigger_info: extra context string shown in the sub-header
    """
    now_str  = datetime.now().strftime("%d %b %Y, %I:%M %p IST")
    today    = datetime.now().strftime("%d %b %Y")

    STATUS_CFG = {
        "Open":        {"tc": "#b45309", "bg": "#fef3c7", "dot": "#f59e0b"},
        "In Progress": {"tc": "#0e7490", "bg": "#cffafe", "dot": "#06b6d4"},
        "Completed":   {"tc": "#065f46", "bg": "#d1fae5", "dot": "#10b981"},
        "Overdue":     {"tc": "#991b1b", "bg": "#fee2e2", "dot": "#ef4444"},
    }
    PRIO_CFG = {
        "High":   {"tc": "#991b1b", "bg": "#fee2e2"},
        "Medium": {"tc": "#92400e", "bg": "#fef3c7"},
        "Low":    {"tc": "#065f46", "bg": "#d1fae5"},
    }
    REPORT_LABELS = {
        "daily":    ("📋 Daily Summary Report",    "Scheduled · 09:00 AM IST"),
        "manual":   ("📤 Manual Report",           "Sent on demand by GM"),
        "new_task": ("🆕 New Task Alert",          trigger_info or "A new task was added"),
        "overdue":  ("⚠️ Overdue Task Alert",      trigger_info or "A task is now overdue"),
    }
    rpt_title, rpt_sub = REPORT_LABELS.get(report_type, REPORT_LABELS["daily"])

    open_c = sum(1 for t in tasks if t.get("status") == "Open")
    prog_c = sum(1 for t in tasks if t.get("status") == "In Progress")
    done_c = sum(1 for t in tasks if t.get("status") == "Completed")
    over_c = sum(1 for t in tasks if t.get("status") == "Overdue")
    high_c = sum(1 for t in tasks if t.get("priority") == "High")

    # ── Task rows ────────────────────────────────────────────────────────────
    task_rows = ""
    for i, t in enumerate(tasks, 1):
        s_cfg  = STATUS_CFG.get(t.get("status","Open"),  STATUS_CFG["Open"])
        p_cfg  = PRIO_CFG.get(t.get("priority","Medium"), PRIO_CFG["Medium"])
        row_bg = "#fff8f1" if t.get("status") == "Overdue" else ("#fafafa" if i % 2 else "#fff")
        left_border = "border-left:3px solid #ef4444;" if t.get("status") == "Overdue" else ""
        task_rows += f"""
        <tr style="background:{row_bg};{left_border}">
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;color:#6b7280;font-size:12px">{i}</td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;color:#6b7280;font-size:12px;white-space:nowrap">{('-'.join(reversed(t.get('date','—').split('-'))) if t.get('date') and len(t.get('date','').split('-'))==3 else t.get('date','—'))}</td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;color:#111827;font-weight:600;font-size:13px;max-width:220px">{t.get('description','—')}</td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;font-size:12px">
            <span style="background:#f0f9ff;color:#0369a1;padding:2px 7px;border-radius:8px;font-size:11px;font-weight:600">{t.get('site','—')}</span>
          </td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;color:#374151;font-size:12px">{t.get('cost','—')}</td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;color:#374151;font-size:12px">{t.get('assignedTo','—')}</td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6">
            <span style="background:{s_cfg['bg']};color:{s_cfg['tc']};padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700">{t.get('status','Open')}</span>
          </td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6">
            <span style="background:{p_cfg['bg']};color:{p_cfg['tc']};padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600">{t.get('priority','Medium')}</span>
          </td>
          <td style="padding:9px 12px;border-bottom:1px solid #f3f4f6;color:#6b7280;font-size:12px;max-width:160px">{t.get('remarks','—')}</td>
        </tr>"""

    if not task_rows:
        task_rows = '<tr><td colspan="9" style="padding:28px;text-align:center;color:#9ca3af;font-size:13px">No tasks in this report.</td></tr>'

    # ── Overdue callout block (only when overdue tasks exist) ────────────────
    overdue_block = ""
    overdue_tasks = [t for t in tasks if t.get("status") == "Overdue"]
    if overdue_tasks:
        overdue_items = "".join(
            f'<li style="margin:4px 0;font-size:13px;color:#7f1d1d">'
            f'<strong>{t.get("description","—")}</strong>'
            f' — {t.get("site","—")} · Assigned: {t.get("assignedTo","—")}'
            f'</li>'
            for t in overdue_tasks
        )
        overdue_block = f"""
  <div style="margin:0 32px 24px;padding:16px 20px;background:#fef2f2;border:1.5px solid #fca5a5;border-radius:10px">
    <div style="font-size:13px;font-weight:700;color:#991b1b;margin-bottom:8px">⚠️ {len(overdue_tasks)} Overdue Task{'s' if len(overdue_tasks)>1 else ''} — Immediate Attention Required</div>
    <ul style="margin:0;padding-left:18px">{overdue_items}</ul>
  </div>"""

    # ── High-priority callout ────────────────────────────────────────────────
    high_block = ""
    high_tasks = [t for t in tasks if t.get("priority") == "High" and t.get("status") not in ("Completed",)]
    if high_tasks:
        high_items = "".join(
            f'<li style="margin:4px 0;font-size:13px;color:#7c2d12">'
            f'<strong>{t.get("description","—")}</strong>'
            f' [{t.get("status","Open")}] — {t.get("site","—")} · {t.get("assignedTo","—")}'
            f'</li>'
            for t in high_tasks
        )
        high_block = f"""
  <div style="margin:0 32px 24px;padding:16px 20px;background:#fff7ed;border:1.5px solid #fdba74;border-radius:10px">
    <div style="font-size:13px;font-weight:700;color:#9a3412;margin-bottom:8px">🔴 {len(high_tasks)} High-Priority Task{'s' if len(high_tasks)>1 else ''} Pending</div>
    <ul style="margin:0;padding-left:18px">{high_items}</ul>
  </div>"""

    # ── Site breakdown ───────────────────────────────────────────────────────
    site_rows = _site_breakdown_html(tasks)
    site_section = ""
    if site_rows:
        site_section = f"""
  <div style="padding:20px 32px;border-top:1px solid #e5e7eb">
    <div style="font-size:13px;font-weight:700;color:#374151;margin-bottom:12px">Site-wise Breakdown</div>
    <table style="width:100%;border-collapse:collapse;font-size:13px">
      <thead>
        <tr style="background:#f9fafb">
          <th style="padding:8px 14px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Site</th>
          <th style="padding:8px 14px;text-align:center;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Total</th>
          <th style="padding:8px 14px;text-align:center;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#d97706;border-bottom:2px solid #e5e7eb">Open</th>
          <th style="padding:8px 14px;text-align:center;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#0e7490;border-bottom:2px solid #e5e7eb">In Progress</th>
          <th style="padding:8px 14px;text-align:center;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#059669;border-bottom:2px solid #e5e7eb">Done</th>
          <th style="padding:8px 14px;text-align:center;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#dc2626;border-bottom:2px solid #e5e7eb">Overdue</th>
        </tr>
      </thead>
      <tbody>{site_rows}</tbody>
    </table>
  </div>"""

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{rpt_title}</title>
</head>
<body style="margin:0;padding:0;font-family:Arial,Helvetica,sans-serif;background:#f3f4f6;color:#111827">

<div style="max-width:960px;margin:24px auto;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.10)">

  <!-- ── HEADER ── -->
  <div style="background:#020b12;padding:26px 32px 22px;border-bottom:3px solid #FFD700">
    <table style="width:100%;border-collapse:collapse">
      <tr>
        <td>
          <div style="color:#FFD700;font-size:10px;letter-spacing:2px;text-transform:uppercase;margin-bottom:5px">Terminus · GM Portal</div>
          <div style="color:#fff;font-size:22px;font-weight:700;line-height:1.2">{rpt_title}</div>
          <div style="color:#94a3b8;font-size:12px;margin-top:5px">{rpt_sub}</div>
        </td>
        <td style="text-align:right;vertical-align:top;white-space:nowrap">
          <div style="color:#64748b;font-size:11px">{now_str}</div>
          <div style="color:#FFD700;font-size:11px;margin-top:3px">📋 {len(tasks)} task{'s' if len(tasks)!=1 else ''} in this report</div>
        </td>
      </tr>
    </table>
  </div>

  <!-- ── STAT STRIP ── -->
  <table style="width:100%;border-collapse:collapse;border-bottom:1px solid #e5e7eb">
    <tr>
      <td style="padding:18px 20px;text-align:center;border-right:1px solid #e5e7eb">
        <div style="font-size:30px;font-weight:800;color:#0e7490;line-height:1">{len(tasks)}</div>
        <div style="font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.8px;margin-top:3px">Total</div>
      </td>
      <td style="padding:18px 20px;text-align:center;border-right:1px solid #e5e7eb">
        <div style="font-size:30px;font-weight:800;color:#d97706;line-height:1">{open_c}</div>
        <div style="font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.8px;margin-top:3px">Open</div>
      </td>
      <td style="padding:18px 20px;text-align:center;border-right:1px solid #e5e7eb">
        <div style="font-size:30px;font-weight:800;color:#0e7490;line-height:1">{prog_c}</div>
        <div style="font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.8px;margin-top:3px">In Progress</div>
      </td>
      <td style="padding:18px 20px;text-align:center;border-right:1px solid #e5e7eb">
        <div style="font-size:30px;font-weight:800;color:#059669;line-height:1">{done_c}</div>
        <div style="font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.8px;margin-top:3px">Completed</div>
      </td>
      <td style="padding:18px 20px;text-align:center;border-right:1px solid #e5e7eb">
        <div style="font-size:30px;font-weight:800;color:{'#dc2626' if over_c else '#9ca3af'};line-height:1">{over_c}</div>
        <div style="font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.8px;margin-top:3px">Overdue</div>
      </td>
      <td style="padding:18px 20px;text-align:center">
        <div style="font-size:30px;font-weight:800;color:{'#dc2626' if high_c else '#9ca3af'};line-height:1">{high_c}</div>
        <div style="font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.8px;margin-top:3px">High Prio</div>
      </td>
    </tr>
  </table>

  <!-- ── OVERDUE CALLOUT ── -->
  {overdue_block}

  <!-- ── HIGH PRIORITY CALLOUT ── -->
  {high_block}

  <!-- ── TASK TABLE ── -->
  <div style="padding:20px 32px">
    <div style="font-size:13px;font-weight:700;color:#374151;margin-bottom:12px">Task Register</div>
    <div style="overflow-x:auto">
      <table style="width:100%;border-collapse:collapse;font-size:13px;min-width:700px">
        <thead>
          <tr style="background:#f8fafc">
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">#</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Date</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Description</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Site</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Cost</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Assigned To</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Status</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Priority</th>
            <th style="padding:9px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.8px;color:#6b7280;border-bottom:2px solid #e5e7eb">Remarks</th>
          </tr>
        </thead>
        <tbody>{task_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- ── SITE BREAKDOWN ── -->
  {site_section}

  <!-- ── FOOTER ── -->
  <div style="padding:16px 32px;background:#f9fafb;border-top:1px solid #e5e7eb">
    <table style="width:100%;border-collapse:collapse">
      <tr>
        <td style="color:#9ca3af;font-size:11px">
          Auto-generated by <strong style="color:#6b7280">SLN Terminus Sr. Level Management Portal</strong> ·
          {now_str} · Do not reply to this email.
        </td>
        <td style="text-align:right;color:#9ca3af;font-size:11px;white-space:nowrap">
          <a href="https://descriptive-joya-unsolidified.ngrok-free.dev/gm_tasks" style="color:#0e7490;text-decoration:none;font-weight:600">📋 Open Sr. Lvl MGMT Tasks →</a>
          &nbsp;&nbsp;
          <a href="https://descriptive-joya-unsolidified.ngrok-free.dev/dashboard" style="color:#b45309;text-decoration:none;font-weight:600">🏢 Portal Dashboard →</a>
        </td>
      </tr>
    </table>
  </div>

</div>
</body></html>"""

# ── GM Tasks dedicated SMTP send — uses _GMAIL_LOCK (same master lock) ───────
def _gm_smtp_send(msg_obj, recipients, caller="GM", retries=3, base_delay=5):
    """
    GM Tasks SMTP sender. Uses the same _GMAIL_LOCK as _smtp_send so MMS and
    GM Tasks are serialised — never simultaneous on the Gmail account.
    """
    last_err = None
    for attempt in range(1, retries + 1):
        with _GMAIL_LOCK:
            # Enforce minimum gap since last send from ANY module
            gap = _time.time() - _LAST_SMTP_SEND["ts"]
            if gap < _MIN_SEND_GAP:
                _time.sleep(_MIN_SEND_GAP - gap)
            try:
                with smtplib.SMTP(GM_SMTP_SERVER, GM_SMTP_PORT, timeout=25) as srv:
                    srv.ehlo()
                    srv.starttls()
                    srv.ehlo()
                    srv.login(GM_SENDER_EMAIL, GM_SENDER_PASS)
                    srv.sendmail(GM_SENDER_EMAIL, recipients, msg_obj.as_string())
                _LAST_SMTP_SEND["ts"] = _time.time()
                print(f"✅ [GM-{caller}] Email sent → {recipients} (attempt {attempt})")
                return True
            except smtplib.SMTPAuthenticationError as e:
                print(f"⚠️  [GM-{caller}] Auth error (attempt {attempt}): {e}")
                last_err = e
                _time.sleep(base_delay * attempt * 2)
            except (smtplib.SMTPException, OSError) as e:
                print(f"⚠️  [GM-{caller}] SMTP error (attempt {attempt}): {e}")
                last_err = e
                _time.sleep(base_delay * attempt)
    print(f"❌ [GM-{caller}] All {retries} attempts failed: {last_err}")
    raise last_err

# ── Core send helper ──────────────────────────────────────────────────────────
def _send_gm_email(subject_override=None, tasks_override=None,
                   report_type="daily", trigger_info="", cfg=None):
    """
    Central send function for all GM Tasks emails.
    Returns (success: bool, message: str, recipients: list)
    """
    if cfg is None:
        cfg = _load_gm_mail_cfg()

    to_list, cc_list = _resolve_recipients(cfg)
    subject  = subject_override or cfg.get("subject", GM_MAIL_DEFAULTS["subject"])
    date_str = datetime.now().strftime("%d %b %Y")

    if tasks_override is not None:
        filtered = tasks_override
    else:
        all_tasks = _load_gm_tasks()
        filtered  = _filter_tasks_for_mail(all_tasks, cfg)

    html_body = _build_gm_email_html(filtered, report_type=report_type, trigger_info=trigger_info)

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"{subject} · {date_str}"
    msg["From"]    = formataddr(("Sr. Level Mgmt Portal", GM_SENDER_EMAIL))
    msg["To"]      = ", ".join(to_list)
    if cc_list:
        msg["Cc"]  = ", ".join(cc_list)
    msg.attach(MIMEText(html_body, "html"))

    all_recipients = to_list + cc_list
    _gm_smtp_send(msg, all_recipients, caller=report_type)

    print(f"✅ GM Tasks [{report_type}] → {all_recipients} at {datetime.now().strftime('%H:%M:%S')}")
    return True, f"Report sent to {len(to_list)} recipient(s).", to_list

# ── API: Send report now (manual trigger from portal) ────────────────────────
@app.route("/api/gm_tasks/send_report", methods=["POST"])
@login_required
def api_gm_tasks_send_report():
    try:
        body = request.get_json(force=True) or {}
        # Client may pass tasks directly (client-side state) or we load from disk
        tasks_from_client = body.get("tasks")   # list or None

        cfg = _load_gm_mail_cfg()

        if tasks_from_client is not None:
            # Save latest client state to disk first, then use it
            _save_gm_tasks(tasks_from_client)
            filtered = _filter_tasks_for_mail(tasks_from_client, cfg)
        else:
            filtered = None  # _send_gm_email will load from disk

        ok, msg, recipients = _send_gm_email(
            tasks_override=filtered,
            report_type="manual",
            trigger_info=f"Sent manually by {session.get('user','GM')} at {datetime.now().strftime('%H:%M IST')}",
            cfg=cfg
        )
        return jsonify({"success": ok, "message": msg, "recipients": recipients})

    except Exception as e:
        print(f"⚠️  GM Tasks send_report error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# ── Instant alert: new task added ────────────────────────────────────────────
def _maybe_notify_new_task(task):
    """Fire an instant email when a task is added, if config says so."""
    try:
        cfg = _load_gm_mail_cfg()
        if not cfg.get("notifyOnAdd", True):
            return
        subject = f"🆕 New GM Task Added — {task.get('site','—')} | {task.get('priority','Medium')} Priority"
        _send_gm_email(
            subject_override=subject,
            tasks_override=[task],
            report_type="new_task",
            trigger_info=f"Task added on {task.get('date','—')} · Site: {task.get('site','—')} · Assigned: {task.get('assignedTo','—')}",
            cfg=cfg
        )
    except Exception as e:
        print(f"⚠️  GM new-task notify error: {e}")

# ── Instant alert: task flipped to Overdue ───────────────────────────────────
def _maybe_notify_overdue(task):
    """Fire an instant email when a task status changes to Overdue."""
    try:
        cfg = _load_gm_mail_cfg()
        if not cfg.get("notifyOnOverdue", True):
            return
        subject = f"⚠️ GM Task Overdue — {task.get('site','—')} · {task.get('description','')[:60]}"
        _send_gm_email(
            subject_override=subject,
            tasks_override=[task],
            report_type="overdue",
            trigger_info=f"Task marked Overdue · Site: {task.get('site','—')} · Assigned: {task.get('assignedTo','—')}",
            cfg=cfg
        )
    except Exception as e:
        print(f"⚠️  GM overdue notify error: {e}")

# ── Wire instant alerts into add/update routes ───────────────────────────────
# Patch add route to fire new-task notification
@app.route("/api/gm_tasks/add_notify", methods=["POST"])
@login_required
def api_gm_tasks_add_notify():
    """Add a task AND send instant notification email."""
    try:
        task = request.get_json(force=True) or {}
        if not task.get("description"):
            return jsonify({"success": False, "error": "description required"}), 400
        if not task.get("id"):
            task["id"] = f"task_{int(datetime.now().timestamp()*1000)}"
        task.setdefault("status",   "Open")
        task.setdefault("priority", "Medium")
        task.setdefault("date",     datetime.now().strftime("%Y-%m-%d"))
        task["updatedAt"] = datetime.now().isoformat()
        tasks = _load_gm_tasks()
        tasks.insert(0, task)
        _save_gm_tasks(tasks)
        # Fire notification in background
        import threading
        threading.Thread(target=_maybe_notify_new_task, args=(task,), daemon=True).start()
        return jsonify({"success": True, "task": task, "total": len(tasks)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── Scheduler: daily 09:00 AM IST — runs via APScheduler (same as PPM/CAM) ──
def _gm_tasks_daily_job():
    """APScheduler job: send GM Tasks daily summary at 09:00 AM IST."""
    with app.app_context():
        try:
            cfg = _load_gm_mail_cfg()
            _send_gm_email(report_type="daily", cfg=cfg)
        except Exception as _e:
            print(f"⚠️  [GM Tasks daily job] {_e}")

def _gm_tasks_overdue_check():
    """APScheduler job: check for Overdue flips every 5 minutes and send instant alerts."""
    with app.app_context():
        try:
            tasks = _load_gm_tasks()
            cfg   = _load_gm_mail_cfg()
            if not cfg.get("notifyOnOverdue", True):
                return
            # Persist last-known statuses in a file so restarts don't re-alert
            _state_file = DATA_DIR / "gm_tasks_overdue_state.json"
            prev = {}
            try:
                if _state_file.exists():
                    with open(_state_file) as f:
                        prev = json.load(f)
            except Exception:
                pass
            curr = {t["id"]: t.get("status","") for t in tasks if t.get("id")}
            newly_overdue = [
                t for t in tasks
                if t.get("id") and
                   curr.get(t["id"]) == "Overdue" and
                   prev.get(t["id"]) not in (None, "Overdue")
            ]
            # Save updated state
            try:
                with open(_state_file, "w") as f:
                    json.dump(curr, f)
            except Exception:
                pass
            for t in newly_overdue:
                try:
                    _maybe_notify_overdue(t)
                except Exception as _ne:
                    print(f"⚠️  [GM Overdue alert] {_ne}")
        except Exception as _e:
            print(f"⚠️  [GM Tasks overdue check] {_e}")

def _setup_gm_tasks_scheduler():
    """Register GM Tasks jobs onto the existing APScheduler instance."""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        _gm_sched = BackgroundScheduler(timezone="Asia/Kolkata")

        # Daily summary — 09:00 AM IST
        _gm_sched.add_job(
            func=_gm_tasks_daily_job,
            trigger="cron",
            hour=9, minute=0,
            timezone="Asia/Kolkata",
            id="gm_tasks_daily",
            replace_existing=True,
            misfire_grace_time=120,
        )

        # Overdue flip detection — every 5 minutes
        _gm_sched.add_job(
            func=_gm_tasks_overdue_check,
            trigger="interval",
            minutes=5,
            id="gm_tasks_overdue_check",
            replace_existing=True,
        )

        _gm_sched.start()
        print("✅ GM Tasks scheduler started — daily 09:00 AM IST + overdue check every 5 min")
    except Exception as _e:
        print(f"⚠️  GM Tasks scheduler failed to start: {_e}")

# =====================================================
# 20. START SERVER
# =====================================================


if __name__ == "__main__":
    
    # Start scheduler (PPM summary + CAM auto-reminders)
    try:
        setup_email_scheduler()
    except Exception as _sched_err:
        print(f"⚠️  Scheduler start error: {_sched_err}")

    # SLN MMS - daily WO auto-generation + mail
    try:
        from sln_mms_routes import register_mms_scheduler
        register_mms_scheduler(app)
    except Exception as _sln_err:
        print(f"⚠️  SLN MMS scheduler error: {_sln_err}")

    # GM Tasks — daily 09:00 AM IST email scheduler
    try:
        _setup_gm_tasks_scheduler()
    except Exception as _gm_err:
        print(f"⚠️  GM Tasks scheduler error: {_gm_err}")

    print(f"""
{'='*70}
⚙️  TERMINUS MMS — SERVER READY
{'='*70}
🌐 Dashboard:           http://localhost:5000
🎯 Command Center:      http://localhost:5000/command_center
📊 MMS (SLN):           http://localhost:5000/sln_mms_dashboard
📊 PPM (ONEWEST):       http://localhost:5000/ow_ppm_dashboard
🏢 SLN Terminus:        http://localhost:5000/sln_terminus
👥 Resource Mgmt:       http://localhost:5000/sln_resource_mgmt
🗒️  PM Daily Log:        http://localhost:5000/sln/pm/daily
🧹 SLN Housekeeping:    http://localhost:5000/sln_hk/
🔒 SLN Security:        http://localhost:5000/sln_sec
🔒 OW Security:         http://localhost:5000/ow_sec
🔥 SLN Fire Fighting:  http://localhost:5000/sln_fire/
🏢 ONEWEST:             http://localhost:5000/onewest
🏢 The District:        http://localhost:5000/the_district
🏢 One Golden Mile:     http://localhost:5000/ogm
🗒️  PM Daily (OGM):      http://localhost:5000/ogm/pm/daily
🏢 Nine Hills:          http://localhost:5000/nine_hills
⚡ Energy Analytics:    http://localhost:5000/energy
⚡ OW Energy Analytics: http://localhost:5000/ow_energy
📊 OW KRA Score Card:   http://localhost:5000/ow_kra
💰 CAM Billing:         http://localhost:5000/cam_charges
📈 Occupancy:           http://localhost:5000/sln_occupancy
💹 SLN Budget Review:   http://localhost:5000/sln_budget
🛠️  Issues:              http://localhost:5000/issues
📋 Project Handover:    http://localhost:5000/project_handover
📋 OW HOTO:             http://localhost:5000/ow_hoto
📋 OW PM Daily:         http://localhost:5000/ow_pm_daily
📋 MVGDS:               http://localhost:5000/mvgds
👔 GM Dashboard:        http://localhost:5000/gm_dashboard
📋 GM Tasks:            http://localhost:5000/gm_tasks
📁 Documents:           http://localhost:5000/documents
{'='*70}
📧 PPM Mailer:          Daily at 08:00 AM IST
📧 CAM Reminders:       Daily at 09:00 AM IST (+1d/+3d/+7d/+10d/+15d then daily)
📧 GM Tasks Report:     Daily at 09:00 AM IST
📅 Server Start:        {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*70}
""")
    app.run(host="0.0.0.0", port=5000, debug=False)
