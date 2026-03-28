"""
ow_energy.py  —  ONEWEST Energy Analytics Blueprint
Prefix: ow_  |  File: static/data/ow_energy_analysys.xlsx
Sheets: ele_con | dg_con | bench_mark | water _con | tank_con

FIX: Uses server's login_required + require_property decorators (session key = "user")
     require_property auto-sets active_property for direct URL / ngrok / mobile access
"""

from flask import (Blueprint, render_template, jsonify, request,
                   send_file, session, redirect, url_for, make_response)
from pathlib import Path
from datetime import datetime, timedelta
import traceback

ow_energy_bp = Blueprint("ow_energy", __name__)

BASE_DIR = Path(__file__).parent.resolve()
OW_XLSX  = BASE_DIR / "static" / "data" / "ow_energy_analysys.xlsx"
DATA_DIR = BASE_DIR / "static" / "data"


# ── Parse Excel ───────────────────────────────────────────────────────────────
def _ow_parse_excel():
    """Parse ow_energy_analysys.xlsx → structured dict for JSON API."""
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl not installed"
    if not OW_XLSX.exists():
        return None, f"File not found: {OW_XLSX}"
    try:
        wb = openpyxl.load_workbook(str(OW_XLSX), data_only=True)
        result = {}

        # ── Sheet 1: ele_con  (S.no | Month | HT Bill Units KVAH) ──────────
        ele = []
        for row in list(wb["ele_con"].iter_rows(values_only=True))[1:]:
            if not isinstance(row[0], (int, float)):
                continue
            m = row[1].strftime("%b %Y") if isinstance(row[1], datetime) else str(row[1] or "—")
            ele.append({"n": int(row[0]), "m": m, "u": float(row[2]) if row[2] else None})
        result["ele"] = ele

        # ── Sheet 2: dg_con  (Sno | Month | DG Run Hrs | kWh | Diesel L) ──
        dg = []
        rows = list(wb["dg_con"].iter_rows(values_only=True))
        start = next((i + 1 for i, r in enumerate(rows) if r[0] == "Sno"), 2)
        for row in rows[start:]:
            if not isinstance(row[0], (int, float)):
                continue
            m = row[1].strftime("%b %Y") if isinstance(row[1], datetime) else str(row[1] or "—")
            rh = row[2]
            if isinstance(rh, timedelta):
                ts = int(rh.total_seconds())
                hd = round(ts / 3600, 2)
                h_str = f"{ts // 3600:02d}:{(ts % 3600) // 60:02d}"
            elif hasattr(rh, "hour"):
                hd = round(rh.hour + rh.minute / 60, 2)
                h_str = f"{rh.hour:02d}:{rh.minute:02d}"
            else:
                hd = float(rh or 0)
                hh = int(hd)
                h_str = f"{hh:02d}:{int((hd - hh) * 60):02d}"
            dg.append({
                "n": int(row[0]), "m": m,
                "h": h_str, "hd": hd,
                "k": float(row[3] or 0),
                "d": float(row[4] or 0),
            })
        result["dg"] = dg

        # ── Sheet 3: bench_mark ──────────────────────────────────────────────
        bm = {}
        for row in wb["bench_mark"].iter_rows(values_only=True):
            if row[0] and row[1] is not None:
                bm[str(row[0]).strip()] = row[1]
        result["benchmark"] = bm

        # ── Sheet 4: water _con  (S.No | Month) ─────────────────────────────
        water = []
        ws_name = "water _con" if "water _con" in wb.sheetnames else "water_con"
        for row in list(wb[ws_name].iter_rows(values_only=True))[1:]:
            if not isinstance(row[0], (int, float)):
                continue
            m = row[1].strftime("%b %Y") if isinstance(row[1], datetime) else str(row[1] or "—")
            v = row[2] if len(row) > 2 else None
            water.append({"n": int(row[0]), "m": m, "v": float(v) if v else None})
        result["water"] = water

        # ── Sheet 5: tank_con  (S.No | Month) ───────────────────────────────
        tanker = []
        for row in list(wb["tank_con"].iter_rows(values_only=True))[1:]:
            if not isinstance(row[0], (int, float)):
                continue
            m = row[1].strftime("%b %Y") if isinstance(row[1], datetime) else str(row[1] or "—")
            tanker.append({
                "n":      int(row[0]),
                "m":      m,
                "trips":  int(row[2])   if len(row) > 2 and row[2] else None,
                "vol":    float(row[3]) if len(row) > 3 and row[3] else None,
                "rate":   float(row[4]) if len(row) > 4 and row[4] else 350,
                "vendor": str(row[5])   if len(row) > 5 and row[5] else "—",
            })
        result["tanker"] = tanker

        return result, None

    except Exception as e:
        return None, f"{e}\n{traceback.format_exc()}"


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES  — decorators injected at registration time via ow_energy_register()
#            so we use the server's real login_required + require_property.
# ══════════════════════════════════════════════════════════════════════════════

def _make_routes(login_required, require_property):
    """
    Build routes using the server's own decorators.
    Called once from ow_energy_register(app).
    """

    @ow_energy_bp.route("/ow_energy")
    @login_required
    @require_property("ONEWEST")
    def ow_energy_page():
        """ONEWEST Energy Analytics — main page."""
        session["active_property"] = "ONEWEST"
        print(f"\n⚡ OW Energy — User: {session.get('user')} | Role: {session.get('role')}")
        return render_template("ow_energy.html")

    @ow_energy_bp.route("/api/ow_energy/data")
    @login_required
    def ow_energy_data():
        """
        Live JSON — called by ow_energy.html on every page load.
        All clients (desktop / mobile / end-user) always get the current file.
        No @require_property so the API works for all property access levels.
        """
        data, err = _ow_parse_excel()
        if err:
            return jsonify({
                "error": err,
                "ele": [], "dg": [], "benchmark": {},
                "water": [], "tanker": []
            }), 200
        return jsonify(data)

    @ow_energy_bp.route("/api/ow_energy/upload", methods=["POST"])
    @login_required
    def ow_energy_upload():
        """
        Replace ow_energy_analysys.xlsx.
        Next /api/ow_energy/data call returns updated values — globally.
        """
        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"success": False, "error": "No file provided"}), 400
        if not f.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({"success": False, "error": "Only .xlsx/.xls accepted"}), 400
        try:
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            f.save(str(OW_XLSX))
            kb = OW_XLSX.stat().st_size // 1024
            return jsonify({
                "success": True,
                "message": f"Updated ({kb} KB). All clients reflect new data on next load."
            })
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    @ow_energy_bp.route("/api/ow_energy/file")
    @login_required
    def ow_energy_file():
        """Download the raw Excel file."""
        if not OW_XLSX.exists():
            return jsonify({"error": "File not found"}), 404
        resp = make_response(send_file(
            str(OW_XLSX),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="ow_energy_analysys.xlsx"
        ))
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp


def ow_energy_register(app, login_required=None, require_property=None):
    """
    Register blueprint onto app.
    Call from server.py AFTER login_required and require_property are defined:

        from ow_energy import ow_energy_register
        ow_energy_register(app, login_required, require_property)

    If decorators are not passed (legacy call), falls back to a safe minimal
    implementation using the server's session key ("user").
    """
    if login_required is None or require_property is None:
        # Fallback: minimal guard — session key "user" matches server.py
        from functools import wraps
        from flask import session as _sess, redirect as _redir, url_for as _urf, request as _req, abort as _abort

        def login_required(f):
            @wraps(f)
            def wrapper(*args, **kwargs):
                if "user" not in _sess:
                    return _redir(_urf("login") + "?next=" + _req.path)
                return f(*args, **kwargs)
            return wrapper

        def require_property(prop):
            def decorator(f):
                @wraps(f)
                def wrapper(*args, **kwargs):
                    if "user" not in _sess:
                        return _redir(_urf("login") + "?next=" + _req.path)
                    bypass = {"admin", "management", "general manager", "property manager"}
                    if (_sess.get("role") or "").lower() in bypass:
                        return f(*args, **kwargs)
                    # Auto-set active_property — fixes direct URL / ngrok / mobile
                    if prop in _sess.get("properties", []):
                        _sess["active_property"] = prop
                        return f(*args, **kwargs)
                    _abort(403)
                return wrapper
            return decorator

        print("⚠️  ow_energy: decorators not passed — using built-in fallback guards")
    else:
        print("✅ ow_energy: using server's login_required + require_property")

    _make_routes(login_required, require_property)
    app.register_blueprint(ow_energy_bp)
    print("✅ Registered: ow_energy_bp  →  /ow_energy")